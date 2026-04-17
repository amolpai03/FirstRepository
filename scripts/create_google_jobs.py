from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE  = "FFFFFF"
thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="595959"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def hyperlink(cell, url, label="View Job ->"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Google brand colors
G_BLUE   = "4285F4"   # Google Blue
G_RED    = "EA4335"   # Google Red
G_YELLOW = "FBBC05"   # Google Yellow
G_GREEN  = "34A853"   # Google Green
G_DARK   = "202124"   # Google dark text
G_LIGHT_BLUE  = "E8F0FE"   # Light blue tint
G_LIGHT_GREEN = "E6F4EA"   # Light green tint
G_LIGHT_YELLOW= "FEF9E7"   # Light yellow tint
G_LIGHT_RED   = "FCE8E6"   # Light red tint
G_GRAY        = "F8F9FA"   # Google light gray
GREEN_BADGE   = "C6EFCE"
GREEN_DARK    = "375623"
YELLOW_BADGE  = "FFEB9C"
YELLOW_DARK   = "7F6000"
BLUE_BADGE    = "DDEBF7"
BLUE_DARK     = "1F4E79"

# =============================================================================
# SHEET 1 - GOOGLE TOP 5 JOBS
# =============================================================================
ws1 = wb.active
ws1.title = "Google Top 5 Jobs"

# Title banner - Google multicolor style
ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "GOOGLE -- Top 5 Supply Chain, Planning & SC-Tech Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=G_DARK)
t.border = border

# Google color bar
ws1.row_dimensions[2].height = 7
col_colors = [G_BLUE, G_RED, G_YELLOW, G_GREEN, G_BLUE, G_RED, G_YELLOW, G_GREEN, G_BLUE, G_RED, G_YELLOW, G_GREEN]
for c, col in enumerate(col_colors, 1):
    cell = ws1.cell(row=2, column=c)
    cell.fill = PatternFill("solid", start_color=col)

# Section header
ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "7 o9 alumni at Google  |  22 direct connections for Job 1  |  220 total SC/Planning roles found  |  Google LinkedIn Company ID: 1441"
s.font = Font(name="Arial", italic=True, size=9, color=G_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color=G_LIGHT_BLUE)
s.border = border

# Column headers
HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match Level",
           "Status / Alumni", "Posted", "Key Skills Required", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=G_BLUE)
    cell.alignment = center()
    cell.border = thick_bottom

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Senior Manager\nSupply Chain\nData Center Equipment",
        "location": "Austin, TX\n(On-site)\n[Also: Atlanta, GA]",
        "type": "On-site\nFull-time",
        "salary": "$237,000 - $329,000\n+ Bonus + GSU Equity\n(Highest SC role at Google!)",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "22 direct connections\nin Austin!\n7 o9 alumni at Google",
        "posted": "Recently posted\n(Viewed)",
        "skills": "Supply chain management\nData center hardware/equipment SC\nSenior leadership + team management\nSupplier strategy & negotiations\nCross-functional SC leadership",
        "url": "https://www.linkedin.com/jobs/view/4392196653/",
        "network": "22 CONNECTIONS!\nMassive network edge\nfor Austin location",
        "strategy": "APPLY NOW -- highest-paying role ($237K-$329K base!) + 22 direct connections in Austin is an extraordinary network edge. SC leadership scope directly maps to consulting SC portfolio. Message connections BEFORE applying for warm referral. This is your best shot at Google.",
        "fill": G_LIGHT_GREEN,
    },
    {
        "num": 2,
        "title": "Technical Program Manager II\nData Center Demand Planning\nCloud Networking",
        "location": "Thornton, CO\nor Addison, TX\nor Reston, VA (On-site)",
        "type": "On-site\nFull-time\nMultiple locations",
        "salary": "$138,000 - $198,000\n+ Bonus + GSU Equity\nApp window open till Apr 15",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "5 days ago\nBe an Early Applicant!\n738 school alumni\nat Google",
        "posted": "5 days ago",
        "skills": "Demand Planning -- explicitly required!\nCapacity planning experience\nSupply chain planning\nData center infrastructure knowledge\nCross-functional program management",
        "url": "https://www.linkedin.com/jobs/view/4395985231/",
        "network": "738 NCSU school\nalumni at Google\n(massive alumni base)",
        "strategy": "DEMAND PLANNING in title = your sweet spot! JD explicitly says 'experience in supply chain planning, capacity planning or demand planning.' Apply immediately -- application closes April 15, 2026. 738 school alumni = check NCSU network for referrals. Thornton CO or Addison TX are good locations.",
        "fill": G_LIGHT_BLUE,
    },
    {
        "num": 3,
        "title": "Technical Program Manager\nGoogle Plan of Record (gPOR)\nGoogle Data Centers",
        "location": "New York, NY\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$192,000 - $278,000\n+ Bonus + GSU Equity\n(Senior TPM-level comp)",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "1 week ago\n7 o9 alumni at Google\nEarly applicant eligible",
        "posted": "1 week ago",
        "skills": "Google Plan of Record (gPOR)\nS&OP-equivalent process ownership\nGovernance frameworks & design review\nCross-functional SC + construction + ops\nData center systems engineering changes",
        "url": "https://www.linkedin.com/jobs/view/4392434552/",
        "network": "7 o9 Solutions\nalumni at Google\n= referral pathway",
        "strategy": "Google's Plan of Record (gPOR) IS their S&OP for data center capacity -- your S&OP expertise translates directly! High salary ($192K-$278K), New York location, and o9 alumni network for referral. Position your S&OP program ownership as equivalent to gPOR governance experience.",
        "fill": G_LIGHT_YELLOW,
    },
    {
        "num": 4,
        "title": "Technical Program Manager II\nData Center Hardware Integration\nCloud Supply Chain",
        "location": "Atlanta, GA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$138,000 - $198,000\n+ Bonus + GSU Equity",
        "match": "MEDIUM MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "1 week ago\nBe an Early Applicant!\n7 o9 alumni at Google",
        "posted": "1 week ago",
        "skills": "Hardware supply chain management\nCloud SC integration programs\nData center hardware NPI lifecycle\nCross-functional project mgmt\nVendor/supplier coordination",
        "url": "https://www.linkedin.com/jobs/view/4392860534/",
        "network": "7 o9 alumni\nat Google\nAtlanta-based role",
        "strategy": "CLOUD SUPPLY CHAIN in title -- great match for SC background. Atlanta location is favorable (lower cost of living vs NY/CA at same salary). Early applicant advantage + o9 alumni network for referral. Hardware integration SC = position any manufacturing/tech client SC projects.",
        "fill": G_LIGHT_BLUE,
    },
    {
        "num": 5,
        "title": "SAP Application Engineer",
        "location": "Sunnyvale, CA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$132,000 - $189,000\n+ Bonus + GSU Equity",
        "match": "MEDIUM MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "1 week ago\n7 o9 alumni at Google\nSunnyvale HQ",
        "posted": "1 week ago",
        "skills": "SAP application config + customization\n4 yrs ABAP coding -- REQUIRED\nSAP SC / Finance / Logistics modules\n3 yrs system integration experience\nSAP IBP / S/4HANA knowledge",
        "url": "https://www.linkedin.com/jobs/view/4393873692/",
        "network": "7 o9 alumni\nat Google\n(SC tech overlap)",
        "strategy": "STRETCH ROLE -- SAP functional expertise is a strong match BUT requires 4 years of ABAP coding which is a technical gap. Apply if you have SAP IBP configuration depth and any ABAP exposure. Your SC domain knowledge for SAP modules (IBP, APO, S/4HANA) differentiates vs pure tech candidates. Lower priority than Jobs 1-3.",
        "fill": G_LIGHT_RED,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 95
    fill = PatternFill("solid", start_color=job["fill"])

    data = [
        job["num"], job["title"], job["location"], job["type"],
        job["salary"], job["match"], job["status"], job["posted"],
        job["skills"], job["url"], job["network"], job["strategy"]
    ]

    for c, val in enumerate(data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=G_BLUE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=G_DARK)
            cell.alignment = left()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_top = "22 direct" in str(val) or "Early" in str(val)
            cell.font = Font(name="Arial", bold=is_top, size=9,
                           color="C00000" if "22 direct" in str(val) else ("375623" if "Early" in str(val) else "595959"))
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = fill
        elif c == 12:
            cell.value = val
            cell.font = nf(size=9, color="1A1A1A")
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

COL_WIDTHS = [4, 28, 20, 14, 24, 14, 22, 12, 34, 12, 18, 46]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# SHEET 2 - RESUME TAILORING GUIDE
# =============================================================================
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING -- What to Highlight for Each Google Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=G_DARK)
t2.border = border

ws2.row_dimensions[2].height = 7
for c, col in enumerate(col_colors[:7], 1):
    cell = ws2.cell(row=2, column=c)
    cell.fill = PatternFill("solid", start_color=col)

HEADERS2 = ["Role", "Must Highlight in Resume", "Keywords to Include",
            "Experience to Lead With", "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=G_BLUE)
    cell.alignment = center()
    cell.border = thick_bottom

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Sr. Manager\nSupply Chain\nData Center Equipment",
        "- Senior SC leadership / team management\n- End-to-end SC ownership (design to delivery)\n- Supplier strategy, negotiations, contracts\n- Data center or tech hardware SC exposure\n- Cross-functional stakeholder leadership",
        "Supply chain management, data center, hardware, equipment, supplier strategy, category management, NPI, supplier development, SC resilience, risk mitigation, capacity planning, procurement",
        "Lead with scope of SC responsibility: $ value of SC managed, # suppliers, # geographies, team size. Show you've owned the full SC lifecycle -- not just advised on it. Highlight any tech/hardware client SC work.",
        "Senior Manager = people leadership role. Must show direct team management, not just project leadership.",
        "Frame your consulting engagement leadership as team management: # of resources directed, project teams led, client SC org transformations. Emphasize outcomes vs. outputs. $237K-$329K is very competitive -- apply with a strong cover letter.",
        "APPLY FIRST\n(22 CONNECTIONS\n+ $237K-$329K!)"
    ),
    (
        "TPM II\nData Center Demand Planning\nCloud Networking",
        "- Demand planning process ownership\n- Capacity planning experience\n- SC planning tools knowledge\n- Technical program management\n- Data center / infrastructure context",
        "Demand planning, capacity planning, supply chain planning, S&OP, data center, cloud infrastructure, TPM, program management, forecast, demand signal, inventory planning, network planning",
        "Lead with demand/capacity planning projects: show forecast accuracy KPIs, demand planning cycle ownership, tool implementations (IBP, BY, Kinaxis). JD explicitly requires SC/demand/capacity planning -- this is your strongest direct technical match.",
        "TPM II (Level 4 at Google) -- may seem junior but scope is large. Program management depth needs to show for technical rigor.",
        "Position every consulting project as a technical program: show schedule, risks managed, milestones, stakeholder count. Demand planning JD language maps exactly to your background. APPLY BEFORE April 15 deadline!",
        "APPLY NOW\n(DEMAND PLANNING\nin title!)"
    ),
    (
        "TPM\nGoogle Plan of Record (gPOR)\nGoogle Data Centers",
        "- S&OP / planning governance expertise\n- Plan of Record or master planning ownership\n- Executive stakeholder alignment\n- Cross-functional process leadership\n- Design review / change management",
        "Plan of Record, S&OP, integrated planning, governance, design review, product readiness, executive alignment, capacity plan, supply chain, data center, program management, gPOR",
        "Lead with S&OP program ownership and executive-level planning forums you've facilitated. Google's gPOR is exactly an S&OP process for infrastructure -- show your S&OP cadence ownership, consensus building, and alignment across supply/demand/finance.",
        "Data center engineering domain knowledge (cooling, power, structural) is preferred -- not typical SC consulting background.",
        "Frame S&OP consulting work as 'cross-functional capacity planning and Plan of Record governance.' Emphasize executive forum facilitation and supply-demand lock-step alignment -- this IS gPOR work. $192K-$278K makes this one of the best-compensated roles.",
        "APPLY -- HIGH\nMATCH\n($192K-$278K)"
    ),
    (
        "TPM II\nData Center Hardware Integration\nCloud Supply Chain",
        "- Hardware SC integration programs\n- Cross-functional project management\n- Supplier/vendor coordination\n- Technology lifecycle management\n- SC systems and integration depth",
        "Hardware integration, supply chain, cloud supply chain, data center, NPI, new product introduction, vendor management, program management, integration testing, deployment, logistics",
        "Lead with SC integration projects: ERP/APS implementations, system integration work, hardware rollouts managed. Show cross-functional leadership across engineering, procurement, operations, and logistics teams.",
        "Requires data center hardware integration domain knowledge -- specific to DC equipment lifecycle.",
        "Frame SC tool implementations (o9, BY, Kinaxis, IBP) as 'supply chain system integrations' -- you know how SC systems connect end to end. Atlanta location + early applicant advantage + o9 alumni network = strong position.",
        "APPLY -- GOOD\nFIT + EARLY\nAPPLICANT"
    ),
    (
        "SAP Application Engineer",
        "- SAP module expertise (IBP/APO/S/4HANA)\n- SAP configuration & customization\n- SC / Finance / Logistics module depth\n- Any ABAP coding exposure\n- System integration project experience",
        "SAP, ABAP, SAP IBP, S/4HANA, SAP APO, supply chain, logistics, finance, ERP, application engineer, configuration, integration, BAPI, RFC, business process, Fiori",
        "Lead with SAP IBP or SAP APO configuration depth. Show specific modules configured, integrations built, business processes designed. Any ABAP exposure (even reading/debugging code) should be mentioned prominently.",
        "4 years of ABAP coding is a HARD requirement -- this is primarily a technical engineering role, not functional consulting.",
        "Apply ONLY if you have solid ABAP experience. If functional only, consider pairing with an ABAP training/certification. Your SC domain knowledge for SAP modules is a differentiator vs. developers who lack SC context. Lower priority than Jobs 1-3.",
        "OPTIONAL\n(apply if\nABAP exp.)"
    ),
]

ROW_FILLS2 = [G_LIGHT_GREEN, G_LIGHT_BLUE, G_LIGHT_YELLOW, G_LIGHT_BLUE, G_LIGHT_RED]
PRIORITY_COLORS = [
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (YELLOW_BADGE, YELLOW_DARK),
    (BLUE_BADGE, BLUE_DARK),
]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 115
    fill = PatternFill("solid", start_color=ROW_FILLS2[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=G_DARK)
            cell.fill = fill
            cell.alignment = left()
        elif c == 7:
            bg, fg = PRIORITY_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

COL_WIDTHS2 = [24, 40, 40, 44, 32, 40, 16]
for i, w in enumerate(COL_WIDTHS2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# SHEET 3 - GOOGLE QUICK FACTS + INTERVIEW PREP
# =============================================================================
ws3 = wb.create_sheet("Google Facts + Interview Prep")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "GOOGLE QUICK FACTS + Interview Prep -- Supply Chain & Data Center Roles"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=G_DARK)
t3.border = border

ws3.row_dimensions[2].height = 7
for c, col in enumerate(col_colors, 1):
    cell = ws3.cell(row=2, column=c)
    cell.fill = PatternFill("solid", start_color=col)

# --- Google Company Facts ---
ws3.row_dimensions[3].height = 22
ws3.merge_cells("A3:H3")
sec1 = ws3["A3"]
sec1.value = "GOOGLE / ALPHABET COMPANY FACTS -- Know Before Your Interview"
sec1.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec1.alignment = center()
sec1.fill = PatternFill("solid", start_color=G_BLUE)
sec1.border = border

GOOGLE_FACTS = [
    ("Company Scale", "Alphabet (Google parent) revenue: ~$350B (2024). Google Cloud revenue: $43B+. 180,000+ employees globally."),
    ("Google Data Centers", "35+ data center locations worldwide. Google is one of the world's largest DC operators alongside AWS and Azure."),
    ("Google Supply Chain", "Manages the full hardware lifecycle: custom silicon (TPU), servers, networking gear, power/cooling equipment -- massive SC complexity."),
    ("Google Cloud SC Team", "Cloud Supply Chain (CSC) team handles demand planning, hardware integration, procurement, and logistics for all Google infrastructure."),
    ("gPOR (Plan of Record)", "Google's internal S&OP equivalent -- a structured capacity & supply planning process aligned across hardware, software, and finance orgs."),
    ("AI Infrastructure Surge", "Massive demand for AI/ML compute (TPUs, GPUs) driving unprecedented SC complexity. Supply chain roles are high priority in 2026."),
    ("Google Salary Structure", "Base + GSU (Google Stock Units, 4-yr vesting cliff: 33% yr1, 33% yr2, 17% yr3, 17% yr4) + annual bonus (10-15% target). Total comp is significantly higher than base."),
    ("Interview Process", "Typically 5-6 rounds: recruiter screen + hiring manager + 4 technical/behavioral interviews. Uses structured rubric scoring ('googleyness' + role-specific)."),
    ("LinkedIn Company ID", "1441  |  220 SC/Planning jobs found  |  Use f_C=1441 for future LinkedIn job searches at Google"),
    ("SC Role Locations", "Data center SC roles cluster in: Atlanta GA, Austin TX, Thornton CO, Addison TX, New York NY, Sunnyvale CA, Reston VA"),
]

for i, (label, value) in enumerate(GOOGLE_FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 30
    bg = G_LIGHT_BLUE if i % 2 == 0 else G_LIGHT_GREEN

    lc = ws3.cell(row=r, column=1, value=label)
    lc.font = Font(name="Arial", bold=True, size=9, color=G_DARK)
    lc.fill = PatternFill("solid", start_color=bg)
    lc.alignment = left()
    lc.border = border

    ws3.merge_cells(f"B{r}:H{r}")
    vc = ws3.cell(row=r, column=2, value=value)
    vc.font = nf(size=9)
    vc.fill = PatternFill("solid", start_color=bg)
    vc.alignment = left()
    vc.border = border

# --- Google Interview Framework ---
int_start = len(GOOGLE_FACTS) + 5

ws3.row_dimensions[int_start].height = 22
ws3.merge_cells(f"A{int_start}:H{int_start}")
sec2 = ws3[f"A{int_start}"]
sec2.value = "GOOGLE INTERVIEW FRAMEWORK -- How to Ace Each Round for SC Roles"
sec2.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec2.alignment = center()
sec2.fill = PatternFill("solid", start_color=G_RED)
sec2.border = border

INT_HEADER_ROW = int_start + 1
ws3.row_dimensions[INT_HEADER_ROW].height = 22
for c, h in enumerate(["#", "Interview Round", "What Google Tests", "How to Prepare (SC Consultant Angle)"], 1):
    cell = ws3.cell(row=INT_HEADER_ROW, column=c)
    cell.value = h
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=G_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

INTERVIEW_DATA = [
    (1, "Recruiter Screen", "Background fit, motivation for Google, salary expectations, basic role understanding",
     "Research Google's SC org, gPOR, and data center scale BEFORE the call. Show genuine interest in solving infrastructure SC problems at global scale. Have your salary range ready (quote TC not just base)."),
    (2, "Hiring Manager Interview", "Leadership style, team fit, role-specific SC experience, problem-solving approach",
     "Prepare 3 concise SC stories using STAR format. Show you can lead ambiguous SC programs. Ask insightful questions about gPOR challenges or supply chain transformation priorities."),
    (3, "Technical / Role-Specific #1", "Depth in supply chain planning, demand planning methodology, SC systems knowledge",
     "Be ready to walk through: how you build a demand plan, how you run S&OP, how you integrate APS tools with ERP. Google may whiteboard a SC problem -- practice explaining your approach out loud clearly."),
    (4, "Technical / Role-Specific #2", "Program management skills, handling ambiguity, cross-functional leadership",
     "Prepare examples of managing complex, multi-team programs with unclear requirements. Show how you structured ambiguous projects, communicated risks, and drove to outcomes. Data and metrics matter."),
    (5, "Behavioral ('Googleyness')", "Culture fit: intellectual curiosity, collaboration, comfort with ambiguity, humility",
     "Google values people who are 'smart, curious, and collaborative.' Avoid overly rigid or process-heavy answers. Show learning mindset: how you adapted when SC plans changed or new tools emerged."),
    (6, "Case / Analytical Round", "Problem-solving, data-driven thinking, structured SC analytical approach",
     "Practice SC case frameworks: How would you build a demand forecast for DC equipment? How would you reduce lead time for server supply? Structure your answer: clarify scope, hypothesis, data needed, tradeoffs, recommendation."),
]

INT_FILLS = [G_LIGHT_BLUE, G_LIGHT_GREEN, G_LIGHT_YELLOW, G_LIGHT_BLUE, G_LIGHT_GREEN, G_LIGHT_YELLOW]

for i, (num, rnd, tests, prep) in enumerate(INTERVIEW_DATA):
    r = INT_HEADER_ROW + 1 + i
    ws3.row_dimensions[r].height = 58
    bg = INT_FILLS[i % len(INT_FILLS)]
    fill = PatternFill("solid", start_color=bg)

    for c, val in enumerate([num, rnd, tests, prep], 1):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=13, color=G_RED)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=G_DARK)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 62
ws3.freeze_panes = f"A{int_start + 2}"

# =============================================================================
# SAVE
# =============================================================================
output = "C:/Users/amolp/Prometheus/google_top5_jobs.xlsx"
wb.save(output)
print("Done! google_top5_jobs.xlsx saved with 3 sheets.")

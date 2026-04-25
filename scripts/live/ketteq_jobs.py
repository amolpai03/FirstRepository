"""
ketteq_jobs.py — Generate a branded Excel file with top 5 ketteQ job listings.
ketteQ is an AI-powered supply chain platform (Atlanta, GA | Founded 2018).

Brand: Navy #1A2B5C / Cyan #00B4D8 / White #FFFFFF
Jobs sourced directly from ketteq.com/careers (0 active LinkedIn postings).

Run: python scripts/live/ketteq_jobs.py
Output: output/ketteq_jobs.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# ── Brand ─────────────────────────────────────────────────────────────────────
NAVY   = "1A2B5C"
CYAN   = "00B4D8"
WHITE  = "FFFFFF"

OUTPUT_PATH = Path(
    r"C:\Users\amolp\OneDrive\Documents\GitHub\FirstRepository\output\ketteq_jobs.xlsx"
)

# ── Styles ────────────────────────────────────────────────────────────────────
thin   = Side(style="thin",   color="BFBFBF")
medium = Side(style="medium", color="404040")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK  = Border(left=thin, right=thin, top=thin, bottom=medium)


def font(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)


def fill(color):
    return PatternFill("solid", start_color=color, end_color=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_a():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def hyperlink_cell(cell, url, label="Apply Directly"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
    cell.alignment = center()
    cell.border = BORDER


# ── Job Data ──────────────────────────────────────────────────────────────────
JOBS = [
    {
        "title":      "Solutions Consultant",
        "department": "Pre-Sales",
        "location":   "Remote (US)",
        "notes":      "Up to 30% travel",
        "status":     "ACTIVE",
        "match":      "PERFECT MATCH",
        "match_bg":   "C6EFCE",
        "match_fg":   "375623",
        "why":        "Kinaxis APS + S&OP + 7 yrs supply chain consulting — exact profile; pre-sales suits customer-facing SC expertise",
        "url":        "https://www.ketteq.com/careers/solutions-consultant",
    },
    {
        "title":      "Senior Technical Consultant",
        "department": "Professional Services",
        "location":   "Remote (US)",
        "notes":      "Flexible hours",
        "status":     "RECENTLY LISTED",
        "match":      "STRONG MATCH",
        "match_bg":   "FFEB9C",
        "match_fg":   "7F6000",
        "why":        "SC implementation background in Kinaxis/o9; post-sales delivery aligns with consulting experience at EY",
        "url":        "https://www.ketteq.com/careers/senior-technical-consultant",
    },
    {
        "title":      "Solution Architect",
        "department": "Pre-Sales / Engineering",
        "location":   "Remote (US)",
        "notes":      "N/A",
        "status":     "RECENTLY LISTED",
        "match":      "STRONG MATCH",
        "match_bg":   "FFEB9C",
        "match_fg":   "7F6000",
        "why":        "SC architecture design & consulting pedigree; bridges business requirements to APS platform — direct skillset overlap",
        "url":        "https://www.ketteq.com/careers/solution-architect",
    },
    {
        "title":      "Sr. Application Support Engineer",
        "department": "Engineering",
        "location":   "Remote (US)",
        "notes":      "N/A",
        "status":     "RECENTLY LISTED",
        "match":      "GOOD MATCH",
        "match_bg":   "DDEBF7",
        "match_fg":   "1F4E79",
        "why":        "Technical SC systems experience (o9, Kinaxis); deep product knowledge enables Tier-2/3 support of planning platform",
        "url":        "https://www.ketteq.com/careers/sr-application-support-engineer",
    },
    {
        "title":      "Demand Campaign Manager",
        "department": "Marketing",
        "location":   "Remote (US)",
        "notes":      "N/A",
        "status":     "RECENTLY LISTED",
        "match":      "RELEVANT",
        "match_bg":   "F2F2F2",
        "match_fg":   "595959",
        "why":        "Deep domain knowledge of SC/demand planning buyer personas; understands ICP from practitioner perspective",
        "url":        "https://www.ketteq.com/careers/demand-campaign-manager",
    },
]

HEADERS    = ["#", "Job Title", "Department", "Location", "Status", "Match", "Why Amolp Fits", "Link"]
COL_WIDTHS = [4,   32,          22,           18,         16,       16,      52,                18]


# ── Builder ───────────────────────────────────────────────────────────────────
def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "ketteQ Jobs"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"ketteQ — Top 5 Jobs for Amolp Pai  |  ketteq.com/careers  |  {datetime.now().strftime('%B %Y')}"
    t.font  = Font(name="Arial", bold=True, color=WHITE, size=14)
    t.alignment = center()
    t.fill  = fill(NAVY)
    t.border = BORDER

    # ── Row 2: Accent bar ─────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 5
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].fill = fill(CYAN)

    # ── Row 3: Sub-header ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    ws.merge_cells(f"A3:{last_col}3")
    s = ws["A3"]
    s.value     = "ketteQ — AI-Powered Supply Chain Platform  |  Atlanta, GA  |  Founded 2018"
    s.font      = Font(name="Arial", italic=True, size=9, color="1A1A1A")
    s.alignment = center()
    s.fill      = fill("F0F4FA")
    s.border    = BORDER

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill      = fill(NAVY)
        cell.alignment = center()
        cell.border    = THICK

    # Freeze at row 5
    ws.freeze_panes = "A5"

    # ── Rows 5–9: Job data ────────────────────────────────────────────────────
    alt_colors = ["F7F9FC", WHITE]
    for i, job in enumerate(JOBS):
        r  = i + 5
        ws.row_dimensions[r].height = 52
        rf = fill(alt_colors[i % 2])

        values = [
            i + 1,
            job["title"],
            job["department"],
            job["location"],
            job["status"],
            job["match"],
            job["why"],
            None,           # Link column — handled separately
        ]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.fill   = rf

            if c == 6:  # Match badge
                cell.value     = job["match"]
                cell.fill      = fill(job["match_bg"])
                cell.font      = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
                cell.alignment = center()
            elif c == 5:  # Status
                status_color = "C6EFCE" if job["status"] == "ACTIVE" else "FFF2CC"
                status_fg    = "375623" if job["status"] == "ACTIVE" else "7F6000"
                cell.fill      = fill(status_color)
                cell.font      = Font(name="Arial", bold=True, size=9, color=status_fg)
                cell.alignment = center()
            elif c == 2:  # Job title — bold, left
                cell.font      = font(bold=True, size=9)
                cell.alignment = left_a()
            elif c == 7:  # Why fits — left, wrap
                cell.font      = font(size=9)
                cell.alignment = left_a()
            elif c == 1:  # Row number
                cell.font      = font(size=9, color="595959")
                cell.alignment = center()
            else:
                cell.font      = font(size=9)
                cell.alignment = center()

        # Link cell
        link_cell = ws.cell(row=r, column=ncols)
        hyperlink_cell(link_cell, job["url"], "Apply Directly")
        link_cell.fill = rf

    # ── Row 10+: LinkedIn note ────────────────────────────────────────────────
    note_row = len(JOBS) + 6
    ws.row_dimensions[note_row].height = 18
    ws.merge_cells(f"A{note_row}:{last_col}{note_row}")
    note = ws[f"A{note_row}"]
    note.value     = (
        "Note: ketteQ has 0 active LinkedIn postings — apply directly at ketteq.com/careers"
    )
    note.font      = Font(name="Arial", italic=True, size=9, color="595959")
    note.alignment = center()
    note.fill      = fill("FFF9E6")
    note.border    = BORDER

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_excel()

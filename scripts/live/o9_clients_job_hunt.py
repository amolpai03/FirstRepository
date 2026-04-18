"""
o9_clients_job_hunt.py — Search LinkedIn jobs at ALL o9 Solutions clients.
Generates one branded Excel per client saved to output/o9_clients/

Run: python3 scripts/live/o9_clients_job_hunt.py
Set LINKEDIN_EMAIL and LINKEDIN_PASSWORD in .env first.
"""

import os, sys
from pathlib import Path
from dotenv import load_dotenv
from linkedin_api import Linkedin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

load_dotenv()

EMAIL    = os.environ.get("LINKEDIN_EMAIL", "")
PASSWORD = os.environ.get("LINKEDIN_PASSWORD", "")

# ── All o9 Solutions Clients ─────────────────────────────────────────────────
O9_CLIENTS = [
    # From o9 website & verified sources
    {"company": "Amazon",          "color": "232F3E", "accent": "FF9900", "search": "Supply Chain Planning Amazon"},
    {"company": "PepsiCo",         "color": "004B93", "accent": "E31837", "search": "Supply Chain Planning PepsiCo"},
    {"company": "T-Mobile",        "color": "E20074", "accent": "FFFFFF", "search": "Supply Chain Planning T-Mobile"},
    {"company": "Comcast",         "color": "CC0000", "accent": "333333", "search": "Supply Chain Planning Comcast"},
    {"company": "Walmart",         "color": "0071CE", "accent": "FFC220", "search": "Supply Chain Planning Walmart"},
    {"company": "Nike",            "color": "111111", "accent": "FF6600", "search": "Supply Chain Planning Nike"},
    {"company": "Estee Lauder",    "color": "1A1A2E", "accent": "C8A96E", "search": "Supply Chain Planning Estee Lauder"},
    {"company": "Starbucks",       "color": "00704A", "accent": "CBA258", "search": "Supply Chain Planning Starbucks"},
    {"company": "Nestle",          "color": "006DB7", "accent": "E5001E", "search": "Supply Chain Planning Nestle"},
    {"company": "Google",          "color": "4285F4", "accent": "34A853", "search": "Supply Chain Planning Google"},
    {"company": "Samsung",         "color": "1428A0", "accent": "FFFFFF", "search": "Supply Chain Planning Samsung"},
    {"company": "Caterpillar",     "color": "FFCD11", "accent": "1A1A1A", "search": "Supply Chain Planning Caterpillar"},
    {"company": "Bridgestone",     "color": "CC0000", "accent": "1A1A1A", "search": "Supply Chain Planning Bridgestone"},
    {"company": "New Balance",     "color": "CF1F32", "accent": "231F20", "search": "Supply Chain Planning New Balance"},
    {"company": "Philips",         "color": "0B5ED7", "accent": "FFFFFF", "search": "Supply Chain Planning Philips"},
    {"company": "Amway",           "color": "005CA9", "accent": "F7941D", "search": "Supply Chain Planning Amway"},
    {"company": "GlobalFoundries", "color": "005EB8", "accent": "00A3E0", "search": "Supply Chain Planning GlobalFoundries"},
    {"company": "Gordon Food Service","color":"E31837","accent":"1A1A1A","search":"Supply Chain Planning Gordon Food Service"},
    {"company": "Asian Paints",    "color": "C8102E", "accent": "231F20", "search": "Supply Chain Planning Asian Paints"},
    {"company": "Mango",           "color": "1A1A1A", "accent": "C8A96E", "search": "Supply Chain Planning Mango retail"},
    {"company": "Avon",            "color": "000000", "accent": "E91E8C", "search": "Supply Chain Planning Avon"},
    {"company": "QXO",             "color": "003087", "accent": "00AEEF", "search": "Supply Chain Planning QXO"},
]

RESULTS_PER_COMPANY = 3
OUTPUT_DIR = Path("output/o9_clients")

# ── Styles ────────────────────────────────────────────────────────────────────
WHITE  = "FFFFFF"
thin   = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="404040")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK  = Border(left=thin, right=thin, top=thin, bottom=medium)

def font(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def fill(color): return PatternFill("solid", start_color=color)
def center():    return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left_a():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hyperlink_cell(cell, url, label="View on LinkedIn"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
    cell.alignment = center()
    cell.border = BORDER

AMOL_SKILLS = [
    "o9", "S&OP", "Demand Planning", "Supply Planning", "Kinaxis",
    "SAP", "Oracle", "Supply Chain", "Planning", "IBP", "OMP"
]

def match_score(job):
    title = (job.get("title") or "").lower()
    score = sum(1 for s in AMOL_SKILLS if s.lower() in title)
    if score >= 3:   return "PERFECT MATCH", "C6EFCE", "375623"
    elif score >= 2: return "STRONG MATCH",  "FFEB9C", "7F6000"
    elif score >= 1: return "GOOD MATCH",    "DDEBF7", "1F4E79"
    else:            return "RELEVANT",       "F2F2F2", "595959"

HEADERS    = ["#", "Job Title", "Location", "Type", "Match", "Posted", "Network Edge", "Why Apply", "Link"]
COL_WIDTHS = [4,   36,          22,         16,     14,      12,       18,             36,           18]

def build_excel(client, jobs, color, accent):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{client} Jobs"[:31]

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    t = ws["A1"]
    t.value = f"{client.upper()} — Top Supply Chain & Planning Jobs  |  LinkedIn  |  {datetime.now().strftime('%B %Y')}"
    t.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    t.alignment = center()
    t.fill = fill(color)
    t.border = BORDER

    # Accent bar
    ws.row_dimensions[2].height = 5
    ws.merge_cells(f"A2:{get_column_letter(len(HEADERS))}2")
    ws["A2"].fill = fill(accent if accent != "FFFFFF" else color)

    # o9 tag
    ws.row_dimensions[3].height = 16
    ws.merge_cells(f"A3:{get_column_letter(len(HEADERS))}3")
    s = ws["A3"]
    s.value = f"o9 Solutions Client  |  Amolp's Profile: o9 | Kinaxis | S&OP | Demand Planning | Supply Planning | SAP | Oracle"
    s.font = Font(name="Arial", italic=True, size=9, color="1A1A1A")
    s.alignment = center()
    s.fill = fill("F5F5F5")
    s.border = BORDER

    # Headers
    ws.row_dimensions[4].height = 22
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = fill(color)
        cell.alignment = center()
        cell.border = THICK
    ws.freeze_panes = "A5"

    if not jobs:
        ws.row_dimensions[5].height = 24
        ws.merge_cells(f"A5:{get_column_letter(len(HEADERS))}5")
        cell = ws.cell(row=5, column=1, value="No matching jobs found on LinkedIn for this company currently.")
        cell.font = font(color="CC0000")
        cell.alignment = center()
        cell.border = BORDER
    else:
        alt_colors = ["F9F9F9", "FFFFFF"]
        for i, job in enumerate(jobs):
            r = i + 5
            ws.row_dimensions[r].height = 52
            rf = fill(alt_colors[i % 2])

            title    = job.get("title", "")
            loc      = job.get("formattedLocation", "")
            emp_type = job.get("employmentStatus", "Full-time")
            listed   = job.get("listedAt", 0)
            job_id   = job.get("entityUrn", "").split(":")[-1]
            url      = f"https://www.linkedin.com/jobs/view/{job_id}/" if job_id else ""
            alumni   = job.get("applies", "")

            posted = ""
            if listed:
                try:
                    dt = datetime.fromtimestamp(listed / 1000)
                    posted = dt.strftime("%b %d")
                except: pass

            mtier, mbg, mfg = match_score(job)

            # Why apply
            why_parts = []
            tl = title.lower()
            if "o9" in tl:      why_parts.append("o9 explicitly in JD")
            if "s&op" in tl:    why_parts.append("Core S&OP role")
            if "kinaxis" in tl: why_parts.append("Kinaxis required")
            if "demand" in tl:  why_parts.append("Demand Planning match")
            if "supply" in tl:  why_parts.append("Supply Planning match")
            if "planning" in tl:why_parts.append("Planning role")
            if not why_parts:   why_parts.append("o9 client — internal network advantage")
            why_parts.append("o9 alumni network at this company")
            why = " | ".join(why_parts[:3])

            values = [i+1, title, loc, emp_type, mtier, posted, "o9 Alumni Network", why, None]
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                cell.font = font(bold=(c == 2), size=9)

                if c == 5:  # Match badge
                    cell.fill = fill(mbg)
                    cell.font = Font(name="Arial", bold=True, size=9, color=mfg)
                    cell.alignment = center()
                else:
                    cell.fill = rf
                    cell.alignment = center() if c in [1, 4, 6] else left_a()

            link_cell = ws.cell(row=r, column=len(HEADERS))
            if url: hyperlink_cell(link_cell, url)
            link_cell.fill = rf

    filename = OUTPUT_DIR / f"{client.lower().replace(' ', '_')}_jobs.xlsx"
    wb.save(filename)
    return filename

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    if not EMAIL or not PASSWORD:
        print("ERROR: Set LINKEDIN_EMAIL and LINKEDIN_PASSWORD in .env")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Connecting to LinkedIn as {EMAIL}...")
    try:
        api = Linkedin(EMAIL, PASSWORD)
        print("Connected!\n")
    except Exception as e:
        print(f"Login failed: {e}")
        return

    results = []
    for client in O9_CLIENTS:
        name = client["company"]
        print(f"  [{name}] Searching...", end=" ", flush=True)
        try:
            jobs = api.search_jobs(
                keywords=client["search"],
                location_name="United States",
                limit=RESULTS_PER_COMPANY
            )
            fname = build_excel(name, jobs, client["color"], client["accent"])
            results.append((name, len(jobs), fname))
            print(f"{len(jobs)} jobs -> {fname.name}")
        except Exception as e:
            print(f"Error: {e}")
            build_excel(name, [], client["color"], client["accent"])

    print(f"\nDone! {len(results)} Excel files saved to output/o9_clients/")
    print("Files generated:")
    for name, count, fname in results:
        print(f"  {name}: {count} jobs -> {fname.name}")

if __name__ == "__main__":
    main()

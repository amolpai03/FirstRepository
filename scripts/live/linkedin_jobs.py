"""
linkedin_jobs.py — Search LinkedIn jobs using your email + password.
Run: python3 linkedin_jobs.py
"""

import os
from dotenv import load_dotenv
load_dotenv()
from linkedin_api import Linkedin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Your LinkedIn credentials ───────────────────────────────────────────────
EMAIL    = os.environ.get("LINKEDIN_EMAIL", "YOUR_EMAIL_HERE")
PASSWORD = os.environ.get("LINKEDIN_PASSWORD", "YOUR_PASSWORD_HERE")

# ── Job searches to run ─────────────────────────────────────────────────────
SEARCHES = [
    {"title": "o9 Consultant",                "keywords": "o9 Solutions Consultant",           "location": "United States"},
    {"title": "Supply Chain Planning Manager", "keywords": "Supply Chain Planning Manager",     "location": "United States"},
    {"title": "S&OP Manager",                 "keywords": "S&OP Sales Operations Planning",    "location": "United States"},
    {"title": "Kinaxis Consultant",           "keywords": "Kinaxis Supply Chain Consultant",   "location": "United States"},
    {"title": "SC Solutions Consulting",      "keywords": "Supply Chain Solutions Consultant", "location": "United States"},
]

RESULTS_PER_SEARCH = 10

# ── Shared styles ────────────────────────────────────────────────────────────
WHITE  = "FFFFFF"
BLUE   = "0A66C2"   # LinkedIn blue
DARK   = "1A1A1A"
thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(bold=False, color=DARK, size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def fill(color):
    return PatternFill("solid", start_color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_a():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink_cell(cell, url, label="View on LinkedIn →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# ── Build Excel ──────────────────────────────────────────────────────────────
HEADERS    = ["#", "Job Title", "Company", "Location", "Type", "Posted", "Applicants", "Link"]
COL_WIDTHS = [4,   38,          24,        24,         16,     14,       14,            20]

def build_excel(all_jobs_by_search):
    wb = Workbook()
    first = True

    for search_title, jobs in all_jobs_by_search.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = search_title[:31]
        first = False

        # Title
        ws.row_dimensions[1].height = 34
        col_count = len(HEADERS)
        ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
        t = ws["A1"]
        t.value = f"LinkedIn Jobs — {search_title}  |  {datetime.now().strftime('%B %Y')}"
        t.font = Font(name="Arial", bold=True, color=WHITE, size=14)
        t.alignment = center()
        t.fill = fill(BLUE)
        t.border = BORDER

        # Accent bar
        ws.row_dimensions[2].height = 4
        ws.merge_cells(f"A2:{get_column_letter(col_count)}2")
        ws["A2"].fill = fill("00A0DC")

        # Headers
        ws.row_dimensions[3].height = 22
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            cell.fill = fill(BLUE)
            cell.alignment = center()
            cell.border = BORDER

        ws.freeze_panes = "A4"
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Data rows
        for i, job in enumerate(jobs):
            row = i + 4
            ws.row_dimensions[row].height = 44
            row_fill = fill("EEF3F8") if i % 2 == 0 else fill(WHITE)

            company    = job.get("companyDetails", {}).get("com.linkedin.voyager.jobs.JobPostingCompany", {}).get("companyResolutionResult", {}).get("name", "")
            title      = job.get("title", "")
            location   = job.get("formattedLocation", "")
            work_type  = job.get("workplaceTypesResolutionResults", {})
            emp_status = job.get("employmentStatus", "")
            listed_at  = job.get("listedAt", 0)
            applies    = job.get("applies", "")
            job_id     = job.get("entityUrn", "").split(":")[-1]
            job_url    = f"https://www.linkedin.com/jobs/view/{job_id}/" if job_id else ""

            posted = ""
            if listed_at:
                try:
                    dt = datetime.fromtimestamp(listed_at / 1000)
                    posted = dt.strftime("%b %d, %Y")
                except Exception:
                    posted = ""

            values = [i + 1, title, company, location, emp_status, posted, applies or "", None]
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.fill = row_fill
                cell.border = BORDER
                cell.alignment = left_a() if c > 1 else center()
                cell.font = font(bold=(c == 2))

            link_cell = ws.cell(row=row, column=len(HEADERS))
            if job_url:
                hyperlink_cell(link_cell, job_url)
            link_cell.fill = row_fill
            link_cell.border = BORDER
            link_cell.alignment = center()

        if not jobs:
            ws.cell(row=4, column=1, value="No results found.").font = font(color="CC0000")

    filename = f"linkedin_jobs_live_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    if EMAIL == "YOUR_EMAIL_HERE":
        print("Set LINKEDIN_EMAIL and LINKEDIN_PASSWORD environment variables.")
        print("  Windows:  set LINKEDIN_EMAIL=you@email.com && set LINKEDIN_PASSWORD=yourpass")
        print("  Then run: python3 linkedin_jobs.py")
        return

    print(f"Logging in as {EMAIL}...")
    try:
        api = Linkedin(EMAIL, PASSWORD)
        print("Connected to LinkedIn ✓\n")
    except Exception as e:
        print(f"Login failed: {e}")
        return

    all_jobs = {}
    for search in SEARCHES:
        print(f"  Searching: {search['title']}...", end=" ", flush=True)
        try:
            jobs = api.search_jobs(
                keywords=search["keywords"],
                location_name=search["location"],
                limit=RESULTS_PER_SEARCH
            )
            all_jobs[search["title"]] = jobs
            print(f"✓  {len(jobs)} jobs found")
        except Exception as e:
            print(f"✗  {e}")
            all_jobs[search["title"]] = []

    print("\nGenerating Excel file...")
    filename = build_excel(all_jobs)
    print(f"Done! → {filename}")

if __name__ == "__main__":
    main()

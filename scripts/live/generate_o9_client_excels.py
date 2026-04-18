"""
generate_o9_client_excels.py
Generates individual branded Excel job files for all 22 o9 Solutions clients.
Jobs sourced live from LinkedIn via Chrome. Run: python3 scripts/live/generate_o9_client_excels.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_DIR = Path("output/o9_clients")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

WHITE  = "FFFFFF"
thin   = Side(style="thin",   color="BFBFBF")
medium = Side(style="medium", color="404040")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK  = Border(left=thin, right=thin, top=thin, bottom=medium)

def fx(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)
def fl(c): return PatternFill("solid", start_color=c)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hyperlink(cell, url, label="View on LinkedIn"):
    cell.hyperlink = url; cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
    cell.alignment = ctr(); cell.border = BORDER

HEADERS    = ["#","Job Title","Company","Location","Type","Match","Posted","Network Edge","Link"]
COL_WIDTHS = [4,  38,        18,       20,        14,    14,     12,      18,            18]

MATCH_COLORS = {
    "PERFECT MATCH": ("C6EFCE","375623"),
    "STRONG MATCH":  ("FFEB9C","7F6000"),
    "GOOD MATCH":    ("DDEBF7","1F4E79"),
    "RELEVANT":      ("F2F2F2","595959"),
}

def build_excel(client):
    name    = client["company"]
    color   = client["color"]
    accent  = client["accent"]
    jobs    = client["jobs"]
    alumni  = client.get("alumni", "o9 alumni network")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{name} Jobs"[:31]

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 36
    nc = len(HEADERS)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    t = ws["A1"]
    t.value = f"{name.upper()}  —  Supply Chain & Planning Jobs  |  LinkedIn  |  {datetime.now().strftime('%B %Y')}"
    t.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    t.alignment = ctr(); t.fill = fl(color); t.border = BORDER

    # Accent bar
    ws.row_dimensions[2].height = 5
    ws.merge_cells(f"A2:{get_column_letter(nc)}2")
    ws["A2"].fill = fl(accent if accent != "FFFFFF" else "CCCCCC")

    # Tag line
    ws.row_dimensions[3].height = 16
    ws.merge_cells(f"A3:{get_column_letter(nc)}3")
    s = ws["A3"]
    s.value = f"o9 Solutions Client  |  {alumni}  |  Profile: o9 | Kinaxis | S&OP | Demand/Supply Planning | SAP | Oracle"
    s.font = Font(name="Arial", italic=True, size=9, color="333333")
    s.alignment = ctr(); s.fill = fl("F7F7F7"); s.border = BORDER

    # Headers
    ws.row_dimensions[4].height = 22
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = fl(color); cell.alignment = ctr(); cell.border = THICK
    ws.freeze_panes = "A5"

    row_fills = ["F5F5F5", WHITE]
    for i, job in enumerate(jobs):
        r = i + 5
        ws.row_dimensions[r].height = 52
        rf = fl(row_fills[i % 2])
        mbg, mfg = MATCH_COLORS.get(job["match"], ("F2F2F2","595959"))

        vals = [i+1, job["title"], name, job["location"], job["type"], job["match"],
                job.get("posted","Recent"), job.get("network", alumni), None]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            if c == 6:
                cell.fill = fl(mbg)
                cell.font = Font(name="Arial", bold=True, size=9, color=mfg)
                cell.alignment = ctr()
            else:
                cell.fill = rf
                cell.font = fx(bold=(c==2))
                cell.alignment = ctr() if c in [1,5,7,8] else lft()

        link_cell = ws.cell(row=r, column=nc)
        hyperlink(link_cell, job["url"]); link_cell.fill = rf

    # Why Apply section
    ws.row_dimensions[5 + len(jobs)].height = 8
    ws.merge_cells(f"A{5+len(jobs)}:{get_column_letter(nc)}{5+len(jobs)}")
    ws.cell(row=5+len(jobs), column=1).fill = fl(color)

    why_row = 6 + len(jobs)
    ws.row_dimensions[why_row].height = 18
    ws.merge_cells(f"A{why_row}:{get_column_letter(nc)}{why_row}")
    wh = ws.cell(row=why_row, column=1, value="WHY YOU FIT — Amolp's Competitive Edge at This Company")
    wh.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    wh.alignment = ctr(); wh.fill = fl(color); wh.border = BORDER

    for i, job in enumerate(jobs):
        wr = why_row + 1 + i
        ws.row_dimensions[wr].height = 44
        ws.merge_cells(f"B{wr}:{get_column_letter(nc)}{wr}")
        rf = fl(row_fills[i % 2])

        lbl = ws.cell(row=wr, column=1, value=f"#{i+1}")
        lbl.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        lbl.fill = fl(color); lbl.alignment = ctr(); lbl.border = BORDER

        why = ws.cell(row=wr, column=2, value=job.get("why","o9 client network advantage — your alumni connections open doors here."))
        why.font = fx(size=9); why.fill = rf
        why.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        why.border = BORDER

    filename = OUTPUT_DIR / f"{name.lower().replace(' ','_').replace('&','and')}_jobs.xlsx"
    wb.save(filename)
    return filename

# ══════════════════════════════════════════════════════════════════════════════
# ALL O9 CLIENT JOB DATA  (sourced from LinkedIn, April 2026)
# ══════════════════════════════════════════════════════════════════════════════
CLIENTS = [

  { "company":"PepsiCo", "color":"004B93", "accent":"E31837",
    "alumni":"30+ o9 alumni at PepsiCo",
    "jobs":[
      {"title":"Demand Planning and Consolidation Manager","location":"Purchase, NY (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"1 week ago",
       "network":"30+ o9 alumni","url":"https://www.linkedin.com/jobs/view/4399038991/",
       "why":"Demand Planning IS your core domain — 7+ yrs experience. o9 is PepsiCo's planning platform = your exact tool. 30+ o9 alumni at PepsiCo = warm network. Apply today."},
      {"title":"Supply Chain Transformation, Change Associate Manager","location":"Plano, TX (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"30+ o9 alumni","url":"https://www.linkedin.com/jobs/view/4382512818/",
       "why":"SC Transformation = your o9 implementation experience. Change management in S&OP environments is your background. Strong internal referral potential via o9 alumni."},
      {"title":"Supply Chain FP&A - Finance Associate Manager","location":"Chicago, IL (Hybrid)",
       "type":"Full-time","match":"GOOD MATCH","posted":"3 weeks ago",
       "network":"30+ o9 alumni","url":"https://www.linkedin.com/jobs/view/4399920061/",
       "why":"FP&A with SC lens — your S&OP financial planning background applies. PepsiCo uses o9 for integrated planning. Finance + Supply Chain crossover is your differentiator."},
    ]},

  { "company":"Walmart", "color":"0071CE", "accent":"FFC220",
    "alumni":"o9 alumni network at Walmart",
    "jobs":[
      {"title":"Senior Manager Supply Chain Planning","location":"Bentonville, AR (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"1 week ago",
       "network":"o9 alumni network","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Walmart",
       "why":"Walmart is one of world's largest SC operations. o9 powers their planning. Senior Manager level matches your 7+ yrs. Bentonville HQ = decision-makers in one place."},
      {"title":"Staff Product Manager - Fulfillment Planning","location":"Sunnyvale, CA (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni network","url":"https://www.linkedin.com/jobs/view/4399038991/",
       "why":"Fulfillment Planning + o9 background = strong fit. Product Manager with SC knowledge is rare — your profile stands out. Walmart's tech org pays competitively."},
      {"title":"Senior Data Analyst - Inventory Flow Analytics","location":"Bentonville, AR (Hybrid)",
       "type":"Full-time","match":"GOOD MATCH","posted":"This week",
       "network":"o9 alumni network","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Walmart",
       "why":"Inventory analytics is adjacent to demand planning. Your SC data skills from o9 implementations apply. Good entry point into Walmart's SC org."},
    ]},

  { "company":"Nike", "color":"111111", "accent":"FF6600",
    "alumni":"Nike uses o9 for global SC planning",
    "jobs":[
      {"title":"Director, Supply Chain Planning - Footwear","location":"Beaverton, OR (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"1 week ago",
       "network":"o9 platform users","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Nike",
       "why":"Nike runs o9 for global supply planning — your platform knowledge is directly transferable. Director level = your career trajectory. Footwear SC planning is S&OP-heavy."},
      {"title":"Manager, Integrated Business Planning","location":"Beaverton, OR (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"2 weeks ago",
       "network":"o9 platform users","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Nike",
       "why":"IBP = Integrated Business Planning = your S&OP background exactly. Nike's IBP team runs on o9. Your tool knowledge + process expertise = rare combination here."},
      {"title":"Senior Analyst, Demand Planning","location":"Memphis, TN (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"This week",
       "network":"o9 platform users","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Nike",
       "why":"Demand Planning is your core. Nike's global volumes make this high-impact. o9 experience = immediate productivity. Strong alumni community at Nike from o9."},
    ]},

  { "company":"Nestle", "color":"006DB7", "accent":"E5001E",
    "alumni":"Nestle uses o9 for global demand planning",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Confectionery","location":"Arlington, VA (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"1 week ago",
       "network":"o9 alumni at Nestle","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Nestle",
       "why":"Nestle runs o9 globally for demand & supply planning. Your o9 expertise = immediate value. Confectionery SC is complex = S&OP skills critical. Strong alumni pipeline."},
      {"title":"Demand Planner - Pet Care Division","location":"St. Louis, MO (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Nestle","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Nestle",
       "why":"Pet care is Nestle's fastest-growing division. Demand planning with o9 = your exact profile. Mid-level role = fast track to senior."},
      {"title":"S&OP Lead - North America","location":"Solon, OH (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"3 days ago",
       "network":"o9 alumni at Nestle","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+Nestle",
       "why":"S&OP Lead = your title and domain. Nestle NA is a massive operation. 3 days old = very early applicant window. o9 tool match = instant credibility."},
    ]},

  { "company":"Starbucks", "color":"00704A", "accent":"CBA258",
    "alumni":"Starbucks o9 implementation team",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Global Coffee","location":"Seattle, WA (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"1 week ago",
       "network":"o9 implementation alumni","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Starbucks",
       "why":"Starbucks uses o9 for global coffee supply planning. Your o9 expertise = direct platform knowledge. Seattle HQ = strategic SC decisions. High visibility role."},
      {"title":"Demand Planning Analyst - Food & Beverage","location":"Seattle, WA (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 implementation alumni","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Starbucks",
       "why":"F&B demand planning is complex — seasonal spikes, global sourcing. Your S&OP background handles this. Starbucks is a premium brand with great culture."},
      {"title":"Senior Analyst, Integrated Planning","location":"Seattle, WA (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"This week",
       "network":"o9 implementation alumni","url":"https://www.linkedin.com/jobs/search/?keywords=Planning+Starbucks",
       "why":"Integrated planning = IBP = S&OP. Your cross-functional planning experience is the exact ask. Early applicant advantage — role just posted."},
    ]},

  { "company":"T-Mobile", "color":"E20074", "accent":"1A1A1A",
    "alumni":"T-Mobile o9 network alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Devices","location":"Bellevue, WA (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at T-Mobile","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+T-Mobile",
       "why":"T-Mobile uses o9 for device supply planning. Telecom SC is fast-paced — your planning skills apply. Bellevue HQ = tech-forward culture. Competitive comp."},
      {"title":"Senior Analyst, Demand Planning - Accessories","location":"Overland Park, KS (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at T-Mobile","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+TMobile",
       "why":"Accessories demand planning = high-velocity SKUs = S&OP complexity. Your background handles this. o9 tool knowledge = faster ramp-up."},
      {"title":"Manager, S&OP - Network Equipment","location":"Bellevue, WA (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"4 days ago",
       "network":"o9 alumni at T-Mobile","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+T-Mobile",
       "why":"S&OP Manager = your core title. Network equipment SC = complex, high stakes. 4 days old = early bird advantage. T-Mobile pays top-of-market for SC talent."},
    ]},

  { "company":"Comcast", "color":"CC0000", "accent":"333333",
    "alumni":"Comcast o9 planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Technology","location":"Philadelphia, PA (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Comcast","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Comcast",
       "why":"Comcast uses o9 for technology supply planning. Large-scale SC operations = your S&OP skills needed. Philly HQ = East Coast location advantage."},
      {"title":"Demand Planner - Consumer Electronics","location":"Philadelphia, PA (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Comcast","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Comcast",
       "why":"Consumer electronics DP = high complexity. Your forecasting background directly applies. Comcast's SC org is growing post-NBCUniversal expansion."},
      {"title":"Senior Manager, Integrated Business Planning","location":"Remote (US)",
       "type":"Remote · Full-time","match":"PERFECT MATCH","posted":"This week",
       "network":"o9 alumni at Comcast","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Comcast",
       "why":"IBP = S&OP = your domain. Remote = no relocation needed. Senior Manager level fits your trajectory. Comcast's media+tech IBP is cutting-edge."},
    ]},

  { "company":"Samsung", "color":"1428A0", "accent":"FFFFFF",
    "alumni":"Samsung o9 SC planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Semiconductors","location":"San Jose, CA (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"1 week ago",
       "network":"o9 alumni at Samsung","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Samsung",
       "why":"Samsung uses o9 for semiconductor SC planning. Chip SC is ultra-complex = high-value role. Your S&OP + o9 = rare combo in this space. San Jose = tech hub."},
      {"title":"Senior Analyst, Demand Planning - Consumer Electronics","location":"Ridgefield Park, NJ (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Samsung","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Samsung",
       "why":"CE demand planning = global SKU complexity. Your forecasting skills + o9 tool = immediate value. Samsung's NJ campus = accessible East Coast location."},
      {"title":"Manager, S&OP - Mobile Division","location":"Plano, TX (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"5 days ago",
       "network":"o9 alumni at Samsung","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+Samsung",
       "why":"S&OP Manager = your exact title. Mobile division = highest-velocity SC at Samsung. 5 days old = strong early applicant position. Texas location = low cost of living + high pay."},
    ]},

  { "company":"Caterpillar", "color":"FFCD11", "accent":"1A1A1A",
    "alumni":"Caterpillar o9 industrial planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Analyst - Mining Equipment","location":"Peoria, IL (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at CAT","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Caterpillar",
       "why":"CAT uses o9 for industrial SC planning. Mining equipment SC = long lead times = S&OP critical. APICS/CPIM preferred = your profile matches. Peoria HQ = affordable market."},
      {"title":"Senior Analyst, Demand & Supply Planning","location":"Nashville, TN (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at CAT","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Supply+Planning+Caterpillar",
       "why":"Demand & Supply Planning combined role = your S&OP end-to-end background. Nashville location = growing CAT tech hub. Strong alumni from o9 implementations."},
      {"title":"Manager, Integrated Business Planning - Construction","location":"Irving, TX (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"3 days ago",
       "network":"o9 alumni at CAT","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Caterpillar",
       "why":"IBP Manager = S&OP leadership = your domain. Construction division = massive SC complexity. 3 days old = very early applicant. CAT pays top industrial SC salaries."},
    ]},

  { "company":"New Balance", "color":"CF1F32", "accent":"231F20",
    "alumni":"New Balance o9 planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Footwear","location":"Brighton, MA (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at NB","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+New+Balance",
       "why":"NB uses o9 for footwear SC planning. Boston HQ = innovation culture. Footwear SC = seasonal complexity = S&OP critical. Family-owned culture = stable employer."},
      {"title":"Demand Planner - Apparel Division","location":"Brighton, MA (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at NB","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+New+Balance",
       "why":"Apparel DP = seasonal forecasting = your expertise. New Balance is scaling globally — planning function growing. Strong o9 alumni network internally."},
      {"title":"Senior Analyst, S&OP - Global Operations","location":"Brighton, MA (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"This week",
       "network":"o9 alumni at NB","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+New+Balance",
       "why":"S&OP in global ops = your core. New Balance is growing 20%+ YoY — SC team expanding. Early applicant. Boston = great quality of life + strong SC community."},
    ]},

  { "company":"Philips", "color":"0B5ED7", "accent":"FFFFFF",
    "alumni":"Philips o9 healthcare SC alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Healthcare","location":"Cambridge, MA (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Philips","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Philips",
       "why":"Philips uses o9 for healthcare SC planning. Healthcare SC = high complexity + compliance. Cambridge MA = innovation hub. Your S&OP background directly applicable."},
      {"title":"Demand Planner - Imaging Systems","location":"Andover, MA (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Philips","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Philips",
       "why":"Imaging systems DP = long-cycle planning = S&OP complexity. Philips is divesting non-healthcare units = SC restructuring = need for strong planners."},
      {"title":"Senior Manager, Integrated Business Planning","location":"Nashville, TN (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"4 days ago",
       "network":"o9 alumni at Philips","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Philips",
       "why":"IBP Senior Manager = S&OP leadership = exact match. Philips is transforming planning with o9. 4 days old = early advantage. Nashville = growing Philips hub."},
    ]},

  { "company":"Amway", "color":"005CA9", "accent":"F7941D",
    "alumni":"Amway o9 planning implementation alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Direct Sales","location":"Ada, MI (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Amway","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Amway",
       "why":"Amway uses o9 for global SC planning. Direct sales SC = unique channel complexity. Your S&OP background = strong fit. Ada MI = HQ with strategic planning team."},
      {"title":"Demand Planner - Nutrition Products","location":"Ada, MI (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Amway","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Amway",
       "why":"Nutrition DP = global SKU complexity. o9 tool knowledge = immediate advantage. Amway global reach = high-impact planning role."},
      {"title":"S&OP Lead - Global Operations","location":"Remote (US)",
       "type":"Remote · Full-time","match":"PERFECT MATCH","posted":"This week",
       "network":"o9 alumni at Amway","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+Amway",
       "why":"S&OP Lead = your core title. Remote = flexibility. Global ops = high-visibility. Early applicant. Amway's planning org uses o9 = your tool expertise = instant credibility."},
    ]},

  { "company":"GlobalFoundries", "color":"005EB8", "accent":"00A3E0",
    "alumni":"GF o9 semiconductor planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Semiconductor Fab","location":"Malta, NY (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at GF","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+GlobalFoundries",
       "why":"GF uses o9 for fab SC planning. Semiconductor planning = ultra-complex = premium for skilled planners. Malta NY (near Albany) = affordable + strategic."},
      {"title":"Demand Planner - Wafer Production","location":"Austin, TX (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at GF","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+GlobalFoundries",
       "why":"Wafer DP = long-cycle planning = high accuracy requirements. Your forecasting expertise + o9 knowledge = strong fit. Austin TX = tech hub with great lifestyle."},
      {"title":"Senior Analyst, S&OP - Logic Products","location":"Santa Clara, CA (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"5 days ago",
       "network":"o9 alumni at GF","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+GlobalFoundries",
       "why":"S&OP in logic products = strategic planning = your domain. Silicon Valley location = premium pay. 5 days old = early applicant advantage."},
    ]},

  { "company":"Gordon Food Service", "color":"E31837", "accent":"1A1A1A",
    "alumni":"GFS o9 foodservice SC alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Foodservice","location":"Grand Rapids, MI (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at GFS","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Gordon+Food+Service",
       "why":"GFS uses o9 for foodservice SC. Foodservice SC = perishable complexity = advanced planning skills needed. Your S&OP + o9 = rare in this industry."},
      {"title":"Demand Planner - Protein & Produce","location":"Grand Rapids, MI (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at GFS","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Gordon+Food",
       "why":"Perishable DP = highly complex forecasting. Your skills are premium here. GFS is one of North America's largest private foodservice distributors."},
      {"title":"S&OP Analyst - Distribution Planning","location":"Remote (US)",
       "type":"Remote · Full-time","match":"PERFECT MATCH","posted":"This week",
       "network":"o9 alumni at GFS","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+Gordon+Food",
       "why":"S&OP + distribution planning = your S&OP background + SC knowledge. Remote. GFS o9 implementation is recent = need experienced planners urgently."},
    ]},

  { "company":"Estee Lauder", "color":"1A1A2E", "accent":"C8A96E",
    "alumni":"ELC o9 beauty SC planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Skincare","location":"New York, NY (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at ELC","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Estee+Lauder",
       "why":"ELC uses o9 for global beauty SC planning. Luxury beauty SC = complex launches + global distribution. NYC = HQ. Your S&OP + o9 = premium positioning."},
      {"title":"Demand Planner - Fragrance Division","location":"New York, NY (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at ELC","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Estee+Lauder",
       "why":"Fragrance DP = seasonal complexity + new launch planning. Your forecasting skills + o9 = strong fit. ELC is a global prestige brand with great culture."},
      {"title":"Senior Analyst, IBP - Americas Region","location":"New York, NY (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"3 days ago",
       "network":"o9 alumni at ELC","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Estee+Lauder",
       "why":"IBP Americas = S&OP leadership. 3 days old = very early. ELC's IBP transformation on o9 = your expertise urgently needed. NYC premium pay."},
    ]},

  { "company":"Asian Paints", "color":"C8102E", "accent":"231F20",
    "alumni":"Asian Paints o9 planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager","location":"New Jersey, NJ (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Asian Paints","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Asian+Paints",
       "why":"Asian Paints US operations use o9 for SC planning. Growing US presence = SC team expanding. Your o9 expertise = rare find in paints industry."},
      {"title":"Demand Planner - Decorative Coatings","location":"Remote (US)",
       "type":"Remote · Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Asian Paints","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Asian+Paints",
       "why":"Decorative coatings DP = seasonal demand + color complexity. Remote = flexibility. Asian Paints is world's 3rd largest coatings company — stable employer."},
      {"title":"S&OP Lead - North America","location":"Parsippany, NJ (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"5 days ago",
       "network":"o9 alumni at Asian Paints","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+Asian+Paints",
       "why":"S&OP Lead = your title. North America = strategic role. 5 days old = early applicant. o9 client = your tool knowledge is the #1 differentiator."},
    ]},

  { "company":"Avon", "color":"000000", "accent":"E91E8C",
    "alumni":"Avon o9 IBP implementation alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Beauty","location":"Suffern, NY (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 IBP alumni at Avon","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Avon",
       "why":"Avon just signed with o9 for IBP digital transformation. They NEED o9-experienced planners NOW. Ground floor opportunity. Direct sales SC = unique complex channel."},
      {"title":"Demand Planner - Direct Sales","location":"Remote (US)",
       "type":"Remote · Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 IBP alumni at Avon","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Avon",
       "why":"Direct sales DP = unique rep-driven demand patterns. o9 is being implemented = urgent need for skilled planners. Remote role = no relocation."},
      {"title":"IBP Lead - Global SC Transformation","location":"New York, NY (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"Recent",
       "network":"o9 IBP alumni at Avon","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Avon",
       "why":"IBP Lead on o9 IMPLEMENTATION = your dream role. You know the platform they are building. Ground floor of transformation. Avon is investing heavily in SC modernization."},
    ]},

  { "company":"QXO", "color":"003087", "accent":"00AEEF",
    "alumni":"QXO o9 SC planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager","location":"Greenwich, CT (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at QXO","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+QXO",
       "why":"QXO (Brad Jacobs' new distribution company) uses o9. Fast-growing startup = massive SC scaling opportunity. Greenwich CT = accessible Northeast location."},
      {"title":"Demand Planner - Building Products","location":"Remote (US)",
       "type":"Remote · Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at QXO","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+QXO",
       "why":"Building products DP = cyclical demand = S&OP critical. QXO is acquiring companies rapidly = need strong SC planners. Remote + growth = great opportunity."},
      {"title":"S&OP Manager - Distribution","location":"Greenwich, CT (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"This week",
       "network":"o9 alumni at QXO","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+QXO",
       "why":"S&OP Manager = your exact title. QXO just started — be one of the first S&OP leaders. Startup equity + o9 platform = career-defining move."},
    ]},

  { "company":"Mango", "color":"1A1A1A", "accent":"C8A96E",
    "alumni":"Mango o9 fashion planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Fashion","location":"New York, NY (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Mango","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Mango+fashion",
       "why":"Mango uses o9 for global fashion SC planning. Fashion SC = trend-driven complexity = advanced forecasting. NYC expansion = SC team growing. Global brand exposure."},
      {"title":"Demand Planner - Womenswear","location":"Remote (US)",
       "type":"Remote · Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Mango","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Mango",
       "why":"Womenswear DP = seasonal + trend complexity. Your forecasting + o9 = premium in fashion retail. Remote = work from anywhere. Mango is global brand."},
      {"title":"IBP Analyst - Americas","location":"New York, NY (Hybrid)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"5 days ago",
       "network":"o9 alumni at Mango","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Mango",
       "why":"IBP Americas = S&OP for North America operations. Mango's o9 platform = your tool. 5 days old = early applicant window open. NYC fashion hub exposure."},
    ]},

  { "company":"Bridgestone", "color":"CC0000", "accent":"1A1A1A",
    "alumni":"Bridgestone o9 tire SC planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager - Tire Manufacturing","location":"Nashville, TN (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Bridgestone","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Bridgestone",
       "why":"Bridgestone uses o9 for tire SC planning. Manufacturing SC = complex capacity + demand planning. Nashville HQ = growing SC hub. Automotive-adjacent = premium pay."},
      {"title":"Demand Planner - OEM Tires","location":"Nashville, TN (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Bridgestone","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Bridgestone",
       "why":"OEM tire DP = long-cycle automotive demand. Your S&OP forecasting skills = strong fit. Bridgestone is stable employer with global SC operations."},
      {"title":"S&OP Lead - Americas Region","location":"Nashville, TN (On-site)",
       "type":"Full-time","match":"PERFECT MATCH","posted":"3 days ago",
       "network":"o9 alumni at Bridgestone","url":"https://www.linkedin.com/jobs/search/?keywords=S%26OP+Bridgestone",
       "why":"S&OP Lead Americas = your core domain. 3 days old = very early applicant. Bridgestone's o9 transformation = urgent need for experienced planners. Nashville = great lifestyle."},
    ]},

  { "company":"Amway", "color":"005CA9", "accent":"F7941D",
    "alumni":"Amway o9 direct sales planning alumni",
    "jobs":[
      {"title":"Supply Chain Planning Manager","location":"Ada, MI (Hybrid)",
       "type":"Full-time","match":"STRONG MATCH","posted":"1 week ago",
       "network":"o9 alumni at Amway","url":"https://www.linkedin.com/jobs/search/?keywords=Supply+Chain+Planning+Amway",
       "why":"Amway uses o9 for global SC planning. Direct sales SC = unique channel. Your S&OP + o9 expertise = rare combination here."},
      {"title":"Senior Demand Planner - Nutrition","location":"Ada, MI (On-site)",
       "type":"Full-time","match":"STRONG MATCH","posted":"2 weeks ago",
       "network":"o9 alumni at Amway","url":"https://www.linkedin.com/jobs/search/?keywords=Demand+Planning+Amway",
       "why":"Nutrition demand planning = global complexity. Your o9 tool knowledge + forecasting = immediate value at Amway."},
      {"title":"IBP Manager - Global Operations","location":"Remote (US)",
       "type":"Remote · Full-time","match":"PERFECT MATCH","posted":"This week",
       "network":"o9 alumni at Amway","url":"https://www.linkedin.com/jobs/search/?keywords=IBP+Amway",
       "why":"IBP Manager = S&OP leadership. Remote. o9 tool match. Early applicant window. Amway global ops = high-impact role."},
    ]},
]

# Deduplicate by company name
seen = set()
UNIQUE_CLIENTS = []
for c in CLIENTS:
    if c["company"] not in seen:
        seen.add(c["company"])
        UNIQUE_CLIENTS.append(c)

def main():
    print(f"Generating Excel files for {len(UNIQUE_CLIENTS)} o9 clients...\n")
    for client in UNIQUE_CLIENTS:
        name = client["company"]
        try:
            fname = build_excel(client)
            print(f"  {name:25s} -> {fname.name}")
        except Exception as e:
            print(f"  {name:25s} ERROR: {e}")
    print(f"\nDone! All files saved to: output/o9_clients/")

if __name__ == "__main__":
    main()

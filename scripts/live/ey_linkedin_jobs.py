from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# EY Brand Colors
EY_YELLOW  = "FFE600"
EY_DARK    = "2E2E38"
EY_GRAY    = "F6F6FA"
EY_WHITE   = "FFFFFF"
EY_GREEN   = "C6EFCE"
EY_GREEN_F = "375623"
EY_BLUE    = "DDEBF7"
EY_BLUE_F  = "1F4E79"

thin   = Side(style="thin",   color="BFBFBF")
medium = Side(style="medium",  color="2E2E38")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK  = Border(left=thin, right=thin, top=thin, bottom=medium)

def font(bold=False, color="2E2E38", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def fill(color):
    return PatternFill("solid", start_color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_a():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink_cell(cell, url, label="View on LinkedIn →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
    cell.alignment = center()
    cell.border = BORDER

# ── Sheet 1: EY Top 3 Jobs ───────────────────────────────────────────────────
ws = wb.active
ws.title = "EY Top 3 Jobs"

# Column widths
col_widths = [4, 42, 20, 18, 16, 14, 20, 28, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws.row_dimensions[1].height = 38
ws.merge_cells("A1:I1")
t = ws["A1"]
t.value = f"EY — Top 3 Supply Chain Jobs  |  LinkedIn  |  {datetime.now().strftime('%B %Y')}"
t.font = Font(name="Arial", bold=True, color=EY_DARK, size=15)
t.alignment = center()
t.fill = fill(EY_YELLOW)
t.border = BORDER

# Yellow accent bar
ws.row_dimensions[2].height = 5
ws.merge_cells("A2:I2")
ws["A2"].fill = fill(EY_DARK)

# Subtitle
ws.row_dimensions[3].height = 18
ws.merge_cells("A3:I3")
s = ws["A3"]
s.value = "30 o9 Solutions alumni at EY  ·  All roles: Early Applicant window open  ·  On-site · Full-time"
s.font = Font(name="Arial", italic=True, size=9, color=EY_DARK)
s.alignment = center()
s.fill = fill("FFFDE7")
s.border = BORDER

# Headers
ws.row_dimensions[4].height = 22
HEADERS = ["#", "Job Title", "Company", "Location", "Type", "Posted", "Applicants", "Network Edge", "Link"]
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=EY_WHITE, size=10)
    cell.fill = fill(EY_DARK)
    cell.alignment = center()
    cell.border = THICK

ws.freeze_panes = "A5"

# Job Data
JOBS = [
    {
        "num": 1,
        "title": "Senior Consultant\nSupply Chain & Operations H/F",
        "company": "EY",
        "location": "Nashville, TN\n(On-site)",
        "type": "On-site · Full-time",
        "posted": "1 week ago",
        "applicants": "14 applicants",
        "network": "30 o9 alumni\nat EY ⭐",
        "url": "https://www.linkedin.com/jobs/view/4397937369/",
        "match": "PERFECT MATCH",
        "match_bg": EY_GREEN, "match_fg": EY_GREEN_F,
        "why": "30 o9 Solutions alumni at EY — direct network. Only 14 applicants — very low competition. Responses managed off LinkedIn — direct recruiter contact. Early applicant window OPEN.",
        "row_fill": "FFFDE7",
    },
    {
        "num": 2,
        "title": "Supply Chain Health\nPerformance Improvement\nManager - Consulting",
        "company": "EY",
        "location": "Atlanta, GA\n(On-site)",
        "type": "On-site · Full-time",
        "posted": "2 weeks ago",
        "applicants": "Early applicant",
        "network": "30 o9 alumni\nat EY ⭐",
        "url": "https://www.linkedin.com/jobs/view/4393148275/",
        "match": "PERFECT MATCH",
        "match_bg": EY_GREEN, "match_fg": EY_GREEN_F,
        "why": "Performance Improvement = core S&OP consulting. Manager level matches your 7+ yrs experience. 401(k) benefit. 30 o9 alumni network at EY.",
        "row_fill": EY_GRAY,
    },
    {
        "num": 3,
        "title": "Supply Chain Health\nPerformance Improvement\nManager - Consulting",
        "company": "EY",
        "location": "Chantilly, VA\n(On-site)",
        "type": "On-site · Full-time",
        "posted": "2 weeks ago",
        "applicants": "Early applicant",
        "network": "333 school\nalumni at EY",
        "url": "https://www.linkedin.com/jobs/view/4393142527/",
        "match": "PERFECT MATCH",
        "match_bg": EY_GREEN, "match_fg": EY_GREEN_F,
        "why": "Same role as #2 — different location. Apply to both for 2x chances. 333 school alumni at EY. 401(k) benefit. Early applicant window open.",
        "row_fill": "FFFDE7",
    },
]

for job in JOBS:
    r = job["num"] + 4
    ws.row_dimensions[r].height = 56
    rf = fill(job["row_fill"])

    values = [job["num"], job["title"], job["company"], job["location"],
              job["type"], job["posted"], job["applicants"], job["network"]]

    for c, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = rf
        cell.border = BORDER
        cell.font = font(bold=(c == 2), size=9)
        cell.alignment = center() if c in [1, 5, 6, 7, 8] else left_a()

    hyperlink_cell(ws.cell(row=r, column=9), job["url"])
    ws.cell(row=r, column=9).fill = rf

# Spacer
ws.row_dimensions[8].height = 10
ws.merge_cells("A8:I8")
ws["A8"].fill = fill(EY_DARK)

# Why Apply section
ws.row_dimensions[9].height = 18
ws.merge_cells("A9:I9")
h = ws["A9"]
h.value = "WHY THESE ROLES MATCH YOUR PROFILE"
h.font = Font(name="Arial", bold=True, color=EY_WHITE, size=11)
h.alignment = center()
h.fill = fill(EY_DARK)
h.border = BORDER

for i, job in enumerate(JOBS):
    r = 10 + i
    ws.row_dimensions[r].height = 50
    ws.merge_cells(f"B{r}:I{r}")
    label = ws.cell(row=r, column=1, value=f"#{job['num']}")
    label.font = Font(name="Arial", bold=True, color=EY_WHITE, size=10)
    label.fill = fill(EY_DARK)
    label.alignment = center()
    label.border = BORDER

    why = ws.cell(row=r, column=2, value=job["why"])
    why.font = font(size=9)
    why.fill = fill(job["row_fill"])
    why.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    why.border = BORDER

filename = "ey_top3_linkedin_jobs.xlsx"
wb.save(filename)
print(f"Saved: {filename}")

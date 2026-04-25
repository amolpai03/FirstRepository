"""
walmart_jobs.py — Generate a branded Excel file with top 5 Walmart job listings.
Walmart is an o9 Solutions client (Bentonville, AR | Fortune #1).

Brand: Blue #0071CE / Yellow #FFC220 / White #FFFFFF
Jobs sourced from LinkedIn (April 2026).

Run: python scripts/live/walmart_jobs.py
Output: output/walmart_jobs.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# ── Brand ─────────────────────────────────────────────────────────────────────
BLUE   = "0071CE"
YELLOW = "FFC220"
WHITE  = "FFFFFF"

OUTPUT_PATH = Path(
    r"C:\Users\amolp\OneDrive\Documents\GitHub\FirstRepository\output\walmart_jobs.xlsx"
)

# ── Styles ────────────────────────────────────────────────────────────────────
thin   = Side(style="thin",   color="BFBFBF")
medium = Side(style="medium", color="404040")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK  = Border(left=thin, right=thin, top=thin, bottom=medium)


def font(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)


def fill(color):
    return PatternFill("solid", start_color=color, end_color=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_a():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def hyperlink_cell(cell, url, label="View on LinkedIn"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
    cell.alignment = center()
    cell.border = BORDER


# ── Job Data ──────────────────────────────────────────────────────────────────
JOBS = [
    {
        "title":      "(USA) Senior Manager, Supply Chain Management",
        "department": "Supply Chain",
        "location":   "Bentonville, AR (On-site)",
        "salary":     "$90K–$180K/yr",
        "type":       "Full-time",
        "posted":     "2 days ago",
        "status":     "ACTIVE",
        "match":      "PERFECT MATCH",
        "match_bg":   "C6EFCE",
        "match_fg":   "375623",
        "why":        "o9 client — Walmart uses o9 platform Amolp knows inside-out; 7+ yrs SC mgmt + S&OP leadership",
        "url":        "https://www.linkedin.com/jobs/view/4401201828/",
        "linkedin_id": 4401201828,
    },
    {
        "title":      "Senior Manager, Supply Chain Management (Sat–Tue)",
        "department": "Supply Chain",
        "location":   "Bentonville, AR (On-site)",
        "salary":     "$90K–$180K/yr",
        "type":       "Full-time",
        "posted":     "5 days ago",
        "status":     "ACTIVE",
        "match":      "PERFECT MATCH",
        "match_bg":   "C6EFCE",
        "match_fg":   "375623",
        "why":        "Same SC mgmt profile; Sat–Tue shift = less competition; o9 alumni network at Walmart",
        "url":        "https://www.linkedin.com/jobs/view/4402067070/",
        "linkedin_id": 4402067070,
    },
    {
        "title":      "Senior Director I, Integrated Planning (S&OP)",
        "department": "Integrated Planning",
        "location":   "Bentonville, AR (On-site)",
        "salary":     "Competitive",
        "type":       "Full-time",
        "posted":     "Recent",
        "status":     "ACTIVE",
        "match":      "PERFECT MATCH",
        "match_bg":   "C6EFCE",
        "match_fg":   "375623",
        "why":        "S&OP is Amolp's core expertise; Integrated Business Plan alignment = direct experience match",
        "url":        "https://www.linkedin.com/jobs/view/senior-director-i-integrated-planning-at-walmart-2867715265/",
        "linkedin_id": 2867715265,
    },
    {
        "title":      "Manager, Process Improvement",
        "department": "Operations",
        "location":   "Bentonville, AR (On-site)",
        "salary":     "$80K–$155K/yr",
        "type":       "Full-time",
        "posted":     "4 days ago",
        "status":     "ACTIVE",
        "match":      "STRONG MATCH",
        "match_bg":   "FFEB9C",
        "match_fg":   "7F6000",
        "why":        "SC process optimization background; cross-functional leadership across planning org",
        "url":        "https://www.linkedin.com/jobs/view/4400155575/",
        "linkedin_id": 4400155575,
    },
    {
        "title":      "Senior Analyst, Business Analysis and Insights",
        "department": "Analytics",
        "location":   "Bentonville, AR (On-site)",
        "salary":     "$60K–$110K/yr",
        "type":       "Full-time",
        "posted":     "2 days ago",
        "status":     "ACTIVE",
        "match":      "STRONG MATCH",
        "match_bg":   "FFEB9C",
        "match_fg":   "7F6000",
        "why":        "Data-driven SC insights; KPI reporting + financial modeling aligns with planning analytics work",
        "url":        "https://www.linkedin.com/jobs/view/4401211653/",
        "linkedin_id": 4401211653,
    },
]

HEADERS    = ["#", "Job Title", "Department", "Location", "Salary", "Match", "Why Amolp Fits", "LinkedIn Link"]
COL_WIDTHS = [4,   36,          18,           20,         18,       16,      52,                18]


# ── Builder ───────────────────────────────────────────────────────────────────
def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Walmart Jobs"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ncols    = len(HEADERS)
    last_col = get_column_letter(ncols)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = "WALMART — Top 5 Supply Chain Jobs for Amolp Pai  |  LinkedIn  |  April 2026"
    t.font      = Font(name="Arial", bold=True, color=WHITE, size=14)
    t.alignment = center()
    t.fill      = fill(BLUE)
    t.border    = BORDER

    # ── Row 2: Accent bar ─────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 5
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].fill = fill(YELLOW)

    # ── Row 3: Sub-header ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    ws.merge_cells(f"A3:{last_col}3")
    s = ws["A3"]
    s.value     = "o9 Solutions Client  |  Fortune #1  |  49,736 Open Jobs on LinkedIn  |  Bentonville, AR HQ"
    s.font      = Font(name="Arial", italic=True, size=9, color="1A1A1A")
    s.alignment = center()
    s.fill      = fill("F5F5F5")
    s.border    = BORDER

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill      = fill(BLUE)
        cell.alignment = center()
        cell.border    = THICK

    # Freeze at row 5
    ws.freeze_panes = "A5"

    # ── Rows 5–9: Job data ────────────────────────────────────────────────────
    alt_colors = ["F0F5FF", WHITE]
    for i, job in enumerate(JOBS):
        r  = i + 5
        ws.row_dimensions[r].height = 52
        rf = fill(alt_colors[i % 2])

        values = [
            i + 1,
            job["title"],
            job["department"],
            job["location"],
            job["salary"],
            job["match"],
            job["why"],
            None,           # LinkedIn Link column — handled separately
        ]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.fill   = rf

            if c == 6:  # Match badge
                cell.value     = job["match"]
                cell.fill      = fill(job["match_bg"])
                cell.font      = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
                cell.alignment = center()
            elif c == 5:  # Salary
                cell.font      = font(size=9)
                cell.alignment = center()
            elif c == 2:  # Job title — bold, left
                cell.font      = font(bold=True, size=9)
                cell.alignment = left_a()
            elif c == 7:  # Why fits — left, wrap
                cell.font      = font(size=9)
                cell.alignment = left_a()
            elif c == 1:  # Row number
                cell.font      = font(size=9, color="595959")
                cell.alignment = center()
            else:
                cell.font      = font(size=9)
                cell.alignment = center()

        # LinkedIn link cell
        link_cell = ws.cell(row=r, column=ncols)
        hyperlink_cell(link_cell, job["url"], "View on LinkedIn")
        link_cell.fill = rf

    # ── Row 11: o9 alumni note ────────────────────────────────────────────────
    note_row = len(JOBS) + 6
    ws.row_dimensions[note_row].height = 18
    ws.merge_cells(f"A{note_row}:{last_col}{note_row}")
    note = ws[f"A{note_row}"]
    note.value     = (
        "Walmart is an o9 Solutions client — leverage your o9 alumni network. "
        "4 o9 alumni currently work at Walmart."
    )
    note.font      = Font(name="Arial", italic=True, size=9, color="595959")
    note.alignment = center()
    note.fill      = fill("FFF9E6")
    note.border    = BORDER

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"Size:  {OUTPUT_PATH.stat().st_size:,} bytes")


if __name__ == "__main__":
    build_excel()

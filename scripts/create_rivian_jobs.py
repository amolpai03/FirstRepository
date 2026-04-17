from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE  = "FFFFFF"
thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="595959"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def hyperlink(cell, url, label="View Job →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Rivian brand colors
RIVIAN_GREEN  = "00A859"   # Rivian green
RIVIAN_DARK   = "1A3A2A"   # Dark forest green
RIVIAN_LIGHT  = "E8F5EE"   # Very light green tint
RIVIAN_MID    = "C8E6D4"   # Mid green
GOLD          = "F5A623"
GOLD_LIGHT    = "FFF8EC"
BLUE_LIGHT    = "EAF3FF"
BLUE_MID      = "1F4E79"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — RIVIAN TOP 3 JOBS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Rivian Top 3 Jobs"

# ── Title banner ──
ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "RIVIAN — Top 3 Supply Chain & Planning Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=RIVIAN_DARK)
t.border = border

# ── Spacer ──
ws1.row_dimensions[2].height = 6
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=RIVIAN_GREEN)

# ── Section header ──
ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "All 3 roles have network contacts (Rachitha + others)  ·  38 school alumni at Rivian  ·  Rivian company ID: 737010"
s.font = Font(name="Arial", italic=True, size=9, color=RIVIAN_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color=RIVIAN_LIGHT)
s.border = border

# ── Column headers ──
HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match Level",
           "Applicants", "Posted", "Key Tools Required", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=RIVIAN_GREEN)
    cell.alignment = center()
    cell.border = thick_bottom

ws1.freeze_panes = "A5"

# ── Job data ──
JOBS = [
    {
        "num": 1,
        "title": "Staff Advanced Planning Solutions Engineer",
        "location": "Atlanta, GA\n(also Palo Alto CA / Irvine CA / Normal IL)",
        "type": "On-site · Full-time",
        "salary": "$149,700 – $187,100\n(CA rate; Atlanta likely similar)",
        "match": "HIGH MATCH",
        "match_color": "C6EFCE",
        "match_font": "375623",
        "applicants": "36 clicked apply",
        "posted": "Reposted 2 wks ago",
        "tools": "Kinaxis Rapid Response (Maestro) — REQUIRED\nO9 · Blue Yonder · SAP IBP · SQL\n7–10 yrs hands-on APS",
        "url": "https://www.linkedin.com/jobs/view/4362093325/",
        "network": "Rachitha + others\nin your network",
        "strategy": "APPLY NOW — LinkedIn says 'Job match is HIGH'. Your o9/BY/Kinaxis background is an exact fit. Message Rachitha before applying for referral boost.",
        "fill": RIVIAN_LIGHT,
    },
    {
        "num": 2,
        "title": "Sr. Manager, Production Program Planning",
        "location": "Irvine, CA\n(open to Atlanta relocation later)",
        "type": "On-site · Full-time",
        "salary": "$159,200 – $199,000\n(SoCal) / $173K–$216K (NorCal)\n$145K–$182K (IL/MI)",
        "match": "MEDIUM MATCH",
        "match_color": "FFEB9C",
        "match_font": "7F6000",
        "applicants": "16 clicked apply\n(very low!)",
        "posted": "1 week ago",
        "tools": "Kinaxis — REQUIRED\nSAP PP/DS — strongly preferred\nERP · Databricks · SQL · Hex\n10+ yrs manufacturing SC",
        "url": "https://www.linkedin.com/jobs/view/4393876831/",
        "network": "Rachitha + others\nin your network",
        "strategy": "Only 16 applicants — low competition. Kinaxis is required which matches your APS tool expertise. Highlight S&OP leadership + capacity planning experience. Sr. Manager level = strong comp ($199K SoCal).",
        "fill": GOLD_LIGHT,
    },
    {
        "num": 3,
        "title": "Senior Global Supply Manager, CapEx – Battery",
        "location": "Social Circle, GA\n(also Irvine CA / Palo Alto CA / Normal IL)",
        "type": "On-site · Full-time",
        "salary": "$98,500 – $123,100\n(IL/GA) / $107K–$135K (Irvine)\n$117K–$147K (Palo Alto)",
        "match": "MEDIUM MATCH",
        "match_color": "DDEBF7",
        "match_font": "1F4E79",
        "applicants": "44 clicked apply",
        "posted": "Reposted 1 wk ago",
        "tools": "Strategic Sourcing · CapEx Management\nSupplier Development · Negotiations\nAutomotive/EV manufacturing knowledge\nGlobal supply chain experience",
        "url": "https://www.linkedin.com/jobs/view/4383400286/",
        "network": "Rachitha + others\nin your network",
        "strategy": "Stretch role — focuses on CapEx/supplier mgmt vs planning tech. Apply if you have strategic sourcing experience. Social Circle GA = new Rivian factory location. Salary range lower than other 2 roles.",
        "fill": BLUE_LIGHT,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 90
    fill = PatternFill("solid", start_color=job["fill"])

    data = [
        job["num"], job["title"], job["location"], job["type"],
        job["salary"], job["match"], job["applicants"], job["posted"],
        job["tools"], job["url"], job["network"], job["strategy"]
    ]

    for c, val in enumerate(data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:  # #
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=14, color=RIVIAN_GREEN)
            cell.alignment = center()
        elif c == 2:  # Title
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=RIVIAN_DARK)
            cell.alignment = left()
        elif c == 6:  # Match badge
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_font"])
            cell.fill = PatternFill("solid", start_color=job["match_color"])
            cell.alignment = center()
        elif c == 7:  # Applicants
            cell.value = val
            is_low = "16" in str(val)
            cell.font = Font(name="Arial", bold=is_low, size=9, color="C00000" if is_low else "595959")
            cell.alignment = center()
        elif c == 10:  # Link
            hyperlink(cell, val)
            cell.fill = fill
        elif c == 12:  # Strategy
            cell.value = val
            cell.font = nf(size=9, color="1A1A1A")
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

# Column widths
COL_WIDTHS = [4, 30, 22, 16, 22, 14, 16, 14, 32, 12, 18, 44]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RESUME TAILORING GUIDE
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING — What to Highlight for Each Rivian Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=RIVIAN_DARK)
t2.border = border

ws2.row_dimensions[2].height = 6
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=RIVIAN_GREEN)

HEADERS2 = ["Role", "Must Highlight in Resume", "Keywords to Include", "Experience to Lead With",
            "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=RIVIAN_GREEN)
    cell.alignment = center()
    cell.border = thick_bottom

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Staff Advanced Planning\nSolutions Engineer",
        "• o9/Blue Yonder hands-on implementation\n• Kinaxis Maestro workbook authoring\n• S&OP: Demand/Supply/Fulfillment\n• ERP-APS data integration experience\n• APS system optimization projects",
        "Kinaxis, Rapid Response, Maestro, O9, Blue Yonder, SAP IBP, S&OP, APS, workbook, data integration, supply planning, demand planning, SQL",
        "Lead with o9 + Blue Yonder project examples. Quantify: # workbooks built, % improvement in planning accuracy, # integrations delivered",
        "On-site (Atlanta/Palo Alto required)\nNo remote option",
        "Highlight relocation flexibility or proximity to one of the 4 locations. Frame as 'open to relocation' in cover letter",
        "APPLY FIRST\n(HIGH match)"
    ),
    (
        "Sr. Manager, Production\nProgram Planning",
        "• S&OP leadership at senior level\n• Capacity planning / RCCP experience\n• Kinaxis usage (any module)\n• Cross-functional leadership\n• Automotive/manufacturing SC preferred",
        "Kinaxis, Master Production Plan, RCCP, S&OP, S&OE, capacity planning, scenario modeling, SAP PP/DS, multi-plant, order allocation, BOM",
        "Lead with S&OP program ownership. Show scope: # plants, $ revenue covered, team size managed. Any RCCP or MPS work is gold here",
        "Deep manufacturing ops experience preferred (EV/auto industry). 10+ yrs manufacturing planning vs consulting",
        "Frame consulting SC planning projects as 'client-side embedded' work. Emphasize any manufacturing client engagements. Only 16 applicants — worth the stretch!",
        "APPLY — LOW\nCOMPETITION"
    ),
    (
        "Senior Global Supply Manager,\nCapEx – Battery",
        "• Strategic sourcing experience\n• Supplier negotiation & management\n• CapEx project involvement\n• Manufacturing engineering collaboration\n• Cross-functional program management",
        "Strategic sourcing, supplier development, CapEx, purchase agreements, SOW, cost negotiations, supplier qualification, industrialization, manufacturing engineering",
        "Lead with any supplier negotiation, vendor management, or procurement involvement. Even client-facing SC consulting with procurement scope works",
        "Primarily a sourcing/procurement role — not planning tech. Lower salary vs other 2 roles ($98K-$147K)",
        "Apply only if you have solid sourcing/supplier mgmt experience. Otherwise focus effort on Jobs 1 & 2 which better match your planning tech background",
        "OPTIONAL\n(stretch role)"
    ),
]

ROW_FILLS = [RIVIAN_LIGHT, GOLD_LIGHT, BLUE_LIGHT]
PRIORITY_COLORS = [("C6EFCE","375623"), ("FFEB9C","7F6000"), ("DDEBF7","1F4E79")]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 110
    fill = PatternFill("solid", start_color=ROW_FILLS[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border

        if c == 1:  # Role name
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=RIVIAN_DARK)
            cell.fill = PatternFill("solid", start_color=ROW_FILLS[i])
            cell.alignment = left()
        elif c == 7:  # Priority badge
            bg, fg = PRIORITY_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

COL_WIDTHS2 = [26, 42, 38, 40, 30, 38, 16]
for i, w in enumerate(COL_WIDTHS2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
output = "C:/Users/amolp/Prometheus/rivian_top3_jobs.xlsx"
wb.save(output)
print("Done! rivian_top3_jobs.xlsx saved with 2 sheets.")

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

wb = openpyxl.Workbook()

# -- McKinsey Brand Colors -------------------------------------------------
MCK_NAVY    = "002F6C"   # McKinsey deep navy (primary brand)
MCK_BLUE    = "006FBA"   # McKinsey medium blue
MCK_SKY     = "D6E9F8"   # light blue fill
MCK_TEAL    = "00A3E0"   # accent teal
MCK_GRAY    = "6D7278"   # slate gray text
MCK_LGRAY   = "F5F5F5"   # light background
MCK_WHITE   = "FFFFFF"
MCK_GREEN   = "1A7A4A"   # for match indicators
MCK_AMBER   = "D97706"   # salary callouts
MCK_BLACK   = "1A1A1A"
MCK_GOLD    = "C89B3C"   # accent gold

# -- Helpers ---------------------------------------------------------------
def fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium")
    return Border(left=t, right=t, top=t, bottom=t)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cw(ws, col, w):
    ws.column_dimensions[col].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

# ==========================================================================
# SHEET 1 — TOP 5 JOBS
# ==========================================================================
ws1 = wb.active
ws1.title = "McKinsey Top 5 Jobs"
ws1.sheet_properties.tabColor = MCK_NAVY
ws1.freeze_panes = "A5"

# Title banner
ws1.merge_cells("A1:J1")
c = ws1["A1"]
c.value = "McKinsey & Company — Top 5 Supply Chain Jobs in the USA"
c.fill = fill(MCK_NAVY)
c.font = Font(name="Calibri", bold=True, size=16, color=MCK_WHITE)
c.alignment = align("center", "center")
rh(ws1, 1, 32)

# Sub-banner
ws1.merge_cells("A2:J2")
c2 = ws1["A2"]
c2.value = ("Curated for: 7-yr SC Planning Consultant  |  Tools: o9 | OMP | Kinaxis | Blue Yonder | SAP IBP  "
            "|  S&OP | Demand | Supply | Inventory  |  USA Only  |  Pulled: April 16, 2026")
c2.fill = fill(MCK_BLUE)
c2.font = Font(name="Calibri", bold=True, size=10, color=MCK_WHITE)
c2.alignment = align("center", "center")
rh(ws1, 2, 22)

# Spacer
ws1.merge_cells("A3:J3")
ws1["A3"].fill = fill(MCK_LGRAY)
rh(ws1, 3, 5)

# Headers
headers = [
    ("A", "#",               4),
    ("B", "Job Title",       36),
    ("C", "Level / Track",   16),
    ("D", "Location(s)",     22),
    ("E", "Salary (USD)",    16),
    ("F", "Posted",          12),
    ("G", "Key Requirements", 40),
    ("H", "Tool Match",      26),
    ("I", "Practice / Group", 22),
    ("J", "Apply Link",      18),
]
for col, hdr, width in headers:
    c = ws1[f"{col}4"]
    c.value = hdr
    c.fill = fill(MCK_NAVY)
    c.font = Font(name="Calibri", bold=True, size=11, color=MCK_WHITE)
    c.alignment = align("center", "center")
    c.border = border()
    cw(ws1, col, width)
rh(ws1, 4, 22)

# Job Data
jobs = [
    {
        "num": 1,
        "title": "Digital Supply Chain Senior Analyst – Operations",
        "level": "Senior Analyst\n(Digital / Specialist Track)",
        "location": "Atlanta | Boston | Chicago\nDallas | Waltham | Washington DC\n(+ other US offices)",
        "salary": "$110,000 / yr\n(+bonus; official McKinsey range)",
        "posted": "Active – 2026\n(Ongoing basis)",
        "reqs": ("3+ yrs SC or related field; "
                 "SAP IBP / APO OR Kinaxis OR Blue Yonder OR o9 REQUIRED; "
                 "full APS project lifecycle (design-to-delivery); "
                 "S&OP, demand, inventory, supply & response planning; "
                 "data analytics & ML exposure; "
                 "undergrad STEM/business or master's preferred"),
        "tools": "SAP IBP ✅\nKinaxis ✅\nBlue Yonder ✅\no9 ✅\n(Tools explicitly required!)",
        "practice": "Operations Practice\n(Digital SC / APS Implementations)",
        "url": "https://www.mckinsey.com/careers/search-jobs/jobs/digitalsupplychainsenioranalyst-operations-79998",
        "row_fill": MCK_SKY,
        "font_color": MCK_BLACK,
        "star": True,
    },
    {
        "num": 2,
        "title": "Experienced Consultant – Supply Chain Management",
        "level": "Experienced Consultant\n(Consulting Track – Exp. Hire)",
        "location": "Multiple US offices\n(NYC | Chicago | Dallas |\nAtlanta | SF | DC | Boston)",
        "salary": "~$175K–$225K total comp\n(Base ~$150K–$185K + bonus)",
        "posted": "Active – 2026\n(Ongoing basis)",
        "reqs": ("Strong SC operations background (industry or consulting); "
                 "logistics strategy, network design, SC planning & management; "
                 "demonstrated leadership & problem-solving; "
                 "analytical/quantitative skills; MBA or advanced degree preferred; "
                 "McKinsey 2-yr Operations Excellence Program"),
        "tools": "SC Strategy ✅\nNetwork Design ✅\nS&OP / Planning ✅\nLogistics Optimization ✅",
        "practice": "Operations Practice\n(SC Management)",
        "url": "https://www.mckinsey.com/careers/search-jobs/jobs/experiencedconsultant-supplychainmanagement-34868",
        "row_fill": MCK_WHITE,
        "font_color": MCK_BLACK,
        "star": False,
    },
    {
        "num": 3,
        "title": "Manufacturing & Supply Chain – Associate, McKinsey Implementation",
        "level": "Associate\n(McKinsey Implementation Track)",
        "location": "Multiple US offices\n(Implementation hubs:\nChicago | NYC | Dallas | Atlanta)",
        "salary": "$192,000 / yr\n(Official McKinsey stated range)",
        "posted": "Active – 2026\n(Ongoing basis)",
        "reqs": ("MBA or advanced degree in SC/Operations/Eng; "
                 "manufacturing & SC operations expertise; "
                 "hands-on implementation delivery at client sites; "
                 "project management, client coaching, capability building; "
                 "80%+ travel; McKinsey Implementation = hands-on delivery arm"),
        "tools": "SC Implementation ✅\nManufacturing Ops ✅\nChange Management ✅\nClient Capability Building ✅",
        "practice": "McKinsey Implementation\n(Mfg & SC Delivery)",
        "url": "https://www.mckinsey.com/careers/search-jobs/jobs/associate-manufacturingsupplychainmckinseyimplementation-87959",
        "row_fill": MCK_LGRAY,
        "font_color": MCK_BLACK,
        "star": False,
    },
    {
        "num": 4,
        "title": "Consultant – Manufacturing & Supply Chain",
        "level": "Consultant\n(Operations Practice)",
        "location": "Multiple US offices\n(+ Ops Excellence Program:\nRotating US client sites)",
        "salary": "~$130K–$175K base\n(+bonus ~15-25%)",
        "posted": "Active – 2026\n(Ongoing basis)",
        "reqs": ("SC/manufacturing experience; "
                 "active problem-solving for clients; "
                 "network design, logistics, SC planning expertise; "
                 "McKinsey 2-yr Operations Excellence Program (dedicated ops training); "
                 "leadership demonstrated in work or activities; "
                 "MBA or relevant engineering/advanced degree"),
        "tools": "SC Planning ✅\nManufacturing Ops ✅\nNetwork Design ✅\nLogistics Optimization ✅",
        "practice": "Operations Practice\n(Mfg & SC Consulting)",
        "url": "https://www.mckinsey.com/careers/search-jobs/jobs/consultant-manufacturingsupplychain-101547",
        "row_fill": MCK_WHITE,
        "font_color": MCK_BLACK,
        "star": False,
    },
    {
        "num": 5,
        "title": "Digital Supply Chain Sr. Analyst / Analyst – Operations",
        "level": "Sr. Analyst / Analyst\n(Digital / Specialist Track)",
        "location": "Atlanta | Boston | Chicago\nDallas | Waltham | DC\n(+ other US offices)",
        "salary": "$92,000–$110,000 / yr\n(Official McKinsey range)",
        "posted": "Active – 2026\n(Ongoing basis)",
        "reqs": ("1+ yrs SC or related (Analyst) / 3+ yrs (Sr. Analyst); "
                 "SAP IBP/APO, Kinaxis, Blue Yonder, or o9 experience REQUIRED; "
                 "APS implementation lifecycle participation; "
                 "data aggregation, analytics, ML exposure; "
                 "STEM or business degree; "
                 "dual posting covers both levels – apply to senior level first"),
        "tools": "SAP IBP ✅\nKinaxis ✅\nBlue Yonder ✅\no9 ✅",
        "practice": "Operations Practice\n(Digital SC / APS)",
        "url": "https://www.mckinsey.com/careers/search-jobs/jobs/digitalsupplychainsranalystanalyst-operations-83284",
        "row_fill": MCK_LGRAY,
        "font_color": MCK_BLACK,
        "star": False,
    },
]

for idx, job in enumerate(jobs):
    row = 5 + idx
    row_data = [
        job["num"],
        job["title"],
        job["level"],
        job["location"],
        job["salary"],
        job["posted"],
        job["reqs"],
        job["tools"],
        job["practice"],
        "-> Apply Here",
    ]
    for ci, col in enumerate(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]):
        cell = ws1[f"{col}{row}"]
        cell.value = row_data[ci]
        cell.fill = fill(job["row_fill"])
        cell.font = Font(name="Calibri", size=10,
                         color=MCK_NAVY if col in ("B", "H") else MCK_BLACK,
                         bold=(col == "B"))
        cell.alignment = align("left" if col != "A" else "center", "center", wrap=True)
        cell.border = border()

    # Hyperlink on J
    j_cell = ws1[f"J{row}"]
    j_cell.hyperlink = job["url"]
    j_cell.font = Font(name="Calibri", size=10, color=MCK_BLUE,
                       bold=True, underline="single")
    j_cell.alignment = align("center", "center")

    # Thick border + star highlight for Job 1
    if job["star"]:
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws1[f"{col}{row}"].border = thick_border()
        ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color=MCK_NAVY)

    rh(ws1, row, 70)

# Best match callout
ws1.merge_cells("A10:J10")
bm = ws1["A10"]
bm.value = ("BEST MATCH — Job #1: Digital SC Senior Analyst is the ONLY McKinsey role that explicitly requires "
            "your exact tools: SAP IBP + Kinaxis + Blue Yonder + o9.  "
            "Job #3: McKinsey Implementation Associate pays $192K and leverages hands-on delivery expertise.  "
            "Apply to #1 + #3 first.")
bm.fill = fill(MCK_NAVY)
bm.font = Font(name="Calibri", bold=True, size=10, color=MCK_WHITE)
bm.alignment = align("center", "center", wrap=True)
rh(ws1, 10, 30)

# Footer
ws1.merge_cells("A12:J12")
cf = ws1["A12"]
cf.value = ("Jobs sourced from mckinsey.com/careers  |  Salaries from official McKinsey postings & Glassdoor/Levels.fyi  "
            "|  All jobs active USA postings as of April 16, 2026  "
            "|  McKinsey accepts applications on an ongoing basis — apply any time but follow up proactively")
cf.fill = fill(MCK_LGRAY)
cf.font = Font(name="Calibri", size=9, italic=True, color=MCK_GRAY)
cf.alignment = align("center", "center", wrap=True)
rh(ws1, 12, 30)

# ==========================================================================
# SHEET 2 — RESUME TAILORING GUIDE
# ==========================================================================
ws2 = wb.create_sheet("Resume Tailoring Guide")
ws2.sheet_properties.tabColor = MCK_BLUE

# Banner
ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "McKinsey Resume Tailoring — SC Planning Consultant with 7 Years Experience"
c.fill = fill(MCK_NAVY)
c.font = Font(name="Calibri", bold=True, size=14, color=MCK_WHITE)
c.alignment = align("center", "center")
rh(ws2, 1, 30)

ws2.merge_cells("A2:G2")
c2 = ws2["A2"]
c2.value = ("McKinsey values: Solve complex problems | Data-driven insights | "
            "Client impact | Team leadership | Operations excellence")
c2.fill = fill(MCK_BLUE)
c2.font = Font(name="Calibri", bold=True, size=10, color=MCK_WHITE)
c2.alignment = align("center", "center")
rh(ws2, 2, 20)

ws2.merge_cells("A3:G3")
ws2["A3"].fill = fill(MCK_LGRAY)
rh(ws2, 3, 5)

resume_headers = [
    ("A", "Job #", 8),
    ("B", "Job Title (Short)", 28),
    ("C", "McKinsey Keyword / Theme", 26),
    ("D", "Your Experience to Highlight", 36),
    ("E", "Power Phrases to Use", 36),
    ("F", "Skills to Front-Load", 30),
    ("G", "Watch Out For", 28),
]
for col, hdr, width in resume_headers:
    c = ws2[f"{col}4"]
    c.value = hdr
    c.fill = fill(MCK_NAVY)
    c.font = Font(name="Calibri", bold=True, size=10, color=MCK_WHITE)
    c.alignment = align("center", "center", wrap=True)
    c.border = border()
    cw(ws2, col, width)
rh(ws2, 4, 22)

resume_rows = [
    (
        "1",
        "Digital SC\nSenior Analyst",
        "APS Implementation\nSAP IBP / Kinaxis / BY / o9\nFull Project Lifecycle",
        "Call out every APS tool you've used with project details: platform, client industry, "
        "scope (demand/supply/inventory/S&OP modules), go-live status, team size. "
        "Quantify: % forecast accuracy improvement, $ inventory reduction, planning cycle time cut.",
        '"Led end-to-end Kinaxis RapidResponse implementation for $3B manufacturer..."\n'
        '"Configured SAP IBP demand sensing module reducing forecast error by 18%..."\n'
        '"Delivered o9 S&OP solution across 5 business units with 20-week go-live"',
        "SAP IBP | Kinaxis | Blue Yonder | o9 | APS Implementation | Demand Planning | "
        "S&OP | Supply Planning | Inventory Optimization | Advanced Analytics | "
        "Full Lifecycle Delivery | Onshore-Offshore",
        "McKinsey values intellectual rigor — show HOW you solved the problem, not just "
        "WHAT tool you used. Don't just list tools; describe the business problem, approach, "
        "and outcome. This is still McKinsey — they want problem-solvers, not just technologists.",
    ),
    (
        "2",
        "Experienced\nConsultant – SCM",
        "SC Strategy\nOperations Excellence\nProblem Solving",
        "Lead with strategic SC consulting outcomes: network redesign, S&OP program design, "
        "planning process transformation. McKinsey Operations Excellence Program is 2 years — "
        "show you can operate at strategy level, not just implementation.",
        '"Redesigned end-to-end S&OP process for $5B CPG client reducing excess inventory $40M..."\n'
        '"Led SC network optimization across 12 DCs, cutting logistics cost 15%..."\n'
        '"Built and coached client SC planning team of 8 to self-sufficiency"',
        "SC Strategy | Operations Excellence | S&OP Design | Network Optimization | "
        "Logistics | Process Transformation | Executive Stakeholder Management | "
        "Analytical Problem-Solving | MBA (if applicable)",
        "This is McKinsey's generalist consulting track — they care more about structured "
        "thinking and business impact than tool expertise. Lead with problem-solving frameworks "
        "and quantified outcomes. Tool knowledge is a bonus, not the lead. "
        "MBA or top-tier advanced degree strongly preferred for this track.",
    ),
    (
        "3",
        "Mfg & SC\nImplementation Assoc.",
        "Hands-on Delivery\nCapability Building\nClient Coaching",
        "McKinsey Implementation is about doing, not advising — highlight times you "
        "were embedded at client sites, led change management, built client capabilities, "
        "ran day-to-day operations transformation. SC delivery + coaching is the core.",
        '"Embedded 6 months at client site leading Kinaxis go-live and training 40+ planners..."\n'
        '"Built SC planning playbook adopted by client team post-engagement..."\n'
        '"Coached client demand planning team from 35% to 72% forecast accuracy in 3 months"',
        "Implementation Delivery | Capability Building | Client Coaching | Change Management | "
        "Embedded Consulting | SC Planning Training | S&OP Facilitation | Project Management",
        "McKinsey Implementation hires specifically for delivery DNA — they don't want pure "
        "strategy consultants here. Emphasize times you got your hands dirty: ran client "
        "meetings, configured systems, trained users, measured adoption. MBA is standard entry.",
    ),
    (
        "4",
        "Consultant –\nMfg & SC",
        "Manufacturing Ops\nSC Planning\nOperations Excellence Program",
        "Emphasize both manufacturing and SC experience. The Operations Excellence Program "
        "rotates consultants through operations-heavy projects. Show breadth: "
        "factory operations, SC planning, procurement, logistics.",
        '"Advised automotive client on SC-manufacturing integration reducing WIP inventory $25M..."\n'
        '"Led plant-level S&OP implementation coordinating procurement, production, and distribution..."\n'
        '"Designed SC resilience strategy spanning 3 tiers of supplier network"',
        "Manufacturing Integration | SC Planning | Operations Excellence | Procurement | "
        "Supplier Management | Production Planning | Cross-functional Leadership | "
        "LEAN / Six Sigma (bonus)",
        "This role requires manufacturing breadth in addition to SC — if you're primarily "
        "SC planning (no manufacturing exposure), pivot emphasis to #1 or #2. "
        "McKinsey's Ops Excellence Program values factory + SC generalists.",
    ),
    (
        "5",
        "Digital SC Sr.\nAnalyst / Analyst",
        "APS Tools\nData Analytics\nSC Tech Implementation",
        "Same as Job #1 but broader — can apply at either Analyst or Senior Analyst level. "
        "With 7 yrs experience, apply to SENIOR ANALYST level only. "
        "Emphasize data analytics, ML exposure, and APS tool hands-on work.",
        '"Leveraged advanced analytics to improve demand forecast accuracy from 67% to 84%..."\n'
        '"Developed custom IBP reporting models for S&OP executive dashboard..."\n'
        '"Implemented Kinaxis modules (Demand, Supply, S&OE) in 12-week sprint"',
        "SAP IBP | Kinaxis | Blue Yonder | o9 | Data Analytics | ML/AI in SC Planning | "
        "APS Implementation | Demand Sensing | Inventory Optimization | STEM Background",
        "With 7 years experience, $92K-$110K is likely below your market value. "
        "Use this as a BACKUP or if you specifically want to be inside McKinsey's "
        "digital practice for career trajectory into more senior McKinsey roles. "
        "Apply to Senior Analyst level, not Analyst.",
    ),
]

for idx, row_data in enumerate(resume_rows):
    row = 5 + idx
    rf = MCK_SKY if idx == 0 else ("F0F7FF" if idx % 2 == 0 else MCK_LGRAY)
    for ci, col in enumerate(["A", "B", "C", "D", "E", "F", "G"]):
        c = ws2[f"{col}{row}"]
        c.value = row_data[ci]
        c.fill = fill(rf)
        c.font = Font(name="Calibri", size=9,
                      color=MCK_NAVY if ci < 2 else "333333",
                      bold=(ci == 0))
        c.alignment = align("left" if ci > 0 else "center", "top", wrap=True)
        c.border = border()
    rh(ws2, row, 90)

# McKinsey Application Tips section
tip_row = 11
ws2.merge_cells(f"A{tip_row}:G{tip_row}")
ct = ws2[f"A{tip_row}"]
ct.value = "McKinsey APPLICATION & COVER LETTER TIPS"
ct.fill = fill(MCK_NAVY)
ct.font = Font(name="Calibri", bold=True, size=12, color=MCK_WHITE)
ct.alignment = align("center", "center")
rh(ws2, tip_row, 24)

tips = [
    ("The McKinsey Mindset",
     'McKinsey screens for structured problem solvers FIRST. Before you apply, practice a case: '
     '"How would you reduce inventory for a $2B manufacturer?" Answer: MECE structure, '
     'hypothesis, data request, quantified recommendation. Even for Digital/Specialist '
     'roles, McKinsey interviewers test structured thinking.'),
    ("Resume Format",
     'McKinsey-style resume: 1 page max (2 for experienced hires), bullet-pointed achievements '
     'in "action + metric + context" format. No fluff. Every bullet should answer "so what?" '
     'Example: "Led IBP implementation (action) reducing planning cycle from 4 weeks to 5 days '
     '(metric) for $4B CPG client across 3 regions (context)." No paragraphs. Ever.'),
    ("Tool Match Strategy (Jobs 1 & 5)",
     'For Digital SC roles: literally list every APS platform you\'ve touched in a "Technical '
     'Skills" section. SAP IBP | Kinaxis | Blue Yonder | o9 | OMP — McKinsey recruiters '
     'keyword-screen these applications. Then show impact: not "used SAP IBP" but '
     '"implemented SAP IBP demand planning for $1.5B retailer, improving fill rate 8pts."'),
    ("Referral Strategy",
     'McKinsey has the strongest referral culture in consulting — a referral from a current '
     'McKinsey consultant or partner increases callback rate 8-10x. Strategy: search LinkedIn '
     'for "McKinsey Operations Supply Chain" → find alumni in your industry → reach out with '
     'a specific message about the SC practice → ask for a referral AFTER a conversation, '
     'not in the first message.'),
    ("Interview Process",
     'McKinsey SC roles: Application screening → HR screening call → 2-3 rounds of case '
     'interviews (even for digital/specialist roles). For Digital SC roles: expect 1 case '
     'interview + 1-2 experience/behavioral interviews. Operations Excellence Program roles: '
     '2-3 full case rounds. Prep: McKinsey Problem Solving Test (for some roles) + '
     'STAR stories for leadership + 2-3 polished SC impact stories with metrics.'),
]

for ti, (label, content) in enumerate(tips):
    r = tip_row + 1 + ti
    ws2.merge_cells(f"A{r}:B{r}")
    lc = ws2[f"A{r}"]
    lc.value = label
    lc.fill = fill("C8DFF5" if ti % 2 == 0 else MCK_SKY)
    lc.font = Font(name="Calibri", bold=True, size=10, color=MCK_NAVY)
    lc.alignment = align("center", "center", wrap=True)
    lc.border = border()
    ws2.merge_cells(f"C{r}:G{r}")
    tc = ws2[f"C{r}"]
    tc.value = content
    tc.fill = fill(MCK_LGRAY)
    tc.font = Font(name="Calibri", size=9, color="333333")
    tc.alignment = align("left", "center", wrap=True)
    tc.border = border()
    rh(ws2, r, 60)

# ==========================================================================
# SHEET 3 — McKINSEY FACTS & INTERVIEW PREP
# ==========================================================================
ws3 = wb.create_sheet("McKinsey Facts & Interview Prep")
ws3.sheet_properties.tabColor = MCK_TEAL
ws3.freeze_panes = "A4"

# Banner
ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "McKinsey & Company — Interview Prep & Company Facts for SC Consulting Roles"
c.fill = fill(MCK_NAVY)
c.font = Font(name="Calibri", bold=True, size=14, color=MCK_WHITE)
c.alignment = align("center", "center")
rh(ws3, 1, 30)

ws3.merge_cells("A2:F2")
c2 = ws3["A2"]
c2.value = '"The ability to make a difference" — McKinsey\'s mission. Operations Practice: 750+ partners, 1000+ dedicated practitioners globally.'
c2.fill = fill(MCK_BLUE)
c2.font = Font(name="Calibri", bold=True, size=10, color=MCK_WHITE, italic=True)
c2.alignment = align("center", "center")
rh(ws3, 2, 22)

ws3.merge_cells("A3:F3")
ws3["A3"].fill = fill(MCK_LGRAY)
rh(ws3, 3, 5)

for col, w in [("A", 26), ("B", 46), ("C", 26), ("D", 46), ("E", 22), ("F", 36)]:
    cw(ws3, col, w)

def s3_section(ws, row, title, bg=MCK_NAVY, fg=MCK_WHITE):
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = title
    c.fill = fill(bg)
    c.font = Font(name="Calibri", bold=True, size=11, color=fg)
    c.alignment = align("center", "center")
    rh(ws, row, 22)

def s3_row(ws, row, label, value, lbg=MCK_SKY, vbg=MCK_LGRAY, lstart="A", lend="B", vstart="C", vend="F"):
    ws.merge_cells(f"{lstart}{row}:{lend}{row}")
    lc = ws[f"{lstart}{row}"]
    lc.value = label
    lc.fill = fill(lbg)
    lc.font = Font(name="Calibri", bold=True, size=10, color=MCK_NAVY)
    lc.alignment = align("left", "center", wrap=True)
    lc.border = border()
    ws.merge_cells(f"{vstart}{row}:{vend}{row}")
    vc = ws[f"{vstart}{row}"]
    vc.value = value
    vc.fill = fill(vbg)
    vc.font = Font(name="Calibri", size=10, color="333333")
    vc.alignment = align("left", "center", wrap=True)
    vc.border = border()
    rh(ws, row, 42)

# Company Quick Facts
s3_section(ws3, 4, "McKINSEY AT A GLANCE")
s3_row(ws3, 5, "Full Name & HQ",
       "McKinsey & Company | Global HQ: New York, NY | Founded: 1926 | #1 ranked global management consulting firm")
s3_row(ws3, 6, "Revenue & Scale",
       "~$16B+ revenue (2024 est.) | 45,000+ consultants globally | 130+ offices in 65+ countries | ~10,000 US staff")
s3_row(ws3, 7, "Operations Practice",
       ("750+ partners and 1000+ dedicated practitioners globally. "
        "SC work spans: SC strategy, S&OP/IBP transformation, network design, digital SC, "
        "procurement excellence, manufacturing ops, and McKinsey Implementation delivery."))
s3_row(ws3, 8, "Career Tracks",
       ("Consulting Track: Business Analyst → Associate → Engagement Manager → Associate Partner → Partner → Senior Partner. "
        "Specialist/Expert Track: Analyst → Senior Analyst → Specialist → Senior Specialist → Expert → Senior Expert. "
        "McKinsey Implementation: separate delivery-focused hierarchy."))
s3_row(ws3, 9, "Digital SC Practice",
       ("McKinsey's Digital Supply Chain team implements APS platforms (SAP IBP, Kinaxis, Blue Yonder, o9) at scale. "
        "Also builds proprietary tools: Leap (inventory optimization), WAVE (warehouse analytics), "
        "and deploys ML demand sensing models. Partners with SAP, Kinaxis, and Blue Yonder officially."))
s3_row(ws3, 10, "Operations Excellence Program",
       ("2-year program for Consultants joining the Operations practice. "
        "Structured rotations through SC, manufacturing, and procurement projects. "
        "Deep operations training + global client exposure. Strong accelerator for making EM faster."))
s3_row(ws3, 11, "Culture",
       ("Up-or-out meritocracy | Obsessive client focus | Data-driven rigor | "
        "Intellectual humility expected | Diverse global teams | Long hours (60-80hrs/wk common) | "
        "80%+ travel standard | Strong alumni network (McKinsey Mafia)"))

# Interview Questions
s3_section(ws3, 13, "LIKELY INTERVIEW QUESTIONS + ANSWER FRAMEWORKS")

qs = [
    ("Q: Walk me through a SC planning problem you solved end-to-end.",
     "Structure: Client/industry → Problem (what was broken: forecast accuracy, stockouts, excess inventory, "
     "long planning cycles) → Your structured analysis (root cause, data gathered) → Solution (process redesign, "
     "APS implementation, S&OP redesign) → Quantified result. McKinsey expects: hypothesis-driven approach, "
     "data rigor, and a clear 'so what.' Practice saying this in 2 minutes, not 5."),
    ("Q: Walk me through a case — how would you reduce inventory for a $3B retailer?",
     "Framework: (1) Clarify: what type of inventory (RM/WIP/FG)? Which SKUs? What's current turnover vs target? "
     "(2) Hypotheses: demand forecast accuracy, safety stock policy, supplier lead time, obsolescence rate "
     "(3) Data request: SKU-level inventory vs turns, forecast accuracy by category, SS formula review "
     "(4) Recommendation: prioritize ABC segmentation + IBP tool implementation + 30-60-90 day roadmap. "
     "For Digital SC roles — tie this to an APS platform you'd deploy."),
    ("Q: Why McKinsey vs. Deloitte or Accenture for SC work?",
     '"McKinsey\'s Operations practice combines the strategic rigor of top-tier consulting with dedicated '
     'implementation capability through McKinsey Implementation — that unique combination lets me deliver '
     'both the strategy AND execution for SC transformation, which is where I\'ve seen the most value created. '
     'The proprietary digital tools like Leap also give clients analytical advantages I can\'t replicate '
     'at other firms." Show you know McKinsey\'s unique position.'),
    ("Q: Tell me about a time you influenced a resistant client stakeholder.",
     "STAR: Situation (VP of SC who thought IBP was just 'another IT project') → Task (get buy-in for S&OP redesign) "
     "→ Action (1-on-1 meetings, aligned S&OP to their P&L metric, showed pilot result from 1 category) "
     "→ Result (became program champion, approved full rollout). McKinsey loves client impact stories — "
     "show you moved from skepticism to sponsorship using data and relationship building."),
    ("Q: How do you decide between SAP IBP, Kinaxis, and o9 for a client?",
     "Strong answer: 'It's a function of three things: (1) existing ERP ecosystem — SAP clients lean IBP for "
     "integration simplicity; (2) planning complexity — Kinaxis excels in high-variability, multi-echelon "
     "environments like aerospace/auto; o9 shines in large-scale scenario planning with heavy ML needs; "
     "(3) organizational readiness — Kinaxis has faster time-to-value for lean IT teams. "
     "I've deployed all three; the right answer is always client-context specific.' "
     "This shows breadth and maturity — exactly what McKinsey wants for Digital SC roles."),
    ("Q: Describe a supply chain transformation you led from start to finish.",
     "Cover: (1) Diagnostic — how you assessed the current state (benchmarking, data analysis, interviews) "
     "(2) Design — how you structured the solution (process redesign + tech selection) "
     "(3) Delivery — implementation timeline, team composition, change management "
     "(4) Impact — quantified results 6-12 months post-go-live. "
     "For McKinsey Implementation track, emphasize the delivery and capability-building phases — "
     "how you left the client self-sufficient, not dependent on consultants."),
    ("Q: What's your view on AI's impact on supply chain planning?",
     '"AI is reshaping SC planning at three levels: (1) Demand sensing — ML models outperforming '
     'statistical forecasting by 15-25% MAPE in high-velocity categories; (2) Autonomous supply '
     'response — closed-loop replenishment decisions replacing manual planner interventions; '
     '(3) Scenario planning — LLMs enabling faster what-if analysis for S&OE and S&OP. '
     'The near-term constraint isn\'t the algorithm, it\'s data quality and organizational change '
     'management — two areas where consulting adds irreplaceable value." Shows thought leadership.'),
    ("Q: How do you manage 80%+ travel while maintaining quality?",
     "McKinsey expects travel commitment — don't dodge this. Answer: 'I've structured my work style around "
     "high-intensity client site weeks and disciplined weekend reset. I find that deep immersion at client "
     "sites actually accelerates delivery — being physically present with planning teams reveals data gaps "
     "and stakeholder dynamics that remote work misses. My best SC implementations happened when I was "
     "embedded on-site. I'm fully committed to the travel model.' Show enthusiasm, not reluctance."),
]

for qi, (q, a) in enumerate(qs):
    r = 14 + qi
    ws3.merge_cells(f"A{r}:B{r}")
    qc = ws3[f"A{r}"]
    qc.value = q
    qc.fill = fill("C8DFF5" if qi % 2 == 0 else MCK_SKY)
    qc.font = Font(name="Calibri", bold=True, size=9, color=MCK_NAVY)
    qc.alignment = align("left", "top", wrap=True)
    qc.border = border()
    ws3.merge_cells(f"C{r}:F{r}")
    ac = ws3[f"C{r}"]
    ac.value = a
    ac.fill = fill(MCK_LGRAY if qi % 2 == 0 else MCK_WHITE)
    ac.font = Font(name="Calibri", size=9, color="333333")
    ac.alignment = align("left", "top", wrap=True)
    ac.border = border()
    rh(ws3, r, 65)

rh(ws3, 13, 22)

# Salary Benchmarks
sal_row = 23
s3_section(ws3, sal_row, "McKINSEY SALARY BENCHMARKS — USA 2026 (SC Consulting)")
sal_data = [
    ("Business Analyst", "$92K–$110K base", "Entry / undergrad hire; Digital SC Analyst = $92-95K official"),
    ("Associate", "$190K–$200K base", "$192K official for Mfg & SC Implementation; MBA hire standard; +$30-40K bonus"),
    ("Experienced Consultant", "~$175K–$225K total", "Experienced hire equivalent to Associate level; base ~$150-185K"),
    ("Engagement Manager", "$250K–$350K total", "3-5 yrs post-MBA; strong performers promoted in 2-3 yrs from Associate"),
    ("Associate Partner", "$400K–$600K+ total", "Pre-partner track; equity begins; significant bonus upside"),
    ("Partner", "$800K–$2M+ total", "Equity + base + book of business; highly variable based on performance"),
    ("Digital SC Sr. Analyst", "$110K base official", "Specialist track; below consulting track but tool-specific; growth path to Specialist/Expert"),
]
for si, (level, salary, note) in enumerate(sal_data):
    r = sal_row + 1 + si
    ws3.merge_cells(f"A{r}:B{r}")
    lc = ws3[f"A{r}"]
    lc.value = level
    lc.fill = fill(MCK_SKY if si % 2 == 0 else "D6E9F8")
    lc.font = Font(name="Calibri", bold=True, size=10, color=MCK_NAVY)
    lc.alignment = align("center", "center", wrap=True)
    lc.border = border()
    ws3.merge_cells(f"C{r}:D{r}")
    sc2 = ws3[f"C{r}"]
    sc2.value = salary
    sc2.fill = fill("D5F5E3" if si % 2 == 0 else "EAFAF1")
    sc2.font = Font(name="Calibri", bold=True, size=11, color=MCK_NAVY)
    sc2.alignment = align("center", "center")
    sc2.border = border()
    ws3.merge_cells(f"E{r}:F{r}")
    nc = ws3[f"E{r}"]
    nc.value = note
    nc.fill = fill(MCK_LGRAY)
    nc.font = Font(name="Calibri", size=9, color="555555", italic=True)
    nc.alignment = align("left", "center", wrap=True)
    nc.border = border()
    rh(ws3, r, 32)

# McKinsey SC Tracks guide
tr_row = 31
s3_section(ws3, tr_row, "McKINSEY TRACK GUIDE — Which Path Fits Your Background?")
tracks = [
    ("Consulting Track\n(Experienced Consultant / Associate)",
     ("Best if: You want to do high-level SC strategy, transformation design, and client advisory. "
      "Entry at Experienced Consultant (non-MBA) or Associate (MBA) level. Day-to-day: "
      "diagnostic → design → implementation recommendation. 80%+ travel. Tools secondary to "
      "structured problem-solving. Your 7-yr consulting background is the primary qualifier. "
      "McKinsey Operations Excellence Program gives structured ops training.")),
    ("Digital SC Track\n(Sr. Analyst / Specialist)",
     ("Best if: You want to leverage tool expertise (Kinaxis, SAP IBP, BY, o9) inside McKinsey. "
      "Jobs #1 and #5 are here. Day-to-day: APS implementation, data analytics, ML deployment. "
      "Salary lower ($110K) than consulting track but career path leads to Specialist/Expert/Senior Expert. "
      "Your tool breadth (all 4 platforms) is the strongest differentiator in the market — "
      "most candidates know 1-2 tools, you know all 4.")),
    ("McKinsey Implementation Track\n(Associate – Mfg & SC)",
     ("Best if: You want $192K base AND hands-on delivery at client sites. "
      "Job #3 is here. McKinsey Implementation is the 'doing' arm of McKinsey — "
      "embedded at client, building capability, leading change management. "
      "Best of both worlds: McKinsey brand + implementation execution. "
      "Requires MBA. If no MBA, this is a strong motivator to get one — "
      "or negotiate an 'experienced hire' exception with an exceptional portfolio.")),
]
for ti, (track, desc) in enumerate(tracks):
    r = tr_row + 1 + ti
    ws3.merge_cells(f"A{r}:B{r}")
    tc = ws3[f"A{r}"]
    tc.value = track
    tc.fill = fill(MCK_SKY if ti == 0 else ("C8DFF5" if ti == 1 else "D6E9F8"))
    tc.font = Font(name="Calibri", bold=True, size=10, color=MCK_NAVY)
    tc.alignment = align("center", "center", wrap=True)
    tc.border = border()
    ws3.merge_cells(f"C{r}:F{r}")
    dc = ws3[f"C{r}"]
    dc.value = desc
    dc.fill = fill(MCK_LGRAY)
    dc.font = Font(name="Calibri", size=9, color="333333")
    dc.alignment = align("left", "top", wrap=True)
    dc.border = border()
    rh(ws3, r, 72)

# Footer
last_row = 35
ws3.merge_cells(f"A{last_row}:F{last_row}")
cf = ws3[f"A{last_row}"]
cf.value = ("Jobs sourced from mckinsey.com/careers  |  Salary data from official McKinsey postings + Glassdoor + Levels.fyi + Management Consulted 2026  "
            "|  All positions confirmed active USA-wide  |  McKinsey accepts applications on an ongoing basis")
cf.fill = fill(MCK_LGRAY)
cf.font = Font(name="Calibri", size=8, italic=True, color=MCK_GRAY)
cf.alignment = align("center", "center", wrap=True)
rh(ws3, last_row, 24)

# Save
output_path = r"C:\Users\amolp\Prometheus\mckinsey_top5_jobs.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")
print("[OK] Sheets: McKinsey Top 5 Jobs | Resume Tailoring Guide | McKinsey Facts & Interview Prep")
print("[OK] Job #1: Digital SC Senior Analyst -- SAP IBP + Kinaxis + BY + o9 ALL required")
print("[OK] Job #3: McKinsey Implementation Associate -- $192K official salary")

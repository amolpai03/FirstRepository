from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE     = "FFFFFF"
BLACK     = "000000"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def nf(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink(cell, url, label="View Job"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

def write_title(ws, title, fill_color, text_color=WHITE, cols=11, row=1, height=34):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, color=text_color, size=14)
    cell.alignment = center()
    cell.fill = PatternFill("solid", start_color=fill_color)

def write_header_row(ws, row, headers, fill_color, text_color=WHITE):
    ws.row_dimensions[row].height = 22
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=text_color, size=10)
        cell.fill = PatternFill("solid", start_color=fill_color)
        cell.alignment = center()
        cell.border = border

def tier_badge(cell, tier):
    colors = {
        "PERFECT MATCH": ("C6EFCE", "375623"),
        "STRONG MATCH":  ("FFEB9C", "7F6000"),
        "GOOD MATCH":    ("DDEBF7", "1F4E79"),
    }
    bg, fg = colors.get(tier, ("F2F2F2", "595959"))
    cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(name="Arial", bold=True, size=9, color=fg)
    cell.alignment = center()
    cell.value = tier
    cell.border = border

def edge_badge(cell, edge_type):
    """Color the network edge badge"""
    colors = {
        "NETWORK":    ("D4EDDA", "155724"),
        "EASY APPLY": ("FFF3CD", "856404"),
        "EARLY BIRD": ("CCE5FF", "004085"),
        "ALUMNI":     ("E2D9F3", "4B0082"),
    }
    bg, fg = colors.get(edge_type, ("F2F2F2", "595959"))
    cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(name="Arial", bold=True, size=9, color=fg)
    cell.alignment = center()
    cell.value = edge_type
    cell.border = border

# ══════════════════════════════════════════════════════════════════════════════
# ALL BEST-CHANCE JOBS
# Columns: title, company, location, type, salary, tier, edge_type, url, why_you_win
# ══════════════════════════════════════════════════════════════════════════════

# (title, company, location, type, salary, tier, edge_type, url, why_youll_get_called)
BEST_JOBS = [
    # ─── NETWORK + PERFECT SKILL MATCH ────────────────────────────────────────
    (
        "Lead o9 Consultant",
        "Infosys",
        "Richardson, TX",
        "On-site · FT",
        "Not listed",
        "PERFECT MATCH",
        "NETWORK",
        "https://www.linkedin.com/jobs/view/4378154088/",
        "3 direct connections · o9 in title = exact skill match · Easy Apply · Actively reviewing — apply TODAY"
    ),
    (
        "Senior o9 Consultant",
        "Infosys",
        "Richardson, TX",
        "On-site · FT",
        "Not listed",
        "PERFECT MATCH",
        "NETWORK",
        "https://www.linkedin.com/jobs/view/4375271313/",
        "14 o9 alumni at Infosys · o9 in title = exact skill match · Easy Apply · Actively reviewing"
    ),
    (
        "Principal O9 Consultant",
        "Infosys",
        "Raleigh, NC",
        "On-site · FT",
        "Not listed",
        "PERFECT MATCH",
        "ALUMNI",
        "https://www.linkedin.com/jobs/view/4388512492/",
        "14 o9 alumni at Infosys · o9 in title · Easy Apply · Principal = 7+ yr exp match"
    ),
    (
        "Senior Demand Planner (Blue Yonder)",
        "Intel",
        "Santa Clara, CA",
        "Hybrid · FT",
        "Not listed",
        "STRONG MATCH",
        "NETWORK",
        "https://www.linkedin.com/jobs/view/4394482391/",
        "Rutvi in your network · 381 school alumni · JD explicitly mentions Blue Yonder · Direct connection advantage"
    ),
    (
        "Senior Demand Planner (Blue Yonder)",
        "Intel",
        "Folsom, CA",
        "Hybrid · FT",
        "Not listed",
        "STRONG MATCH",
        "NETWORK",
        "https://www.linkedin.com/jobs/view/4394478440/",
        "Rutvi in your network · 381 school alumni · JD explicitly mentions Blue Yonder · Same role, 2nd location = 2x shot"
    ),
    (
        "Sr Supply & Demand Planner",
        "Kohler Co.",
        "Kohler, WI",
        "On-site · FT",
        "Not listed",
        "STRONG MATCH",
        "NETWORK",
        "https://www.linkedin.com/jobs/view/4366386259/",
        "Taylor Magri (recruiter) is in your network · o9 alumni at Kohler · Message recruiter directly for fast-track"
    ),
    # ─── PERFECT SKILL MATCH — o9/OMP/BY IN TITLE ────────────────────────────
    (
        "O9 Solutions Consultant",
        "InfoVision Inc.",
        "Remote (US)",
        "Remote · FT",
        "Not listed",
        "PERFECT MATCH",
        "EASY APPLY",
        "https://www.linkedin.com/jobs/view/4383145463/",
        "o9 in title · Remote · Easy Apply · Actively reviewing applicants — low competition window"
    ),
    (
        "O9 Supply Chain Consultant (SCM)",
        "InfoVision Inc.",
        "Remote (US)",
        "Remote · FT",
        "Not listed",
        "PERFECT MATCH",
        "EASY APPLY",
        "https://www.linkedin.com/jobs/view/4385958110/",
        "o9 in title · Remote · Easy Apply · Actively reviewing — same hiring team, dual application strategy"
    ),
    (
        "SC Planning Manager (o9/OMP/BY/JDA)",
        "Deloitte",
        "Charlotte, NC + Multiple",
        "Hybrid · FT",
        "Not listed",
        "PERFECT MATCH",
        "ALUMNI",
        "https://apply.deloitte.com/careers/JobDetail/Supply-Chain-Planning-Manager-experience-with-o9-OMP-or-Blue-Yonder-JDA/119515",
        "o9 + OMP + BY all in job title · 28 o9 alumni at Deloitte · 8+ yrs SC = your experience level · Manager level"
    ),
    (
        "SC Planning Sr Consultant (o9/OMP/BY)",
        "Deloitte",
        "Huntsville, AL",
        "Hybrid · FT",
        "$95K-$135K",
        "PERFECT MATCH",
        "ALUMNI",
        "https://www.linkedin.com/jobs/view/4373925993/",
        "o9 + OMP + BY in title · 5+ yrs SC fit · 28 o9 alumni · $95K-$135K + Deloitte bonus = strong comp"
    ),
    (
        "SC Tech - OMP Planning Manager (Location OPEN)",
        "EY",
        "Remote-Flexible (US)",
        "Hybrid · FT",
        "Not listed",
        "PERFECT MATCH",
        "ALUMNI",
        "https://seasonalworks.labor.ny.gov/albany-ny/supply-chain-tech-omp-planning-manager-location-open/4AA121E9ECFB4969AB54DA59E1622059/job/",
        "OMP in job title · Location OPEN = remote option · 30 o9 alumni at EY · S&OP/IBP focus = core skills"
    ),
    (
        "o9 Consulting Manager",
        "NTT DATA North America",
        "Georgia (Remote)",
        "Remote · FT",
        "Not listed",
        "PERFECT MATCH",
        "EARLY BIRD",
        "https://www.linkedin.com/jobs/view/4390662733/",
        "o9 IN job title · Remote · 13 alumni · Be an early applicant · Manager level = 7+ yr exp match"
    ),
    (
        "O9 Technical Consultant",
        "Xaxis Solutions",
        "Berkeley Heights, NJ",
        "On-site · FT",
        "Not listed",
        "STRONG MATCH",
        "EASY APPLY",
        "https://www.linkedin.com/jobs/view/4398494386/",
        "o9 in title · Easy Apply · Review time ~1 week · Small boutique firm = faster hiring process"
    ),
    # ─── BIG SALARY + EXACT TOOL MATCH ───────────────────────────────────────
    (
        "SAP SC IBP Solution Architect Sr. Manager",
        "PwC",
        "New York, NY + 11 locations",
        "Hybrid · FT",
        "$130K-$256K",
        "PERFECT MATCH",
        "ALUMNI",
        "https://jobs.us.pwc.com/job/new-york/sap-supply-chain-ibp-solution-architect-sr-manager/932/67751281280",
        "IBP Solution Architect = exact title · $130K-$256K range · 11 US locations = location flexibility · Sr Manager level"
    ),
    (
        "Solutions Architect - Kinaxis",
        "Genpact",
        "Remote (US)",
        "Remote · FT",
        "$150K-$180K",
        "PERFECT MATCH",
        "EASY APPLY",
        "https://www.linkedin.com/jobs/view/4330227649/",
        "Vipul Gupta (Global TA) visible · Easy Apply · Actively reviewing · $150K-$180K · Remote · Solution Architect = exact role"
    ),
    (
        "SC Planning & Exec Sr Manager (o9/Kinaxis/BY/SAP IBP)",
        "Accenture",
        "USA (Multiple)",
        "Hybrid · FT",
        "$140K-$200K",
        "PERFECT MATCH",
        "ALUMNI",
        "https://www.accenture.com/us-en/careers/jobdetails?id=R00174637_en",
        "o9 + Kinaxis + BY + SAP IBP all in JD · $140K-$200K · Location flexible · Sr Manager = 7+ yr exp match"
    ),
    (
        "SC Planning Mgmt Consultant (o9/Kinaxis/SAP IBP/BY)",
        "Accenture",
        "Midwest / Chicago, IL",
        "Hybrid · FT",
        "$110K-$160K",
        "PERFECT MATCH",
        "ALUMNI",
        "https://www.accenture.com/us-en/careers/jobdetails?id=R00209619_en",
        "o9 + Kinaxis + SAP IBP + BY all required · 2+ yrs consulting exp · Management Consultant = achievable level"
    ),
    (
        "SAP IBP/TM/PP Senior Manager",
        "PwC",
        "Atlanta, GA",
        "Hybrid · FT",
        "$130K-$256K",
        "STRONG MATCH",
        "ALUMNI",
        "https://jobs.us.pwc.com/job/atlanta/sap-ibp-tm-pp-senior-manager/932/78996929616",
        "SAP IBP + planning tools · $130K-$256K · Atlanta hub · Sr Manager = 7 yr exp match · Big 4 brand"
    ),
    # ─── EASY APPLY + ACTIVELY REVIEWING (LOW COMPETITION) ───────────────────
    (
        "Senior SAP IBP Consultant",
        "SELECCION Consulting",
        "East Brunswick, NJ",
        "Hybrid · FT",
        "Not listed",
        "STRONG MATCH",
        "EARLY BIRD",
        "https://www.linkedin.com/jobs/view/4400587273/",
        "0 applicants at time of posting · Easy Apply · Krishna S. recruiter visible · First-mover massive advantage"
    ),
    (
        "Sr Solution Consultant - Kinaxis",
        "Pacer Group",
        "Remote (US)",
        "Remote · Contract",
        "$82-$92/hr (~$190K ann.)",
        "STRONG MATCH",
        "EASY APPLY",
        "https://www.linkedin.com/jobs/view/4393864119/",
        "$82-$92/hr = ~$170K-$191K annualized · Remote · Easy Apply · Contract = faster hiring, less competition"
    ),
    (
        "Sr Functional Solution Architect",
        "Blue Yonder",
        "Remote (Coppell, TX)",
        "Remote · FT",
        "Not listed",
        "PERFECT MATCH",
        "EARLY BIRD",
        "https://www.linkedin.com/jobs/view/4394225134/",
        "Blue Yonder vendor direct · Solution Architect = exact title · Be an early applicant window open · Remote"
    ),
    (
        "Senior Solution Designer",
        "Korber Supply Chain",
        "Remote (US)",
        "Remote · FT",
        "Not listed",
        "PERFECT MATCH",
        "EARLY BIRD",
        "https://www.linkedin.com/jobs/view/4382197450/",
        "Michael Unser (Director, Solution Design) visible · SC platform consulting · Remote · HIGH MATCH profile fit"
    ),
    # ─── INDUSTRY PLANNING — SCHOOL ALUMNI ADVANTAGE ─────────────────────────
    (
        "Sr Supply Demand Planner",
        "Intuitive",
        "Sunnyvale, CA",
        "On-site · FT",
        "$118.7K-$170.7K",
        "STRONG MATCH",
        "ALUMNI",
        "https://www.linkedin.com/jobs/view/4345274465/",
        "2 o9 alumni work there · $118.7K-$170.7K · Medical robotics SC = unique resume differentiator"
    ),
    (
        "Senior Supply Chain Planner",
        "Micron Technology",
        "Boise, ID",
        "On-site · FT",
        "Not listed",
        "STRONG MATCH",
        "ALUMNI",
        "https://www.linkedin.com/jobs/view/4386872480/",
        "50 school alumni at Micron · Semiconductor SC planning · IBP/S&OP focus · Alumni referral possible"
    ),
    (
        "Sr Analyst, Supply Chain Planning",
        "Analog Devices",
        "Wilmington, MA",
        "Hybrid · FT",
        "Not listed",
        "STRONG MATCH",
        "ALUMNI",
        "https://www.linkedin.com/jobs/view/4371933606/",
        "140 school alumni = highest alumni density on list · IBP/S&OP · Hybrid · Alumni referral very likely"
    ),
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER BEST CHANCES LIST
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Best Chances"

COLS = 10
write_title(
    ws1,
    "YOUR BEST SHOT — 25 Jobs Most Likely to Get You an Interview (April 2026)",
    "1A3C34",
    text_color="FFFFFF",
    cols=COLS
)
ws1.row_dimensions[2].height = 6

HEADERS = ["#", "Job Title", "Company", "Location", "Type", "Salary", "Match Level", "Edge", "Link", "Why You'll Get Called"]
write_header_row(ws1, 3, HEADERS, "2D6A4F")

# Edge type priority grouping (Network > PERFECT MATCH > Easy Apply > Early Bird > Alumni)
for i, job in enumerate(BEST_JOBS):
    r = i + 4
    title, company, location, jtype, salary, tier, edge, url, why = job

    # Row base fill by edge type
    fill_map = {
        "NETWORK":    "F0FFF4",  # soft green — network = highest priority
        "EASY APPLY": "FFFBF0",  # soft yellow
        "EARLY BIRD": "F0F7FF",  # soft blue
        "ALUMNI":     "FAF0FF",  # soft purple
    }
    base_fill = fill_map.get(edge, "FFFFFF")
    row_fill = PatternFill("solid", start_color=base_fill)

    ws1.row_dimensions[r].height = 42

    for c in range(1, COLS + 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border

        if c == 1:
            cell.value = i + 1
            cell.font = nf(bold=True, size=10)
            cell.fill = row_fill
            cell.alignment = center()
        elif c == 2:
            cell.value = title
            cell.font = nf(bold=True, size=9)
            cell.fill = row_fill
            cell.alignment = left()
        elif c == 3:
            cell.value = company
            cell.font = nf(size=9)
            cell.fill = row_fill
            cell.alignment = left()
        elif c == 4:
            cell.value = location
            cell.font = nf(size=9)
            cell.fill = row_fill
            cell.alignment = left()
        elif c == 5:
            cell.value = jtype
            cell.font = nf(size=9)
            cell.fill = row_fill
            cell.alignment = left()
        elif c == 6:
            cell.value = salary
            cell.font = nf(bold=True if "$" in str(salary) else False, size=9, color="155724" if "$" in str(salary) else "595959")
            cell.fill = row_fill
            cell.alignment = left()
        elif c == 7:
            tier_badge(cell, tier)
        elif c == 8:
            edge_badge(cell, edge)
        elif c == 9:
            hyperlink(cell, url)
            cell.fill = row_fill
        elif c == 10:
            cell.value = why
            cell.font = nf(size=9)
            cell.fill = row_fill
            cell.alignment = left()

ws1.freeze_panes = "A4"
COL_WIDTHS = [4, 34, 18, 18, 14, 18, 14, 11, 12, 48]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Legend row
leg_row = len(BEST_JOBS) + 5
ws1.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=COLS)
leg = ws1.cell(row=leg_row, column=1,
    value="COLOR KEY: Green = Network connection  |  Yellow = Easy Apply open  |  Blue = Early bird advantage  |  Purple = Strong alumni presence")
leg.font = Font(name="Arial", italic=True, size=9, color="595959")
leg.alignment = center()
leg.fill = PatternFill("solid", start_color="F9F9F9")
leg.border = border

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — NETWORK LEVERAGE (Use Your Connections!)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Network Leverage")
write_title(ws2, "YOUR NETWORK ADVANTAGE — Apply With Warm Intro or Direct Message", "1F5C2E", cols=8)
ws2.row_dimensions[2].height = 6
write_header_row(ws2, 3, ["#", "Job Title", "Company", "Location", "Salary", "Your Connection", "Link", "Action to Take"], "2D6A4F")

NETWORK_JOBS = [
    ("Lead o9 Consultant",              "Infosys",     "Richardson, TX",    "Not listed",     "3 direct LinkedIn connections",       "https://www.linkedin.com/jobs/view/4378154088/", "Message all 3 connections — ask for referral before applying"),
    ("Senior o9 Consultant",            "Infosys",     "Richardson, TX",    "Not listed",     "14 o9 alumni at Infosys",             "https://www.linkedin.com/jobs/view/4375271313/", "Find o9 alums on LinkedIn · Ask for internal referral"),
    ("Senior Demand Planner (BY)",      "Intel",       "Santa Clara, CA",   "Not listed",     "Rutvi in your network",              "https://www.linkedin.com/jobs/view/4394482391/", "Message Rutvi FIRST. Get intel on team, then apply"),
    ("Senior Demand Planner (BY)",      "Intel",       "Folsom, CA",        "Not listed",     "Rutvi in your network",              "https://www.linkedin.com/jobs/view/4394478440/", "Apply to BOTH Intel locations — same connection helps both"),
    ("Sr Supply & Demand Planner",      "Kohler Co.",  "Kohler, WI",        "Not listed",     "Taylor Magri (Recruiter) in network", "https://www.linkedin.com/jobs/view/4366386259/", "DM Taylor Magri directly on LinkedIn before applying"),
    ("SC Planning Manager (o9/OMP/BY)", "Deloitte",    "Charlotte, NC+",    "Not listed",     "28 o9 alumni at Deloitte",           "https://apply.deloitte.com/careers/JobDetail/Supply-Chain-Planning-Manager-experience-with-o9-OMP-or-Blue-Yonder-JDA/119515", "Find SC practice alumni, get referral code — Deloitte referrals prioritized"),
    ("SC Planning Sr Consultant",       "Deloitte",    "Huntsville, AL",    "$95K-$135K",     "28 o9 alumni at Deloitte",           "https://www.linkedin.com/jobs/view/4373925993/", "Apply + activate any Deloitte alumni connection for referral"),
    ("SAP IBP Solution Architect Sr Mgr","PwC",         "NY + 11 locations", "$130K-$256K",    "Alumni network (check LinkedIn)",    "https://jobs.us.pwc.com/job/new-york/sap-supply-chain-ibp-solution-architect-sr-manager/932/67751281280", "Search PwC SC practice alumni, request informational call"),
    ("Sr Analyst SC Planning",          "Analog Dev.", "Wilmington, MA",    "Not listed",     "140 school alumni at ADI!",          "https://www.linkedin.com/jobs/view/4371933606/", "HIGHEST alumni density — reach out to 2-3 alumni for referral"),
    ("Senior SC Planner",               "Micron",      "Boise, ID",         "Not listed",     "50 school alumni at Micron",         "https://www.linkedin.com/jobs/view/4386872480/", "50 alumni — find SC/planning team members, request referral"),
]

for i, job in enumerate(NETWORK_JOBS):
    r = i + 4
    title, company, location, salary, connection, url, action = job
    fill = PatternFill("solid", start_color="E8F5E9" if i % 2 == 0 else "F1F8F2")
    ws2.row_dimensions[r].height = 44

    for c in range(1, 9):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = i + 1; cell.font = nf(bold=True); cell.alignment = center()
        elif c == 2:
            cell.value = title; cell.font = nf(bold=True, size=9); cell.alignment = left()
        elif c == 3:
            cell.value = company; cell.font = nf(size=9); cell.alignment = left()
        elif c == 4:
            cell.value = location; cell.font = nf(size=9); cell.alignment = left()
        elif c == 5:
            cell.value = salary; cell.font = nf(size=9, color="155724" if "$" in salary else "595959"); cell.alignment = left()
        elif c == 6:
            cell.value = connection; cell.font = nf(bold=True, size=9, color="1F5C2E"); cell.alignment = left()
        elif c == 7:
            hyperlink(cell, url); cell.fill = fill
        elif c == 8:
            cell.value = action; cell.font = nf(size=9); cell.alignment = left()

ws2.freeze_panes = "A4"
for i, w in enumerate([4, 30, 16, 16, 14, 36, 12, 46], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — PERFECT SKILL MATCH (Apply First!)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Perfect Skill Match")
write_title(ws3, "PERFECT SKILL MATCH — o9 / OMP / Blue Yonder / IBP Directly in Job Title", "B8860B", text_color="FFFFFF", cols=8)
ws3.row_dimensions[2].height = 6
write_header_row(ws3, 3, ["#", "Job Title", "Company", "Location", "Type", "Salary", "Link", "Why Perfect"], "7F6000", text_color="FFFFFF")

PERFECT_JOBS = [
    ("Lead o9 Consultant",                     "Infosys",             "Richardson, TX",        "On-site · FT",   "Not listed",          "https://www.linkedin.com/jobs/view/4378154088/",     "o9 IS the job · 7+ yrs = Principal/Lead level fit · 3 connections"),
    ("Senior o9 Consultant",                   "Infosys",             "Richardson, TX",        "On-site · FT",   "Not listed",          "https://www.linkedin.com/jobs/view/4375271313/",     "o9 IS the job · Senior = your experience band"),
    ("Principal O9 Consultant",                "Infosys",             "Raleigh, NC",           "On-site · FT",   "Not listed",          "https://www.linkedin.com/jobs/view/4388512492/",     "o9 IS the job · Principal level = 7+ yrs · 14 alumni"),
    ("O9 Solutions Consultant",                "InfoVision Inc.",     "Remote (US)",           "Remote · FT",    "Not listed",          "https://www.linkedin.com/jobs/view/4383145463/",     "o9 + Solutions Consultant = your exact background"),
    ("O9 Supply Chain Consultant (SCM)",       "InfoVision Inc.",     "Remote (US)",           "Remote · FT",    "Not listed",          "https://www.linkedin.com/jobs/view/4385958110/",     "o9 IS the job · Remote · Same team as above listing"),
    ("SC Planning Manager (o9/OMP/BY/JDA)",    "Deloitte",            "Charlotte, NC+",        "Hybrid · FT",    "Not listed",          "https://apply.deloitte.com/careers/JobDetail/Supply-Chain-Planning-Manager-experience-with-o9-OMP-or-Blue-Yonder-JDA/119515", "o9 + OMP + BY ALL in title = designed for your resume"),
    ("SC Planning Sr Consultant (o9/OMP/BY)",  "Deloitte",            "Huntsville, AL",        "Hybrid · FT",    "$95K-$135K",          "https://www.linkedin.com/jobs/view/4373925993/",     "o9 + OMP + BY in title · $95K-$135K base + bonus"),
    ("SC Tech - OMP Planning Mgr",             "EY",                  "Remote-Flexible (US)",  "Hybrid · FT",    "Not listed",          "https://seasonalworks.labor.ny.gov/albany-ny/supply-chain-tech-omp-planning-manager-location-open/4AA121E9ECFB4969AB54DA59E1622059/job/", "OMP in title · Remote OK · Manager level"),
    ("o9 Consulting Manager",                  "NTT DATA",            "Georgia (Remote)",      "Remote · FT",    "Not listed",          "https://www.linkedin.com/jobs/view/4390662733/",     "o9 in title · Consulting Manager = your seniority · Remote"),
    ("SAP SC IBP Solution Architect Sr Mgr",   "PwC",                 "NY + 11 US locations",  "Hybrid · FT",    "$130K-$256K",         "https://jobs.us.pwc.com/job/new-york/sap-supply-chain-ibp-solution-architect-sr-manager/932/67751281280", "IBP Solution Architect Sr Mgr · $130K-$256K · 11 locations"),
    ("SC Planning & Exec Sr Mgr (o9/Kinaxis/BY/IBP)", "Accenture",   "USA (Multiple)",        "Hybrid · FT",    "$140K-$200K",         "https://www.accenture.com/us-en/careers/jobdetails?id=R00174637_en", "All tools in 1 JD · $140K-$200K · Location flex"),
    ("SC Planning Mgmt Consultant (o9/Kinaxis/IBP/BY)", "Accenture", "Chicago, IL",           "Hybrid · FT",    "$110K-$160K",         "https://www.accenture.com/us-en/careers/jobdetails?id=R00209619_en", "All tools in 1 JD · $110K-$160K · Mgmt Consultant"),
    ("Sr Functional Solution Architect",       "Blue Yonder",         "Remote (Coppell, TX)",  "Remote · FT",    "Not listed",          "https://www.linkedin.com/jobs/view/4394225134/",     "BY vendor direct · Solution Architect = exact title match"),
    ("Solutions Architect - Kinaxis",          "Genpact",             "Remote (US)",           "Remote · FT",    "$150K-$180K",         "https://www.linkedin.com/jobs/view/4330227649/",     "SA title · $150K-$180K · Recruiter visible · Easy Apply"),
    ("Senior Solution Designer",               "Korber SC",           "Remote (US)",           "Remote · FT",    "Not listed",          "https://www.linkedin.com/jobs/view/4382197450/",     "Director contact visible · SC platform consulting · Remote"),
]

for i, job in enumerate(PERFECT_JOBS):
    r = i + 4
    title, company, location, jtype, salary, url, why = job
    fill = PatternFill("solid", start_color="FFF9E6" if i % 2 == 0 else "FFFDE7")
    ws3.row_dimensions[r].height = 44

    for c in range(1, 9):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = i + 1; cell.font = nf(bold=True); cell.alignment = center()
        elif c == 2:
            cell.value = title; cell.font = nf(bold=True, size=9, color="7F4000"); cell.alignment = left()
        elif c == 3:
            cell.value = company; cell.font = nf(size=9); cell.alignment = left()
        elif c == 4:
            cell.value = location; cell.font = nf(size=9); cell.alignment = left()
        elif c == 5:
            cell.value = jtype; cell.font = nf(size=9); cell.alignment = left()
        elif c == 6:
            cell.value = salary; cell.font = nf(bold=True if "$" in salary else False, size=9, color="155724" if "$" in salary else "595959"); cell.alignment = left()
        elif c == 7:
            hyperlink(cell, url); cell.fill = fill
        elif c == 8:
            cell.value = why; cell.font = nf(size=9); cell.alignment = left()

ws3.freeze_panes = "A4"
for i, w in enumerate([4, 36, 18, 18, 14, 14, 12, 44], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — APPLY NOW (Easy Apply + 0-Low Applicant Count)
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Apply NOW")
write_title(ws4, "APPLY NOW — Easy Apply + Low Applicant Count + Actively Reviewing", "C00000", cols=8)
ws4.row_dimensions[2].height = 6
write_header_row(ws4, 3, ["#", "Job Title", "Company", "Location", "Salary / Rate", "Competition Level", "Link", "Apply Strategy"], "C00000")

APPLY_NOW = [
    ("Senior SAP IBP Consultant",              "SELECCION Consulting", "East Brunswick, NJ", "Not listed",          "0 APPLICANTS at posting",     "https://www.linkedin.com/jobs/view/4400587273/", "APPLY IMMEDIATELY · Krishna S. recruiter visible · Easy Apply · First mover wins"),
    ("O9 Solutions Consultant",                "InfoVision Inc.",      "Remote (US)",        "Not listed",          "Low — actively reviewing",    "https://www.linkedin.com/jobs/view/4383145463/", "Easy Apply · o9 exact match · Remote · Apply + follow up in 3 days"),
    ("O9 Supply Chain Consultant (SCM)",       "InfoVision Inc.",      "Remote (US)",        "Not listed",          "Low — actively reviewing",    "https://www.linkedin.com/jobs/view/4385958110/", "Easy Apply · Same team as above · Apply to both for double exposure"),
    ("Sr Solution Consultant - Kinaxis",       "Pacer Group",          "Remote (US)",        "$82-$92/hr (~$190K)", "45 applicants (moderate)",    "https://www.linkedin.com/jobs/view/4393864119/", "Easy Apply · Contract = faster hire · $190K ann · Apply today"),
    ("Solutions Architect - Kinaxis",          "Genpact",              "Remote (US)",        "$150K-$180K",         "Moderate — still reviewing",  "https://www.linkedin.com/jobs/view/4330227649/", "Easy Apply · Vipul Gupta recruiter visible · DM recruiter after applying"),
    ("Blue Yonder Project Manager",            "Maven Workforce Inc.", "California (Remote)","Not listed",          "Low — 1 hr old at posting",   "https://www.linkedin.com/jobs/view/4400574301/", "Easy Apply · Low competition · BY skills transfer well"),
    ("O9 Technical Consultant",                "Xaxis Solutions",      "Berkeley Heights, NJ","Not listed",         "Low — boutique firm",         "https://www.linkedin.com/jobs/view/4398494386/", "Easy Apply · Small firm = personal review · ~1 week turnaround"),
    ("Platform Architect - Kinaxis",           "Kinaxis (Direct)",     "Remote (US)",        "Not listed",          "Low — posted 46 mins at search","https://www.linkedin.com/jobs/view/4398156298/", "Kinaxis direct · Early applicant window · Platform Architect = SC tech blend"),
    ("Sr Functional Solution Architect",       "Blue Yonder",          "Remote (Coppell TX)","Not listed",          "Early applicant window",      "https://www.linkedin.com/jobs/view/4394225134/", "BY vendor direct · Early bird · Solution Architect = exact match"),
    ("SAP IBP Consultant",                     "via Dice",             "Remote (US)",        "Not listed",          "Low — remote contract",       "https://www.linkedin.com/jobs/view/4398118825/", "Easy Apply · Remote · Contract = fast hire track"),
]

for i, job in enumerate(APPLY_NOW):
    r = i + 4
    title, company, location, salary, competition, url, strategy = job
    fill = PatternFill("solid", start_color="FFF0F0" if i % 2 == 0 else "FFF8F8")
    ws4.row_dimensions[r].height = 44

    for c in range(1, 9):
        cell = ws4.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = i + 1; cell.font = nf(bold=True); cell.alignment = center()
        elif c == 2:
            cell.value = title; cell.font = nf(bold=True, size=9, color="7B0000"); cell.alignment = left()
        elif c == 3:
            cell.value = company; cell.font = nf(size=9); cell.alignment = left()
        elif c == 4:
            cell.value = location; cell.font = nf(size=9); cell.alignment = left()
        elif c == 5:
            cell.value = salary; cell.font = nf(bold=True if "$" in salary else False, size=9, color="155724" if "$" in salary else "595959"); cell.alignment = left()
        elif c == 6:
            cell.value = competition
            if "0 APPLICANTS" in competition:
                cell.font = Font(name="Arial", bold=True, size=9, color="C00000")
                cell.fill = PatternFill("solid", start_color="FFE0E0")
            else:
                cell.font = nf(size=9, color="595959")
            cell.alignment = left()
        elif c == 7:
            hyperlink(cell, url); cell.fill = fill
        elif c == 8:
            cell.value = strategy; cell.font = nf(size=9); cell.alignment = left()

ws4.freeze_panes = "A4"
for i, w in enumerate([4, 34, 18, 16, 20, 24, 12, 46], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
output_path = "C:/Users/amolp/Prometheus/job_listings_best_chances.xlsx"
wb.save(output_path)
print(f"Done! job_listings_best_chances.xlsx saved with 4 sheets and 25 curated jobs.")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bot = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="404040"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink(cell, url, label="View Job →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Mastercard brand palette
MC_RED       = "EB001B"   # Mastercard red
MC_ORANGE    = "F79E1B"   # Mastercard orange/gold
MC_DARK      = "1A1A2E"   # Very dark navy
MC_LIGHT_RED = "FFF0F0"   # Light red tint
MC_LIGHT_ORG = "FFF8EE"   # Light orange tint
MC_LIGHT_NVY = "EEF3FF"   # Light navy tint
MC_MID       = "FFE5CC"   # Mid orange

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTERCARD TOP 3 JOBS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Mastercard Top 3 Jobs"

# Title banner
ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "MASTERCARD — Top 3 Matching Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=MC_DARK)
t.border = border

# Red-orange gradient spacer
ws1.row_dimensions[2].height = 5
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=MC_RED)

# Sub-banner
ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "7 o9 alumni at Mastercard  ·  23 school alumni  ·  Mastercard Advisors = world-class consulting division  ·  Company ID: 3015"
s.font = Font(name="Arial", italic=True, size=9, color=MC_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color="FFF5F5")
s.border = border

# Column headers
HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match",
           "Applicants", "Posted", "Key Skills Required", "Link", "Network Edge", "Why Apply"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=MC_RED)
    cell.alignment = center()
    cell.border = thick_bot

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Consultant, Advisors & Consulting Services,\nStrategy & Transformation",
        "location": "Chicago, IL\n(also Arlington VA / Boston MA\n/ Purchase NY / San Francisco CA)",
        "type": "Hybrid · Full-time",
        "salary": "$97,000 – $139,000\n+ benefits + bonus",
        "match": "STRONG MATCH",
        "match_bg": "C6EFCE", "match_fg": "375623",
        "applicants": "100+ clicked apply",
        "posted": "1 week ago",
        "skills": "Strategy consulting · Data & analytics\nClient/stakeholder management\nProject management · PowerPoint/Excel\nBusiness intelligence · Logical thinking",
        "url": "https://www.linkedin.com/jobs/view/4395163669/",
        "network": "7 o9 alumni at MC\n23 school alumni",
        "why": "Mastercard Advisors = elite consulting arm. Your SC planning consulting background directly maps to Strategy & Transformation work. 5 US locations = flexibility. $97K-$139K base + Mastercard bonus = strong package.",
        "fill": MC_LIGHT_RED,
    },
    {
        "num": 2,
        "title": "Director, Supply Chain\nSustainability Strategy",
        "location": "Purchase, NY\n(also Chicago, IL)\nHybrid",
        "type": "Hybrid · Full-time",
        "salary": "$163,000 – $269,000\n(Purchase NY)\n$142,000 – $234,000 (Chicago)",
        "match": "GOOD MATCH",
        "match_bg": "FFEB9C", "match_fg": "7F6000",
        "applicants": "Fresh — 1 day ago!\n(Low applicants)",
        "posted": "1 day ago",
        "skills": "Supply chain strategy · Scope 3 / GHG\nChange management · Procurement\nSustainability programs · Stakeholder mgmt\nCarbon visibility & emissions accounting",
        "url": "https://www.linkedin.com/jobs/view/4375911195/",
        "network": "7 o9 alumni at MC\n23 school alumni",
        "why": "Supply Chain in title + Director level = strong seniority match. Posted YESTERDAY = minimal competition. $163K-$269K is the highest salary of the 3 roles. Reports to VP Environmental Sustainability. Leverage SC planning expertise, highlight any ESG/sustainability exposure.",
        "fill": MC_LIGHT_ORG,
    },
    {
        "num": 3,
        "title": "Manager, Services\nProgram Enablement",
        "location": "Purchase, NY\nHybrid",
        "type": "Hybrid · Full-time",
        "salary": "$130,000 – $254,000\n+ benefits + bonus",
        "match": "GOOD MATCH",
        "match_bg": "DDEBF7", "match_fg": "1F4E79",
        "applicants": "1 week ago\n(Moderate competition)",
        "posted": "1 week ago",
        "skills": "Program management · Services operations\nEnablement strategy · Cross-functional leadership\nData-driven decision making\nStakeholder engagement",
        "url": "https://www.linkedin.com/jobs/view/4397161628/",
        "network": "23 school alumni\nat Mastercard",
        "why": "Excellent salary range $130K-$254K. Program Enablement = driving services capability across Mastercard's payments ecosystem. SC planning consulting = excellent fit for program leadership. 23 school alumni = strong referral opportunity.",
        "fill": MC_LIGHT_NVY,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 100
    base_fill = PatternFill("solid", start_color=job["fill"])

    vals = [job["num"], job["title"], job["location"], job["type"], job["salary"],
            job["match"], job["applicants"], job["posted"], job["skills"],
            job["url"], job["network"], job["why"]]

    for c, val in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = base_fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=MC_RED)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=MC_DARK)
            cell.alignment = left()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_fresh = "1 day" in str(val) or "0 applicants" in str(val).lower()
            cell.font = Font(name="Arial", bold=is_fresh, size=9,
                           color=MC_RED if is_fresh else "595959")
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = base_fill
        elif c == 5:
            cell.value = val
            cell.font = nf(bold=True, size=9, color="155724")
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

COL_WIDTHS = [4, 30, 22, 14, 22, 14, 18, 14, 34, 12, 20, 44]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RESUME TAILORING GUIDE
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING — What to Highlight for Each Mastercard Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=MC_DARK)
t2.border = border

ws2.row_dimensions[2].height = 5
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=MC_RED)

HEADERS2 = ["Role", "Highlight in Resume", "Keywords to Include",
            "Lead With This Experience", "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=MC_RED)
    cell.alignment = center()
    cell.border = thick_bot

ws2.freeze_panes = "A4"

RESUME_ROWS = [
    (
        "Consultant, Advisors &\nConsulting Services,\nStrategy & Transformation",
        "• SC planning consulting project experience\n• Client-facing advisory & presentations\n• Data-driven strategy recommendations\n• Cross-industry problem solving\n• Tools: Excel, PowerPoint, data analysis",
        "Strategy, transformation, consulting, client management, data analytics, business intelligence, stakeholder engagement, structured thinking, problem solving",
        "Lead with SC planning consulting projects — show how you drove client strategy using data/analytics. Quantify: # clients served, $ impact delivered, project scale",
        "Role is generalist consulting — not SC planning tech specific. 100+ applicants = high competition",
        "Frame your o9/BY/Kinaxis work as 'technology-enabled strategy consulting'. Mastercard Advisors values cross-industry consulting depth. 5 locations = apply to ALL for more chances",
        "APPLY — GOOD FIT\n$97K-$139K"
    ),
    (
        "Director, Supply Chain\nSustainability Strategy",
        "• Any sustainability/ESG project work\n• Supply chain strategy ownership\n• Change management initiatives\n• Scope 3 / emissions knowledge\n• Procurement & sourcing exposure",
        "Supply chain, sustainability, decarbonization, Scope 3, GHG, carbon, procurement, change management, stakeholder engagement, ESG, environmental strategy",
        "Lead with supply chain transformation projects. Highlight any sustainability angle from your projects (e.g., inventory optimization = waste reduction). Director = 8-10+ yr experience",
        "Requires deep ESG/Scope 3 expertise. Sustainability focus is a stretch from planning tech",
        "Only 1 day old — apply IMMEDIATELY before competition builds. $163K-$269K is exceptional. Frame SC planning as 'supply chain optimization reducing emissions'. Any ESG-adjacent work is gold",
        "APPLY NOW\n(Fresh posting!)\n$163K-$269K"
    ),
    (
        "Manager, Services\nProgram Enablement",
        "• Program management experience\n• Cross-functional leadership\n• Enabling teams/services at scale\n• Data-driven planning & reporting\n• Stakeholder management across org",
        "Program management, enablement, services operations, cross-functional, stakeholder management, planning, execution, KPIs, program governance, leadership",
        "Lead with any program management or project governance work. Show scope of programs managed: # stakeholders, budget managed, outcomes delivered. SC planning implementations = program management experience",
        "Mastercard-specific services context not well-known. Role is payments/financial services focused vs supply chain",
        "Frame SC planning project implementations as large-scale program management. 23 school alumni = strong referral pool. $130K-$254K wide range = room for negotiation",
        "GOOD FIT\n$130K-$254K"
    ),
]

ROW_FILLS   = [MC_LIGHT_RED, MC_LIGHT_ORG, MC_LIGHT_NVY]
PRI_COLORS  = [("C6EFCE","375623"), ("FFEB9C","7F6000"), ("DDEBF7","1F4E79")]

for i, row_data in enumerate(RESUME_ROWS):
    r = i + 4
    ws2.row_dimensions[r].height = 120
    fill = PatternFill("solid", start_color=ROW_FILLS[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=MC_DARK)
            cell.fill = fill
            cell.alignment = left()
        elif c == 7:
            bg, fg = PRI_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

for i, w in enumerate([26, 40, 40, 42, 30, 42, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — MASTERCARD QUICK FACTS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Mastercard Quick Facts")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "MASTERCARD — Company & Advisors Division Quick Reference"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=MC_DARK)
t3.border = border

ws3.row_dimensions[2].height = 5
ws3.merge_cells("A2:D2")
ws3["A2"].fill = PatternFill("solid", start_color=MC_RED)

facts = [
    ("Company",         "Mastercard Incorporated — Global payments technology company"),
    ("HQ",              "Purchase, New York (also major offices Chicago, NYC, Arlington VA, SF)"),
    ("Employees",       "~33,000 globally · Fortune 500 · NYSE: MA"),
    ("Revenue",         "$25.1B (2024) — consistently growing ~10-15% YoY"),
    ("o9 Alumni",       "7 o9 Solutions alumni work at Mastercard — strong internal SC network"),
    ("School Alumni",   "23 school alumni at Mastercard — high referral probability"),
    ("Advisors Div.",   "Mastercard Advisors = 1,000+ consultants · Combines payments data + strategy consulting"),
    ("Advisors Focus",  "Strategy & Transformation · Performance Analytics · Business Experimentation · Marketing · Program Mgmt"),
    ("Why MC?",         "Global brand · Payments data advantage · SC + analytics consulting = rare combo · Strong comp"),
    ("Interview Tip",   "Mastercard loves 'data-driven decision making' — quantify EVERYTHING in your resume & interviews"),
    ("Salary Ranges",   "Consultant: $97K-$139K | Director SC: $163K-$269K | Manager Enablement: $130K-$254K"),
    ("Benefits",        "401k match · 16 wks parental leave · 25 days PTO + 5 personal · Tuition reimbursement · Health/dental/vision"),
    ("LinkedIn ID",     "3015 · linkedin.com/company/mastercard"),
    ("Careers Page",    "careers.mastercard.com"),
]

ws3.row_dimensions[3].height = 22
for c, h in enumerate(["Category", "Details", "", ""], 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=MC_RED)
    cell.alignment = center()
    cell.border = thick_bot

for i, (cat, detail) in enumerate(facts):
    r = i + 4
    ws3.row_dimensions[r].height = 28
    alt_fill = PatternFill("solid", start_color="FFF5F5" if i % 2 == 0 else WHITE)

    c1 = ws3.cell(row=r, column=1, value=cat)
    c1.font = nf(bold=True, size=9, color=MC_DARK)
    c1.fill = alt_fill
    c1.alignment = left()
    c1.border = border

    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c2 = ws3.cell(row=r, column=2, value=detail)
    c2.font = nf(size=9)
    c2.fill = alt_fill
    c2.alignment = left()
    c2.border = border

ws3.freeze_panes = "A4"
for i, w in enumerate([20, 80, 1, 1], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
output = "C:/Users/amolp/Prometheus/mastercard_top3_jobs.xlsx"
wb.save(output)
print("Done! mastercard_top3_jobs.xlsx saved with 3 sheets.")

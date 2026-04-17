from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="595959"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def hyperlink(cell, url, label="View Job ->"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Intel brand colors
INTEL_BLUE     = "0071C5"
INTEL_DARK     = "003C71"
INTEL_LIGHT    = "E8F4FD"
INTEL_TEAL     = "00AEEF"
INTEL_GRAY     = "F5F5F5"
INTEL_ELECTRIC = "00C7FD"

GREEN_BADGE  = "C6EFCE"
GREEN_DARK   = "375623"
YELLOW_BADGE = "FFEB9C"
YELLOW_DARK  = "7F6000"
BLUE_BADGE   = "DDEBF7"
BLUE_DARK    = "1F4E79"
ORANGE_BADGE = "FCE4D6"
ORANGE_DARK  = "833C00"

# =============================================================================
# SHEET 1 - INTEL TOP 5 JOBS
# =============================================================================
ws1 = wb.active
ws1.title = "Intel Top 5 Jobs"

ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "INTEL CORPORATION -- Top 5 Supply Chain & Planning Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=INTEL_DARK)
t.border = border

ws1.row_dimensions[2].height = 7
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=INTEL_BLUE)

ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "7 o9 alumni at Intel  |  10 direct connections for Job 4  |  64 total SC/Planning roles found  |  Intel LinkedIn Company ID: 1053  |  NOTE: Jobs 3 & 4 are Temp Contracts"
s.font = Font(name="Arial", italic=True, size=9, color=INTEL_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color=INTEL_LIGHT)
s.border = border

HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match Level",
           "Status / Alumni", "Posted", "Key Skills Required", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=INTEL_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Director\nFoundry Customer Planning",
        "location": "Santa Clara, CA\n(Hybrid)",
        "type": "Full-Time\nHybrid\nDIRECTOR LEVEL",
        "salary": "$220,320 - $311,040\n(Base only)\nDirector-level comp",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "381 school alumni\nat Intel\n3 weeks ago",
        "posted": "3 weeks ago",
        "skills": "End-to-end demand planning\nCapacity management & supply assurance\nForecast accuracy ownership\nPricing strategy coordination\nCross-fab execution (assembly, test, suppliers)\n12+ yrs experience (BS) required",
        "url": "https://www.linkedin.com/jobs/view/4386119770/",
        "network": "381 NCSU alumni\nat Intel\n(large network base)",
        "strategy": "MOST STRATEGIC ROLE -- Director-level S&OP/demand/capacity planning leadership ($220K-$311K). Owns Intel Foundry's entire customer delivery planning including forecast accuracy, supply assurance, and cycle-time optimization. Your SC consulting breadth = cross-industry planning expertise. Requires 12+ years -- frame 7 yrs consulting as 15+ yrs industry equivalent impact.",
        "fill": INTEL_LIGHT,
    },
    {
        "num": 2,
        "title": "Substrates Strategic\nCapacity Planner",
        "location": "Hillsboro, OR\n(Hybrid)",
        "type": "Full-Time\nHybrid",
        "salary": "$128,080 - $254,320\n(Wide band reflects\nlevel flexibility)",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "3 days ago!\nBe an Early Applicant\n7 o9 alumni at Intel",
        "posted": "3 days ago",
        "skills": "Strategic capacity planning\nSemiconductor substrate SC knowledge\nRoadmap development (short/long-term)\nSupplier ecosystem management\nDemand-supply balancing\n3+ yrs semiconductor SC preferred",
        "url": "https://www.linkedin.com/jobs/view/4369847841/",
        "network": "7 o9 Solutions\nalumni at Intel\n= referral pathway",
        "strategy": "FRESH POSTING (3 days old!) + o9 alumni network = strong position. Capacity planning for substrate supply chain -- directly maps to demand/supply planning expertise. Salary range is wide ($128K-$254K) suggesting level flexibility. Apply immediately + message an o9 alum at Intel for referral before posting fills.",
        "fill": "EBF5FB",
    },
    {
        "num": 3,
        "title": "Supply Planning\nBusiness Architect\n(Temporary Contract)",
        "location": "Folsom, CA\n(Hybrid)\n[Also: Phoenix, AZ]",
        "type": "TEMP CONTRACT\nHybrid\n(Contract-to-hire\npotential)",
        "salary": "~$90,000 - $160,000\n(estimated contract rate)\nFull range: $52K-$200K",
        "match": "PERFECT MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "1 week ago\nBe an Early Applicant\n7 o9 alumni at Intel",
        "posted": "1 week ago",
        "skills": "SC planning process architecture\n5+ yrs SC planning environment\nPower BI / Tableau / data viz\nBusiness process design\nAPS/ERP system knowledge\nIntel Foundry supply chain context",
        "url": "https://www.linkedin.com/jobs/view/4394475419/",
        "network": "7 o9 alumni\nat Intel\nFolsom CA location",
        "strategy": "YOUR TITLE IN THE JOB TITLE -- 'Business Architect' = exactly your Solutions Consultant/Architect background. SC planning environment + business process design = core consulting skill. NOTE: Temp contract -- treat as a foot-in-the-door at Intel. Many Intel contracts convert to FTE. 7 o9 alumni for referral. Apply to BOTH Folsom CA + Phoenix AZ listings.",
        "fill": "FDFEFE",
    },
    {
        "num": 4,
        "title": "Supply Planning Integrator\n(Temporary Contract)",
        "location": "Folsom, CA\n(Hybrid)\n[Also: Hillsboro, OR]",
        "type": "TEMP CONTRACT\nHybrid\n(Contract-to-hire\npotential)",
        "salary": "~$80,000 - $120,000\n(estimated contract rate)\nGlassdoor est: $67K-$94K",
        "match": "HIGH MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "10 DIRECT CONNECTIONS\n1 week ago\nBe an Early Applicant",
        "posted": "1 week ago",
        "skills": "SC planning integration\nNew product/customer onboarding\nIntel Foundry Supply Chain Enablement\nCross-functional coordination\nPlanning system processes\nIBP / APS / ERP familiarity",
        "url": "https://www.linkedin.com/jobs/view/4394488236/",
        "network": "10 DIRECT\nCONNECTIONS!\nFolsom CA location",
        "strategy": "10 DIRECT CONNECTIONS = biggest network edge in this search! Integrating new products/customers into Intel Foundry's supply chain planning = APS implementation experience is DIRECTLY relevant. NOTE: Temp contract -- but 10 connections gives massive referral advantage. Message connections NOW, apply in Folsom and Hillsboro. Contract = fast path into Intel SC team.",
        "fill": "FEFDF0",
    },
    {
        "num": 5,
        "title": "Supply Chain Engineer\nFab Equipment Sourcing",
        "location": "Phoenix, AZ\n(On-site)",
        "type": "Full-Time\nOn-site",
        "salary": "$120,860 - $231,670\n(Base only)\nFull-time perm role",
        "match": "MEDIUM MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "6 days ago\nBe an Early Applicant\n7 o9 alumni at Intel",
        "posted": "6 days ago",
        "skills": "Supply chain engineering\nFab equipment supplier management\nSC performance KPIs & controls\nSupplier process validation\nReliability / yield / cost optimization\nEngineering or SC degree preferred",
        "url": "https://www.linkedin.com/jobs/view/4385974365/",
        "network": "7 o9 alumni\nat Intel\nPhoenix AZ cluster",
        "strategy": "FULL-TIME role (vs. temp contracts) + early applicant + o9 alumni network. SC engineering focus on fab equipment sourcing -- requires semiconductor supplier management depth. Apply if you have manufacturing/engineering client SC experience. Phoenix AZ is Intel's largest US campus (growing rapidly). Lower priority vs. Jobs 1-3 but only full-time option with o9 alumni edge.",
        "fill": INTEL_GRAY,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 100
    fill = PatternFill("solid", start_color=job["fill"])

    data = [
        job["num"], job["title"], job["location"], job["type"],
        job["salary"], job["match"], job["status"], job["posted"],
        job["skills"], job["url"], job["network"], job["strategy"]
    ]

    for c, val in enumerate(data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=INTEL_BLUE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=INTEL_DARK)
            cell.alignment = left()
        elif c == 4:  # Type - flag contracts
            cell.value = val
            is_contract = "TEMP" in str(val)
            cell.font = Font(name="Arial", bold=is_contract, size=9,
                           color="C00000" if is_contract else "595959")
            cell.fill = PatternFill("solid", start_color=ORANGE_BADGE if is_contract else job["fill"])
            cell.alignment = center()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_hot = "10 DIRECT" in str(val) or "3 days" in str(val)
            cell.font = Font(name="Arial", bold=is_hot, size=9,
                           color="C00000" if "10 DIRECT" in str(val) else ("375623" if "Early" in str(val) else "595959"))
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = fill
        elif c == 12:
            cell.value = val
            cell.font = nf(size=9, color="1A1A1A")
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

COL_WIDTHS = [4, 26, 18, 16, 22, 14, 22, 12, 34, 12, 18, 48]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# SHEET 2 - RESUME TAILORING GUIDE
# =============================================================================
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING -- What to Highlight for Each Intel Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=INTEL_DARK)
t2.border = border

ws2.row_dimensions[2].height = 7
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=INTEL_BLUE)

HEADERS2 = ["Role", "Must Highlight in Resume", "Keywords to Include",
            "Experience to Lead With", "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=INTEL_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Director\nFoundry Customer Planning",
        "- Director-level SC leadership (team + org)\n- End-to-end S&OP cycle ownership\n- Demand planning + capacity management\n- Executive stakeholder management\n- Forecast accuracy KPIs you've owned\n- Cross-fab / multi-site SC coordination",
        "Director, demand planning, S&OP, capacity management, forecast accuracy, supply assurance, customer delivery, foundry, Intel Foundry, IDM 2.0, cycle time, revenue protection, supply chain strategy, pricing",
        "Lead with executive-level SC planning outcomes: $ revenue protected, forecast accuracy %, on-time delivery %, # manufacturing sites coordinated. Frame consulting client engagements as end-to-end SC transformations you led -- not just advised. Show breadth across demand/supply/finance/operations.",
        "12+ years experience required -- 7 yrs consulting may be perceived as short. Director = people manager role.",
        "Frame 7 yrs consulting as accelerated impact: each 1-2 yr client engagement = full SC transformation lifecycle that takes internal PMs 3-5 yrs to complete. Emphasize org leadership: junior consultant mentorship, client team leadership, project team management. Apply anyway -- JD experience ranges are aspirational.",
        "APPLY -- HIGHEST\nSALARY ROLE\n($220K-$311K)"
    ),
    (
        "Substrates Strategic\nCapacity Planner",
        "- Strategic capacity planning experience\n- Demand-supply balancing methodologies\n- Supplier ecosystem management\n- SC roadmap development (short + long term)\n- Planning tools (IBP, Kinaxis, BY) depth\n- Cross-functional coordination skills",
        "Capacity planning, strategic planning, substrate, semiconductor, supply chain, roadmap, demand-supply balance, supplier management, forecast, long-range planning, IBP, S&OP, APS, planning tools",
        "Lead with capacity planning projects: show planning horizon (13-week, 18-month, 5-year), # SKUs/products planned, $ value of supply commitments managed. Highlight any semiconductor or high-tech manufacturing client SC work. Tool depth in IBP/Kinaxis/BY directly addresses requirements.",
        "3+ yrs semiconductor substrate SC experience preferred -- not typical for SC planning consultants.",
        "Position tech/electronics manufacturing client SC projects as semiconductor-adjacent. Substrate = advanced packaging materials SC -- frame any component sourcing or tech manufacturing experience. 3-day-old posting + o9 alumni referral = apply NOW before it's filled.",
        "APPLY NOW\n(3 DAYS OLD\n+ o9 alumni!)"
    ),
    (
        "Supply Planning\nBusiness Architect\n(Contract)",
        "- Business architecture / process design\n- SC planning system implementation\n- APS/IBP/ERP configuration experience\n- Data visualization (Power BI / Tableau)\n- Intel Foundry SC process knowledge\n- 5+ yrs SC planning environment",
        "Business architect, supply planning, process architecture, APS, IBP, S/4HANA, ERP, Power BI, Tableau, D3, data visualization, SC planning, business process design, Intel Foundry, SCE, planning systems",
        "Lead with SC planning architecture projects: ERP/APS implementations, business process design, planning system rollouts. Show Power BI/Tableau dashboard work. 'Business Architect' in the JD = your exact skill set as a Solutions Architect/Consultant. Emphasize you've designed SC planning processes, not just used them.",
        "Temp contract (not full-time perm). Power BI/Tableau data viz experience is listed as a key requirement.",
        "Contract = lower barrier to entry and faster hiring process. Treat as strategic foot-in-the-door at Intel. Brush up Power BI/Tableau skills before interview. Intel contracts frequently convert to FTE. Apply to BOTH Folsom CA and Phoenix AZ listings to maximize chances.",
        "APPLY -- PERFECT\nTITLE MATCH\n(also apply Phoenix)"
    ),
    (
        "Supply Planning Integrator\n(Contract)",
        "- Planning system integration experience\n- New product/customer onboarding in SC\n- APS/ERP planning process knowledge\n- Cross-functional project management\n- Intel Foundry SC process context",
        "Supply planning, integrator, new product introduction, NPI, customer onboarding, SC integration, planning process, Intel Foundry, supply chain enablement, cross-functional, program management, APS, ERP",
        "Lead with SC system integration projects: new customer/product onboarding in APS tools, planning process design for new product families, cross-functional coordination for SC launches. Your APS implementation experience (o9, BY, Kinaxis, IBP) = direct qualification for 'planning integrator' role.",
        "Temp contract. Role is focused on Intel Foundry operations which has specific semiconductor fab SC context.",
        "10 direct connections = huge advantage. Message connections BEFORE applying. Contract = fast entry path. Intel Foundry is growing rapidly (new fab in Ohio, Arizona) -- planning integrator roles critical for IDM 2.0 strategy. Frame APS tool implementations as 'SC planning integration' work.",
        "APPLY NOW\n(10 DIRECT\nCONNECTIONS!)"
    ),
    (
        "Supply Chain Engineer\nFab Equipment Sourcing",
        "- Supply chain engineering experience\n- Supplier management & qualification\n- SC performance metrics / KPIs\n- Engineering background (preferred)\n- Process validation experience",
        "Supply chain engineer, fab equipment, sourcing, supplier management, KPI, performance, process validation, reliability, yield, cost optimization, semiconductor, equipment, supplier qualification",
        "Lead with any manufacturing equipment SC work from consulting clients. Show supplier management outcomes: # suppliers managed, $ savings achieved, quality improvement %. Engineering + SC combination is key here -- if you have any engineering background, feature it prominently.",
        "Requires engineering degree (EE, Chemical, Materials, etc.) and semiconductor fab equipment knowledge -- quite specialized.",
        "Lower priority role unless you have strong engineering + semiconductor supplier background. Full-time perm role is an advantage vs. temp contracts. Phoenix AZ = Intel's fastest-growing campus. If other roles don't pan out, this is a solid full-time fallback with o9 alumni network.",
        "FALLBACK OPTION\n(full-time perm\nvs. contracts)"
    ),
]

ROW_FILLS2 = [INTEL_LIGHT, "EBF5FB", "FDFEFE", "FEFDF0", INTEL_GRAY]
PRIORITY_COLORS = [
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (YELLOW_BADGE, YELLOW_DARK),
    (BLUE_BADGE, BLUE_DARK),
]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 115
    fill = PatternFill("solid", start_color=ROW_FILLS2[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=INTEL_DARK)
            cell.fill = fill
            cell.alignment = left()
        elif c == 7:
            bg, fg = PRIORITY_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

COL_WIDTHS2 = [24, 40, 40, 44, 32, 40, 16]
for i, w in enumerate(COL_WIDTHS2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# SHEET 3 - INTEL QUICK FACTS + IDM 2.0 CONTEXT
# =============================================================================
ws3 = wb.create_sheet("Intel Facts + Interview Prep")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "INTEL CORPORATION QUICK FACTS + Interview Prep -- SC & Foundry Roles"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=INTEL_DARK)
t3.border = border

ws3.row_dimensions[2].height = 7
ws3.merge_cells("A2:H2")
ws3["A2"].fill = PatternFill("solid", start_color=INTEL_BLUE)

# --- Intel Company Facts ---
ws3.row_dimensions[3].height = 22
ws3.merge_cells("A3:H3")
sec1 = ws3["A3"]
sec1.value = "INTEL COMPANY & FOUNDRY FACTS -- Critical Context Before Your Interview"
sec1.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec1.alignment = center()
sec1.fill = PatternFill("solid", start_color=INTEL_BLUE)
sec1.border = border

INTEL_FACTS = [
    ("Company Overview", "Intel Corp: ~$54B revenue (2024). ~100,000 employees. World's largest semiconductor chip designer and manufacturer. Nasdaq: INTC."),
    ("IDM 2.0 Strategy", "Intel's transformation: IDM 2.0 = Integrated Device Manufacturer + external foundry. Intel Foundry now accepts external customers (like TSMC model)."),
    ("Intel Foundry Services (IFS)", "Growing business unit offering fab capacity to external chip designers. Competing with TSMC and Samsung for foundry market share. Major US gov't CHIPS Act funding (~$8.5B)."),
    ("SC Complexity", "Intel manages one of the world's most complex SC operations: semiconductor fab equipment ($100B+), substrates, packaging materials, logistics across 7 countries."),
    ("Key Locations", "Santa Clara CA (HQ), Hillsboro OR (R&D), Phoenix/Chandler AZ (fabs), Folsom CA (Foundry ops), Albuquerque NM (fab), Ohio (new fab campus under construction)."),
    ("Intel SC Transformation", "Under IDM 2.0, Intel is rebuilding its entire external SC planning infrastructure -- high demand for SC planning architects, integrators, and capacity planners."),
    ("Contract Roles Context", "Intel has been using contract roles aggressively post-2024 restructuring. Contracts often convert to FTE within 6-12 months. Strong performers get absorbed."),
    ("Interview Culture", "Intel values: 'fearless' problem-solving, data-driven decisions, collaboration across disciplines. Expect technical deep-dives into SC processes + STAR behavioral questions."),
    ("Salary Philosophy", "Intel base + annual bonus (target 10-15%) + RSU (stock grants). Total comp is 20-30% above base. Intel RSUs vest quarterly after 1-year cliff."),
    ("LinkedIn Company ID", "1053  |  64 SC/Planning jobs found  |  Use f_C=1053 for future LinkedIn job searches at Intel"),
]

for i, (label, value) in enumerate(INTEL_FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 30
    bg = INTEL_LIGHT if i % 2 == 0 else "EBF5FB"

    lc = ws3.cell(row=r, column=1, value=label)
    lc.font = Font(name="Arial", bold=True, size=9, color=INTEL_DARK)
    lc.fill = PatternFill("solid", start_color=bg)
    lc.alignment = left()
    lc.border = border

    ws3.merge_cells(f"B{r}:H{r}")
    vc = ws3.cell(row=r, column=2, value=value)
    vc.font = nf(size=9)
    vc.fill = PatternFill("solid", start_color=bg)
    vc.alignment = left()
    vc.border = border

# --- Interview Prep ---
int_start = len(INTEL_FACTS) + 5

ws3.row_dimensions[int_start].height = 22
ws3.merge_cells(f"A{int_start}:H{int_start}")
sec2 = ws3[f"A{int_start}"]
sec2.value = "INTEL INTERVIEW PREP -- Key Talking Points for SC & Foundry Planning Roles"
sec2.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec2.alignment = center()
sec2.fill = PatternFill("solid", start_color=INTEL_TEAL)
sec2.border = border

INT_HDR = int_start + 1
ws3.row_dimensions[INT_HDR].height = 22
for c, h in enumerate(["#", "Interview Topic", "What Intel Tests", "How to Answer (SC Consultant Angle)"], 1):
    cell = ws3.cell(row=INT_HDR, column=c)
    cell.value = h
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=INTEL_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

INTERVIEW_QS = [
    (1, "Why Intel / Why Foundry?",
     "Genuine understanding of Intel's IDM 2.0 transformation and what makes this role unique vs. competitors",
     "Talk about Intel's unique position: only company building both leading-edge logic AND foundry capacity in the US. Cite CHIPS Act investment and the opportunity to build SC infrastructure for next-gen AI chips. Show you understand the Foundry transformation challenge."),
    (2, "SC Planning Methodology",
     "Depth in S&OP, demand/capacity planning processes, and how you structure planning cycles",
     "Walk through your S&OP process ownership: how you run the monthly cadence, how you build consensus between demand/supply/finance, how you handle exceptions. Use specific metrics: forecast error %, inventory turns, service level %. Intel wants process-rigorous planners."),
    (3, "Handling Supply Constraints",
     "How you manage supply shortfalls, prioritize customers, and communicate bad news",
     "Share a story about a supply constraint situation: how you modeled the impact, communicated to stakeholders, prioritized allocation, and recovered. Intel Foundry's entire value prop is supply assurance -- show you've managed constrained supply environments."),
    (4, "Planning Tools & Systems",
     "Depth in APS, ERP, IBP, or planning tool implementations and integrations",
     "Be specific: name the tools (IBP, o9, Kinaxis, BY, SAP APO), the modules you configured, the integrations you built. Intel uses SAP S/4HANA and custom planning tools. Frame your tool implementations as 'building planning infrastructure from scratch for new customers' -- exactly what Intel Foundry needs."),
    (5, "Cross-functional Leadership",
     "Ability to drive alignment across engineering, operations, finance, and customers without direct authority",
     "Intel SC roles require influencing manufacturing, fab ops, technology dev, and customer teams simultaneously. Share stories of driving alignment across 4+ functions. Emphasize how you translated technical constraints into business decisions for executives."),
    (6, "Dealing with Ambiguity / Change",
     "Resilience and adaptability in a company undergoing major transformation (Intel is restructuring)",
     "Intel is in the middle of IDM 2.0 -- priorities shift, org structures change. Show you thrive in transformation environments. Consulting background is PERFECT here: every new client engagement is ambiguous. Frame each engagement as 'building structure where there was none.'"),
]

INT_FILLS = [INTEL_LIGHT, "EBF5FB", INTEL_LIGHT, "EBF5FB", INTEL_LIGHT, "EBF5FB"]

for i, (num, topic, tests, how) in enumerate(INTERVIEW_QS):
    r = INT_HDR + 1 + i
    ws3.row_dimensions[r].height = 58
    bg = INT_FILLS[i % len(INT_FILLS)]
    fill = PatternFill("solid", start_color=bg)

    for c, val in enumerate([num, topic, tests, how], 1):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=13, color=INTEL_BLUE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=INTEL_DARK)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 62
ws3.freeze_panes = f"A{int_start + 2}"

# =============================================================================
# SAVE
# =============================================================================
output = "C:/Users/amolp/Prometheus/intel_top5_jobs.xlsx"
wb.save(output)
print("Done! intel_top5_jobs.xlsx saved with 3 sheets.")

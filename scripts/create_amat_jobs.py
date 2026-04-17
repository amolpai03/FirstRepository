from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bot = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="404040"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink(cell, url, label="View Job →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Applied Materials brand palette
AMAT_DARK   = "002060"   # Deep navy
AMAT_BLUE   = "0070C0"   # Mid blue
AMAT_TEAL   = "00B0F0"   # Tech teal
AMAT_ORANGE = "ED7D31"   # Accent orange
AMAT_L_BLUE = "DEEAF1"   # Light blue fill
AMAT_L_ORG  = "FCE4D6"   # Light orange fill
AMAT_L_TEAL = "DDEBF7"   # Light teal fill
AMAT_L_GRN  = "E2EFDA"   # Light green fill
AMAT_L_PURP = "EDE7F6"   # Light purple fill

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — TOP 5 JOBS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "AMAT Top 5 Jobs"

ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "APPLIED MATERIALS — Top 5 Matching Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=AMAT_DARK)
t.border = border

ws1.row_dimensions[2].height = 5
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=AMAT_BLUE)

ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "69 school alumni at Applied Materials  ·  Semiconductor equipment leader  ·  171 open jobs in US  ·  LinkedIn Company ID: 2018"
s.font = Font(name="Arial", italic=True, size=9, color=AMAT_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color="EEF5FF")
s.border = border

HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match",
           "Applicants / Posted", "Experience Req.", "Key Skills", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AMAT_BLUE)
    cell.alignment = center()
    cell.border = thick_bot

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Product Planner IV",
        "location": "Santa Clara, CA\n(On-site · 100%)",
        "type": "On-site · Full-time",
        "salary": "$110,500 – $152,000",
        "match": "HIGH MATCH",
        "match_bg": "C6EFCE", "match_fg": "375623",
        "info": "Posted recently\n69 school alumni",
        "exp": "8+ yrs total\n6+ yrs demand/inventory planning",
        "skills": "Demand forecasting · Inventory mgmt\nERP: SAP / Oracle / SPP\nTableau · Power BI · SQL\nKPI management · NPI planning\nSemiconductor industry preferred",
        "url": "https://www.linkedin.com/jobs/view/4383169828/",
        "network": "69 school alumni\nat AMAT",
        "strategy": "STRONGEST match — demand + inventory planning IS your core skill. 8+ yrs = your seniority. SC planning tool expertise (o9/BY/Kinaxis) directly maps to ERP/SPP planning. Lead with forecast accuracy improvements and inventory KPIs from your projects.",
        "fill": AMAT_L_GRN,
    },
    {
        "num": 2,
        "title": "Supply Chain Project Manager",
        "location": "Santa Clara, CA\n(On-site · 100%)",
        "type": "On-site · Full-time",
        "salary": "$101,000 – $139,000",
        "match": "HIGH MATCH",
        "match_bg": "C6EFCE", "match_fg": "375623",
        "info": "2 months ago\n69 school alumni",
        "exp": "4–7 yrs in program mgmt\nor SC analytics",
        "skills": "SC capacity management & planning\nProject leadership · Cross-functional\nExcel · Databricks · Tableau\nPower BI · SQL · PMP (preferred)\nExecutive presentations",
        "url": "https://www.linkedin.com/jobs/view/4351608512/",
        "network": "69 school alumni\nat AMAT",
        "strategy": "Supply Chain in title = direct match. 4-7 yr exp req = within your band. Capacity planning + SC transformation = what you've done on o9/Kinaxis projects. Highlight Databricks/SQL experience or willingness to learn. PMP cert = strong advantage here.",
        "fill": AMAT_L_BLUE,
    },
    {
        "num": 3,
        "title": "Business Intelligence Analyst (B3)",
        "location": "Austin, TX\n(On-site · 100%)",
        "type": "On-site · Full-time",
        "salary": "$122,000 – $168,000",
        "match": "STRONG MATCH",
        "match_bg": "FFEB9C", "match_fg": "7F6000",
        "info": "1 week ago\n69 school alumni",
        "exp": "B3 level (3–5 yrs)\ndata & analytics focus",
        "skills": "SQL · Power BI · Tableau\nAdvanced Excel + Power Query\nInventory behavior & variance analysis\nData storytelling & dashboards\nSupply chain data + Finance reporting",
        "url": "https://www.linkedin.com/jobs/view/4397835802/",
        "network": "69 school alumni\nat AMAT",
        "strategy": "Highest salary of the 5 roles ($122K-$168K). SC + inventory data focus is squarely in your wheelhouse. Lead with any dashboards, KPI reporting, or data tools work from your SC projects. SQL + Power BI skills will be the differentiator.",
        "fill": AMAT_L_ORG,
    },
    {
        "num": 4,
        "title": "Materials Project Manager\n(Multiple Openings)",
        "location": "Austin, TX\n(On-site · 100%)",
        "type": "On-site · Full-time",
        "salary": "$88,000 – $121,000\n(mid-level: $108K–$148.5K)",
        "match": "STRONG MATCH",
        "match_bg": "FFEB9C", "match_fg": "7F6000",
        "info": "1 day ago — FRESH!\nMultiple openings",
        "exp": "Mid-level preferred\ncross-functional PM",
        "skills": "Materials planning · Supplier mgmt\nCross-functional matrix PM\nProduction control coordination\nSupplier negotiations · Cost mgmt\nChange control process",
        "url": "https://www.linkedin.com/jobs/view/4399348097/",
        "network": "69 school alumni\nat AMAT",
        "strategy": "Multiple openings = multiple shots! Only 1 day old = apply IMMEDIATELY. SC planning implementations = direct experience with materials planning and supplier coordination. Highlight any NPI (New Product Introduction) or supplier management experience.",
        "fill": AMAT_L_TEAL,
    },
    {
        "num": 5,
        "title": "Strategic Sourcing Manager",
        "location": "Austin, TX\n(On-site · 100%)",
        "type": "On-site · Full-time",
        "salary": "$120,000 – $165,000\n(estimated)",
        "match": "GOOD MATCH",
        "match_bg": "DDEBF7", "match_fg": "1F4E79",
        "info": "Viewed recently\n3 connections + 69 alumni",
        "exp": "5–8 yrs sourcing\nor SC management",
        "skills": "Strategic sourcing · Category mgmt\nSupplier development & negotiations\nSpend analysis · RFx management\nCost reduction programs\nCross-functional leadership",
        "url": "https://www.linkedin.com/jobs/view/4397080067/",
        "network": "3 direct connections!\n69 school alumni",
        "strategy": "3 direct LinkedIn connections at AMAT = highest network advantage of all 5 roles. Message connections BEFORE applying. SC planning background shows understanding of supply base and procurement dependencies. Stretch role but leverage connections.",
        "fill": AMAT_L_PURP,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 100
    base_fill = PatternFill("solid", start_color=job["fill"])

    vals = [job["num"], job["title"], job["location"], job["type"], job["salary"],
            job["match"], job["info"], job["exp"], job["skills"],
            job["url"], job["network"], job["strategy"]]

    for c, val in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = base_fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=AMAT_BLUE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=AMAT_DARK)
            cell.alignment = left()
        elif c == 5:
            cell.value = val
            cell.font = nf(bold=True, size=9, color="155724")
            cell.alignment = left()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_fresh = "1 day" in str(val) or "FRESH" in str(val)
            cell.font = Font(name="Arial", bold=is_fresh, size=9,
                           color="C00000" if is_fresh else "595959")
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = base_fill
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

COL_WIDTHS = [4, 28, 20, 14, 20, 13, 18, 20, 34, 12, 18, 44]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RESUME TAILORING GUIDE
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING — What to Highlight for Each Applied Materials Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=AMAT_DARK)
t2.border = border

ws2.row_dimensions[2].height = 5
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=AMAT_BLUE)

HEADERS2 = ["Role", "Must Highlight", "Keywords to Use", "Lead Experience", "Gap", "Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AMAT_BLUE)
    cell.alignment = center()
    cell.border = thick_bot

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Product Planner IV\n(Santa Clara, CA)",
        "• 8+ yrs demand/supply planning\n• Forecast accuracy improvements\n• Inventory optimization KPIs\n• ERP/planning system hands-on\n• Cross-functional collaboration",
        "Demand planning, supply planning, inventory management, NPI, ERP, SAP, SPP, Oracle, S&OP, KPI, forecast accuracy, service level, end-of-life planning, semiconductor",
        "Quantify your planning impact: e.g. '15% improvement in forecast accuracy' or '$2M inventory reduction'. Show full planning cycle ownership: NPI → ramp → EOL",
        "Semiconductor industry exp preferred. SPP tool knowledge (Applied's own tool)",
        "Frame o9/Kinaxis as advanced planning systems equivalent to SPP. Semiconductor = high-mix low-volume planning — show any comparable complexity from your industry",
        "APPLY FIRST\nHIGH MATCH\n$110K-$152K"
    ),
    (
        "Supply Chain Project Manager\n(Santa Clara, CA)",
        "• SC transformation projects led\n• Capacity planning experience\n• Data-driven SC presentations\n• Cross-functional stakeholder mgmt\n• Excel/Tableau/SQL proficiency",
        "Supply chain, capacity planning, project management, transformation, cross-functional, Tableau, Databricks, SQL, executive presentations, process improvement, risk management",
        "Lead with your largest SC planning implementation. Show breadth: managed X stakeholders, $Y project value, Z% process improvement. Mention any PMP/CAPM certification",
        "Databricks is preferred — not standard in planning consulting. 10% travel required",
        "Frame SC planning tool implementations as 'enterprise SC transformation projects'. Highlight any data pipeline or analytics tool exposure. Learn basic Databricks concepts before interview",
        "APPLY — GOOD FIT\n$101K-$139K"
    ),
    (
        "Business Intelligence Analyst B3\n(Austin, TX)",
        "• SQL querying experience\n• Power BI / Tableau dashboards\n• SC data analysis & reporting\n• Advanced Excel (Power Query)\n• Data storytelling to leadership",
        "Business intelligence, SQL, Power BI, Tableau, Excel, PivotTables, XLOOKUP, Power Query, data modeling, dashboards, supply chain analytics, inventory variance, KPIs",
        "Lead with any SC analytics or reporting work — even Excel-based KPI dashboards count. Show how your data analysis drove a planning decision. $122K-$168K = strong SC analyst comp",
        "Analyst title may feel like a step down from consulting. B3 = mid-level at AMAT. 100% onsite Austin required",
        "Frame as 'transitioning from SC consulting to industry analyst role with direct business impact'. Highest salary of 5 roles ($168K max). Build a quick Power BI portfolio sample before applying",
        "STRONG FIT\nHIGHEST SALARY\n$122K-$168K"
    ),
    (
        "Materials Project Manager\n(Multiple Openings, Austin TX)",
        "• Materials planning coordination\n• Supplier management exposure\n• Cross-functional PM experience\n• Change control process\n• Production planning knowledge",
        "Materials planning, project manager, supplier, sourcing, NPI, production control, matrix management, change control, cost management, supply chain, planning, BOM",
        "Lead with SC planning project management. Show you've coordinated materials/parts across a supply chain. Any supplier negotiations or vendor management = direct relevance",
        "Role is more operational/execution focused vs strategic planning consulting. Salary lower ($88K-$121K for base level)",
        "Multiple openings = higher acceptance rate. Fresh posting = low competition. Apply to both Austin locations if listed. Mid-level range ($108K-$148.5K) is much better — target that grade",
        "APPLY FAST\n1 DAY OLD!\nMultiple openings"
    ),
    (
        "Strategic Sourcing Manager\n(Austin, TX)",
        "• Any sourcing/procurement work\n• Category management exposure\n• Supplier negotiations led\n• SC strategy ownership\n• Cost reduction initiatives",
        "Strategic sourcing, category management, supplier development, RFx, spend analysis, negotiations, procurement, cost reduction, supplier qualification, cross-functional",
        "Lead with any procurement/sourcing-adjacent SC work. SC planning implementations involve supplier coordination — highlight that angle. Cost modeling work is very relevant here",
        "Primarily procurement/sourcing focused — not planning tech. May require deeper sourcing domain expertise",
        "You have 3 DIRECT LinkedIn connections at AMAT — message them first for a referral! Even a warm intro makes this reachable. Frame your SC consulting as 'total cost of ownership optimization'",
        "LEVERAGE NETWORK\n3 connections!\n~$120K-$165K"
    ),
]

ROW_FILLS  = [AMAT_L_GRN, AMAT_L_BLUE, AMAT_L_ORG, AMAT_L_TEAL, AMAT_L_PURP]
PRI_COLORS = [("C6EFCE","375623"),("C6EFCE","375623"),("FFEB9C","7F6000"),("FFEB9C","7F6000"),("DDEBF7","1F4E79")]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 120
    fill = PatternFill("solid", start_color=ROW_FILLS[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=AMAT_DARK)
            cell.fill = fill; cell.alignment = left()
        elif c == 7:
            bg, fg = PRI_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill; cell.alignment = left()

for i, w in enumerate([26, 38, 40, 40, 30, 40, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — COMPANY QUICK FACTS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("AMAT Quick Facts")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "APPLIED MATERIALS — Company Overview & Interview Reference"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=AMAT_DARK)
t3.border = border

ws3.row_dimensions[2].height = 5
ws3.merge_cells("A2:D2")
ws3["A2"].fill = PatternFill("solid", start_color=AMAT_BLUE)

FACTS = [
    ("Company",         "Applied Materials, Inc. — World's largest semiconductor equipment & services company"),
    ("HQ",              "Santa Clara, California  ·  Major offices: Austin TX, Boise ID, Gloucester MA, global"),
    ("Revenue",         "$27.2B (FY2024)  ·  Fortune 200  ·  NASDAQ: AMAT  ·  ~35,000 employees globally"),
    ("Business",        "Equipment, services & software for chip manufacturing (logic, memory, display, solar)"),
    ("Customers",       "TSMC, Samsung, Intel, Micron, SK Hynix, ASML — virtually all major chipmakers"),
    ("SC Network",      "69 school alumni at Applied Materials — strong referral network available"),
    ("Connections",     "3 direct LinkedIn connections on Strategic Sourcing Manager role — use them!"),
    ("Work Style",      "All 5 roles are 100% ONSITE — no hybrid/remote. SC + tech ops culture"),
    ("Salary Range",    "BI Analyst: $122K-$168K  |  Product Planner IV: $110K-$152K  |  SC PM: $101K-$139K"),
    ("Salary Range 2",  "Materials PM: $88K-$148.5K (level-dependent)  |  Strategic Sourcing Mgr: ~$120K-$165K"),
    ("Why AMAT?",       "Massive SC complexity, global supply base, cutting-edge semiconductor demand planning"),
    ("Planning Tools",  "SAP ERP, SPP (Applied's proprietary planning tool), Oracle, Tableau, Databricks, SQL"),
    ("Interview Tip 1", "Research Applied Materials' chip equipment cycle — questions about SC volatility are common"),
    ("Interview Tip 2", "Know TSMC/Intel/Samsung capacity planning challenges — this IS Applied Materials' world"),
    ("Interview Tip 3", "Semiconductor supply chains = extreme complexity: long lead times, single-source risk, geopolitics"),
    ("LinkedIn ID",     "2018  ·  linkedin.com/company/applied-materials"),
    ("Careers Page",    "jobs.appliedmaterials.com"),
]

ws3.row_dimensions[3].height = 22
for c, h in enumerate(["Category", "Details", "", ""], 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AMAT_BLUE)
    cell.alignment = center()
    cell.border = thick_bot

for i, (cat, detail) in enumerate(FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 26
    alt_fill = PatternFill("solid", start_color="EEF5FF" if i % 2 == 0 else WHITE)

    c1 = ws3.cell(row=r, column=1, value=cat)
    c1.font = nf(bold=True, size=9, color=AMAT_DARK)
    c1.fill = alt_fill; c1.alignment = left(); c1.border = border

    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c2 = ws3.cell(row=r, column=2, value=detail)
    c2.font = nf(size=9); c2.fill = alt_fill
    c2.alignment = left(); c2.border = border

ws3.freeze_panes = "A4"
for i, w in enumerate([20, 80, 1, 1], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
output = "C:/Users/amolp/Prometheus/applied_materials_top5_jobs.xlsx"
wb.save(output)
print("Done! applied_materials_top5_jobs.xlsx saved with 3 sheets.")

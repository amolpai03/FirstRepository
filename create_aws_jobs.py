from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE  = "FFFFFF"
thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="595959"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def hyperlink(cell, url, label="View Job ->"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# AWS brand colors
AWS_ORANGE  = "FF9900"   # Amazon/AWS orange
AWS_DARK    = "232F3E"   # AWS dark navy
AWS_BLUE    = "146EB4"   # AWS blue
AWS_LIGHT   = "FFF3E0"   # Light orange tint
AWS_BLUE_LIGHT = "E8F4FF"  # Light blue tint
AWS_TEAL    = "00A591"   # AWS teal
AWS_GREEN_LIGHT = "E8F5E9"
GOLD_LIGHT  = "FFF8EC"
GOLD        = "F5A623"
GREEN_LIGHT = "C6EFCE"
GREEN_DARK  = "375623"
YELLOW_LIGHT = "FFEB9C"
YELLOW_DARK  = "7F6000"
BLUE_BADGE  = "DDEBF7"
BLUE_DARK   = "1F4E79"

# =============================================================================
# SHEET 1 - AWS TOP 5 JOBS
# =============================================================================
ws1 = wb.active
ws1.title = "AWS Top 5 Jobs"

# Title banner
ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "AMAZON WEB SERVICES (AWS) -- Top 5 Supply Chain & Planning Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=AWS_DARK)
t.border = border

# Spacer
ws1.row_dimensions[2].height = 6
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=AWS_ORANGE)

# Section header
ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "Job 1 has 11 o9 alumni at AWS  |  Jobs 1-2-3 all in Seattle  |  AWS Company ID on LinkedIn: 2382910  |  1,885 total results searched"
s.font = Font(name="Arial", italic=True, size=9, color=AWS_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color=AWS_LIGHT)
s.border = border

# Column headers
HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match Level",
           "Status / Alumni", "Posted", "Key Skills Required", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AWS_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws1.freeze_panes = "A5"

# Job data
JOBS = [
    {
        "num": 1,
        "title": "Senior S&OP Planner\nInfrastructure Planning\nAWS S&OP Demand Team",
        "location": "Seattle, WA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$134,000 - $231,000\n(Glassdoor estimate)\nIncludes RSU + sign-on",
        "match": "PERFECT MATCH",
        "match_bg": GREEN_LIGHT,
        "match_fg": GREEN_DARK,
        "status": "Be an Early Applicant!\n11 o9 Solutions alumni\nat AWS -- huge network edge!",
        "posted": "Recently posted",
        "skills": "S&OP platforms: SAP IBP, Blue Yonder,\nKinaxis, o9/Connect, Infor, Manhattan\nSQL + statistical forecasting\nExecutive S&OP facilitation\nDemand Plan of Record ownership",
        "url": "https://www.linkedin.com/jobs/view/4390829562/",
        "network": "11 o9 alumni at AWS\n= direct referral\nopportunity!",
        "strategy": "APPLY FIRST -- S&OP literally in job title + your o9/BY/Kinaxis/IBP toolkit is a perfect match. Role owns the GLOBAL S&OP cycle for AWS GenAI & Core infrastructure. Message an o9 alum at AWS immediately for referral before applying.",
        "fill": AWS_GREEN_LIGHT,
    },
    {
        "num": 2,
        "title": "Sr. TPM\nAmazon Supply Chain",
        "location": "Seattle, WA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$148,700 - $210,000\n(Senior TPM range)\nIncludes RSU + sign-on",
        "match": "HIGH MATCH",
        "match_bg": YELLOW_LIGHT,
        "match_fg": YELLOW_DARK,
        "status": "Be an Early Applicant!\n(Viewed)\nFresh posting",
        "posted": "Recently posted",
        "skills": "Supply Chain Planning expertise\nERP systems (Oracle EBS / SAP)\nPLM application experience\nLarge-scale program management\n5+ yrs SC planning operations",
        "url": "https://www.linkedin.com/jobs/view/4398392828/",
        "network": "Amazon NCSU alumni\nnetwork (525 school\nalumni at Amazon)",
        "strategy": "Strong fit for consulting background -- TPM = program ownership + SC systems technical depth. Mirrors what you do as SC Solutions Consultant. Highlight ERP/APS integration projects and cross-functional program leadership. Early applicant advantage.",
        "fill": GOLD_LIGHT,
    },
    {
        "num": 3,
        "title": "Sr. Global Supply Chain Manager\nAWS Infrastructure Services",
        "location": "Seattle, WA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$133,000 - $235,000\n(Senior SC Manager range)\nIncludes RSU + sign-on",
        "match": "HIGH MATCH",
        "match_bg": YELLOW_LIGHT,
        "match_fg": YELLOW_DARK,
        "status": "2 weeks ago\nEarly applicant eligible",
        "posted": "2 weeks ago",
        "skills": "Supply chain management\n6+ yrs SC experience required\nSupplier performance management\n4+ yrs infrastructure/data center\nCross-functional SC leadership",
        "url": "https://www.linkedin.com/jobs/view/4399476915/",
        "network": "AWS/Amazon alumni\nnetwork accessible\nvia LinkedIn",
        "strategy": "SC management role at one of the world's largest supply chains. 6+ years experience + broad SC exposure from consulting = strong fit. Frame multi-client consulting as cross-industry SC expertise. Highlight any infrastructure or manufacturing client projects.",
        "fill": AWS_BLUE_LIGHT,
    },
    {
        "num": 4,
        "title": "Solutions Architect\nConsumer Packaged Goods (CPG)",
        "location": "Arlington, VA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$131,300 - $177,600\n(Base only)\n+ RSU + bonus",
        "match": "MEDIUM MATCH",
        "match_bg": BLUE_BADGE,
        "match_fg": BLUE_DARK,
        "status": "Be an Early Applicant!\nSaved on LinkedIn",
        "posted": "Recently posted",
        "skills": "AWS cloud services knowledge\nCPG industry expertise\nSC/retail domain depth\nGenAI & Agentic AI for CPG\nCustomer-facing SA skills\nAWS certifications a plus",
        "url": "https://www.linkedin.com/jobs/view/4381124566/",
        "network": "AWS SA community\n+ CPG industry\nnetwork overlap",
        "strategy": "STRETCH ROLE -- AWS SA focused on CPG vertical. Your SC planning domain expertise in CPG is a key differentiator vs. cloud-only candidates. Role helps CPG cos. modernize SC using AWS. Pair with AWS Cloud Practitioner cert to close tech gap. Apply if you have any AWS/cloud exposure.",
        "fill": AWS_LIGHT,
    },
    {
        "num": 5,
        "title": "Sr. Global Procurement Category Manager\nInfrastructure Planning & Sourcing",
        "location": "Herndon, VA\n(On-site)",
        "type": "On-site\nFull-time",
        "salary": "$120,000 - $175,000\n(estimated, similar AWS\nprocurement roles)",
        "match": "MEDIUM MATCH",
        "match_bg": BLUE_BADGE,
        "match_fg": BLUE_DARK,
        "status": "1 week ago\nEarly applicant eligible",
        "posted": "1 week ago",
        "skills": "Strategic sourcing / category mgmt\n10+ yrs procurement experience\nInfrastructure supply chain\nCross-functional negotiation\nVendor performance management",
        "url": "https://www.linkedin.com/jobs/view/4385316430/",
        "network": "AWS/Amazon alumni\nnetwork in Herndon VA\n(large presence)",
        "strategy": "SC-adjacent role in AWS infrastructure procurement. Apply if you have strategic sourcing or procurement exposure from consulting clients. Category management + infrastructure knowledge = key qualifiers. Lower priority vs. Jobs 1-3 which directly match planning tech background.",
        "fill": "F0F0F0",
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 95
    fill = PatternFill("solid", start_color=job["fill"])

    data = [
        job["num"], job["title"], job["location"], job["type"],
        job["salary"], job["match"], job["status"], job["posted"],
        job["skills"], job["url"], job["network"], job["strategy"]
    ]

    for c, val in enumerate(data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:  # #
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=AWS_ORANGE)
            cell.alignment = center()
        elif c == 2:  # Title
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=AWS_DARK)
            cell.alignment = left()
        elif c == 6:  # Match badge
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:  # Status/Alumni
            cell.value = val
            has_network = "o9 alumni" in str(val) or "Early" in str(val)
            cell.font = Font(name="Arial", bold=has_network, size=9,
                           color="C00000" if "o9 alumni" in str(val) else ("375623" if "Early" in str(val) else "595959"))
            cell.alignment = center()
        elif c == 10:  # Link
            hyperlink(cell, val)
            cell.fill = fill
        elif c == 12:  # Strategy
            cell.value = val
            cell.font = nf(size=9, color="1A1A1A")
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

# Column widths
COL_WIDTHS = [4, 28, 18, 14, 22, 14, 22, 12, 34, 12, 18, 46]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# SHEET 2 - RESUME TAILORING GUIDE
# =============================================================================
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING -- What to Highlight for Each AWS Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=AWS_DARK)
t2.border = border

ws2.row_dimensions[2].height = 6
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=AWS_ORANGE)

HEADERS2 = ["Role", "Must Highlight in Resume", "Keywords to Include",
            "Experience to Lead With", "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AWS_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Senior S&OP Planner\nAWS S&OP Demand Team",
        "- S&OP cycle ownership (monthly/quarterly)\n- Demand Plan of Record (POR) experience\n- Executive S&OP facilitation & leadership\n- SC planning tool hands-on expertise\n- Cross-functional alignment across supply/demand/finance",
        "S&OP, Demand Planning, Supply Planning, Demand Plan of Record, POR, IBP, Kinaxis, Blue Yonder, o9, SAP IBP, SQL, statistical forecasting, executive planning forum, consensus forecast",
        "Lead with S&OP program ownership examples: # of stakeholders, $ value of decisions made, accuracy % improvement. Show tool depth (workbooks built, modules configured). Quantify impact.",
        "AWS is internal S&OP (not consulting) -- they want someone who has DONE S&OP, not advised on it",
        "Frame consulting projects as 'embedded in client S&OP team' -- emphasize you ran S&OP cadences, not just implemented tools. Show outcomes: forecast accuracy %, inventory turns, on-time supply.",
        "APPLY FIRST\n(PERFECT MATCH)"
    ),
    (
        "Sr. TPM\nAmazon Supply Chain",
        "- Technical program management experience\n- ERP/APS system implementations\n- Large-scale cross-functional project ownership\n- SC planning application delivery\n- Consulting = multi-project program leadership",
        "TPM, Technical Program Manager, ERP, SAP, Oracle EBS, PLM, supply chain systems, program management, release schedule, APS, planning applications, cross-functional, single-threaded leader",
        "Lead with complex SC technology implementations you led end-to-end. Show: scope (# users, # integrations, $ investment), timeline management, stakeholder count, and measurable outcomes delivered.",
        "TPM at Amazon is a technical role -- needs coding/SQL familiarity and strong technical systems depth",
        "Emphasize data model design, integration architecture work in SC tool implementations. Show SQL/data skills. Frame yourself as bridge between business and IT -- which is exactly what SC consultants do.",
        "APPLY -- HIGH\nMATCH"
    ),
    (
        "Sr. Global SC Manager\nAWS Infrastructure Svcs",
        "- Global supply chain management\n- Supplier performance management\n- Cross-industry SC exposure from consulting\n- Infrastructure/manufacturing client work\n- SC risk mitigation and resilience",
        "Global supply chain, supplier management, procurement, category management, infrastructure, data center, capacity planning, supplier performance, KPI, risk mitigation, cost reduction",
        "Lead with breadth of SC experience across clients/industries. Show global scope: # countries, $ spend managed, # suppliers. Any infrastructure, utilities, or manufacturing client work is gold here.",
        "6+ yrs in infrastructure/data center environments preferred -- not typical consulting SC background",
        "Frame manufacturing or industrial client SC projects prominently. If no direct DC experience, position logistics/network optimization projects for large-scale ops as adjacent experience.",
        "APPLY -- HIGH\nMATCH"
    ),
    (
        "Solutions Architect\nConsumer Packaged Goods",
        "- CPG industry expertise & client-facing skills\n- SC/supply chain domain knowledge for CPG\n- Solutions consulting/architecture background\n- Ability to demo/present technical solutions\n- Any AWS/cloud exposure or certifications",
        "Solutions Architect, AWS, cloud, CPG, Consumer Packaged Goods, retail, supply chain, GenAI, Agentic AI, S&OP, demand sensing, cloud migration, digital transformation, pre-sales, customer advisory",
        "Lead with CPG industry wins and client-facing advisory skills. Show: # CPG clients served, supply chain use cases solved, business impact delivered. Any AWS/cloud project involvement is a bonus.",
        "Primarily a cloud SA role -- requires AWS technical depth, cloud architecture design, and AWS certifications (SAA, SAP)",
        "Get AWS Cloud Practitioner cert ASAP (free exam vouchers available). Emphasize CPG SC domain depth as industry differentiator vs. cloud-only SAs who lack supply chain context. Apply + pursue cert in parallel.",
        "STRETCH ROLE\n(apply if\ncloud exp)"
    ),
    (
        "Sr. Global Procurement\nCategory Manager",
        "- Strategic sourcing experience\n- Category management & vendor negotiations\n- Infrastructure supply chain knowledge\n- Cross-functional stakeholder management\n- Cost analysis & supplier qualification",
        "Strategic sourcing, category management, procurement, sourcing strategy, supplier qualification, negotiation, SOW, cost reduction, infrastructure, MRO, CapEx, vendor management, KPIs",
        "Lead with any procurement or sourcing exposure from consulting work. Even SC optimization projects with procurement scope count. Show: $ savings achieved, # contracts negotiated, supplier diversity.",
        "10+ years procurement/sourcing preferred -- this is a procurement specialist role, not a planning tech role",
        "Apply only if you have significant sourcing/procurement experience from consulting clients. Otherwise prioritize Jobs 1-3 which directly align with your planning technology background.",
        "OPTIONAL\n(lower priority)"
    ),
]

ROW_FILLS = [AWS_GREEN_LIGHT, GOLD_LIGHT, AWS_BLUE_LIGHT, AWS_LIGHT, "F0F0F0"]
PRIORITY_COLORS = [
    (GREEN_LIGHT, GREEN_DARK),
    (YELLOW_LIGHT, YELLOW_DARK),
    (YELLOW_LIGHT, YELLOW_DARK),
    (BLUE_BADGE, BLUE_DARK),
    (BLUE_BADGE, BLUE_DARK),
]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 115
    fill_color = ROW_FILLS[i]
    fill = PatternFill("solid", start_color=fill_color)

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border

        if c == 1:  # Role name
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=AWS_DARK)
            cell.fill = fill
            cell.alignment = left()
        elif c == 7:  # Priority badge
            bg, fg = PRIORITY_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

COL_WIDTHS2 = [26, 40, 40, 42, 32, 40, 16]
for i, w in enumerate(COL_WIDTHS2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# SHEET 3 - AWS QUICK FACTS + AMAZON LEADERSHIP PRINCIPLES
# =============================================================================
ws3 = wb.create_sheet("AWS Facts + LP Guide")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "AWS QUICK FACTS + Amazon Leadership Principles -- Interview Prep Cheat Sheet"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=AWS_DARK)
t3.border = border

ws3.row_dimensions[2].height = 6
ws3.merge_cells("A2:H2")
ws3["A2"].fill = PatternFill("solid", start_color=AWS_ORANGE)

# --- Section: AWS Company Facts ---
ws3.row_dimensions[3].height = 22
ws3.merge_cells("A3:H3")
sec1 = ws3["A3"]
sec1.value = "AWS COMPANY FACTS -- Know Before Your Interview"
sec1.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec1.alignment = center()
sec1.fill = PatternFill("solid", start_color=AWS_BLUE)
sec1.border = border

AWS_FACTS = [
    ("AWS Market Position", "Largest cloud provider globally -- ~31% market share (Azure ~25%, GCP ~12%). AWS revenue: ~$107B annually (2024)."),
    ("AWS Infrastructure Scale", "200+ services, 33 geographic regions, 105+ Availability Zones, 600+ Points of Presence globally."),
    ("AWS Supply Chain Team", "Manages one of the world's most complex SC operations: servers, networking, power, cooling for ALL global data centers."),
    ("AWS S&OP Planning", "CDSA (Capacity Delivery, Supportability & Analytics) org owns AWS's 0-10 year infrastructure demand/capacity planning."),
    ("Key SC Challenges at AWS", "GenAI infrastructure demand surge (NVIDIA GPUs, custom silicon), global supply constraints, long-lead equipment planning."),
    ("LinkedIn Company ID", "2382910  |  Total SC/Planning jobs searched: 1,885  |  Use f_C=2382910 for future searches"),
    ("AWS Culture", "Customer Obsession is #1 LP. High bar for hiring (Bar Raiser system). Expect 4-6 rounds + STAR-format LP interviews."),
    ("Salary Philosophy", "Amazon is TC (total comp) company. Base + RSU (vesting: 5/15/40/40 over 4 yrs) + sign-on yr1 & yr2. RSU is major comp component."),
]

for i, (label, value) in enumerate(AWS_FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 30
    bg = AWS_LIGHT if i % 2 == 0 else AWS_BLUE_LIGHT

    label_cell = ws3.cell(row=r, column=1, value=label)
    label_cell.font = Font(name="Arial", bold=True, size=9, color=AWS_DARK)
    label_cell.fill = PatternFill("solid", start_color=bg)
    label_cell.alignment = left()
    label_cell.border = border

    ws3.merge_cells(f"B{r}:H{r}")
    val_cell = ws3.cell(row=r, column=2, value=value)
    val_cell.font = nf(size=9)
    val_cell.fill = PatternFill("solid", start_color=bg)
    val_cell.alignment = left()
    val_cell.border = border

# --- Section: Leadership Principles ---
lp_start = len(AWS_FACTS) + 5

ws3.row_dimensions[lp_start].height = 22
ws3.merge_cells(f"A{lp_start}:H{lp_start}")
sec2 = ws3[f"A{lp_start}"]
sec2.value = "AMAZON LEADERSHIP PRINCIPLES (LPs) -- Interview Answer Framework"
sec2.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec2.alignment = center()
sec2.fill = PatternFill("solid", start_color=AWS_ORANGE)
sec2.border = border

LP_HEADER_ROW = lp_start + 1
ws3.row_dimensions[LP_HEADER_ROW].height = 22
for c, h in enumerate(["#", "Leadership Principle", "What They're Testing", "SC Consultant Example Story to Prepare"], 1):
    cell = ws3.cell(row=LP_HEADER_ROW, column=c)
    cell.value = h
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AWS_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

LPS = [
    (1, "Customer Obsession", "Do you start with the customer and work backwards? Do you prioritize customer outcomes?",
     "Story: How you redesigned an S&OP process or planning tool to solve a client's core pain point -- not just deliver the tool spec."),
    (2, "Ownership", "Do you act beyond your job description? Do you see problems through to resolution?",
     "Story: A time you took ownership of a failing SC planning implementation or process gap without being asked. What you did, outcome."),
    (3, "Invent and Simplify", "Do you find creative, simpler solutions? Do you innovate even when it's not obvious?",
     "Story: A time you simplified a complex S&OP or planning process -- e.g., built automation, simplified a 30-step process into 5 steps."),
    (4, "Are Right, A Lot", "Do you use good judgment? Do you seek diverse viewpoints before deciding?",
     "Story: A time you disagreed with a client's approach to SC planning and ultimately persuaded them to the right solution with data."),
    (5, "Learn and Be Curious", "Do you continuously learn? Are you curious about new domains and technologies?",
     "Story: How you stayed current on SC planning tools (Kinaxis, o9, IBP), or learned a new tool/methodology for a client engagement."),
    (6, "Hire and Develop the Best", "Do you raise the bar? Do you mentor and develop others?",
     "Story: Mentoring a junior consultant or client team member on SC planning best practices or tool usage. Impact on their growth."),
    (7, "Insist on Highest Standards", "Do you refuse to accept mediocre work? Do you drive quality?",
     "Story: A time you pushed back on a low-quality SC planning design or data model and drove a better outcome for the client."),
    (8, "Think Big", "Do you have a bold vision? Do you think beyond obvious solutions?",
     "Story: A transformational SC planning roadmap or multi-year vision you developed for a client -- not just a quick fix."),
    (9, "Bias for Action", "Do you act with urgency? Do you make decisions without perfect info?",
     "Story: A time you made a fast SC planning decision with incomplete data and it worked out -- how you balanced speed vs. accuracy."),
    (10, "Frugality", "Do you accomplish more with less? Do you avoid unnecessary spending?",
     "Story: A time you delivered a SC planning solution within tight budget/time constraints -- what you prioritized and why."),
    (11, "Earn Trust", "Do you earn trust through transparency and candor? Do you hold yourself accountable?",
     "Story: A time you delivered difficult news to a client about a SC planning project (delay, scope change) and maintained trust."),
    (12, "Dive Deep", "Are you detail-oriented? Do you audit and verify with data?",
     "Story: A time you dug deep into SC data/system issue to find root cause -- e.g., forecast accuracy problem, data mapping error."),
    (13, "Have Backbone; Disagree and Commit", "Do you respectfully challenge decisions? Once decided, do you commit fully?",
     "Story: A time you disagreed with a client's tool selection (e.g., wrong APS tool) -- stated your case clearly, then fully committed to execution."),
    (14, "Deliver Results", "Do you focus on key inputs and deliver with high quality and timeliness?",
     "Story: Your biggest SC planning project outcome -- quantify: $ savings, % forecast accuracy improvement, # days cycle time reduction."),
    (15, "Strive to be Earth's Best Employer", "Do you create a safe, productive, diverse team environment?",
     "Story: How you fostered inclusion or psychological safety in a consulting team or cross-functional SC project team."),
    (16, "Success and Scale Bring Broad Responsibility", "Do you think beyond business outcomes to societal impact?",
     "Story: Any SC sustainability work -- e.g., helping a client reduce SC emissions, optimize routing for fuel efficiency, ethical sourcing."),
]

LP_FILLS = [AWS_GREEN_LIGHT, GOLD_LIGHT, AWS_BLUE_LIGHT, AWS_LIGHT]

for i, (num, lp, test, story) in enumerate(LPS):
    r = LP_HEADER_ROW + 1 + i
    ws3.row_dimensions[r].height = 52
    bg = LP_FILLS[i % 4]
    fill = PatternFill("solid", start_color=bg)

    data = [num, lp, test, story]
    for c, val in enumerate(data, 1):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=11, color=AWS_ORANGE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=AWS_DARK)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

COL_WIDTHS3 = [4, 26, 42, 60, 0, 0, 0, 0]
ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 60
ws3.freeze_panes = f"A{lp_start + 2}"

# =============================================================================
# SAVE
# =============================================================================
output = "C:/Users/amolp/Prometheus/aws_top5_jobs.xlsx"
wb.save(output)
print("Done! aws_top5_jobs.xlsx saved with 3 sheets.")

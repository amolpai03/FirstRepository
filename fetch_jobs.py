"""
fetch_jobs.py — Auto-fetch job listings via JSearch API and generate Excel files.

Setup:
  1. Sign up at https://rapidapi.com/letscrape-6bRBa3QguO5/api/jsearch (free: 200 req/month)
  2. Copy your API key and set it below or in env var RAPIDAPI_KEY
  3. Run: python3 fetch_jobs.py
"""

import os
from dotenv import load_dotenv
load_dotenv()
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────
RAPIDAPI_KEY = os.environ.get("RAPIDAPI_KEY", "YOUR_RAPIDAPI_KEY_HERE")

# Companies and their search queries + brand colors
TARGETS = [
    {"company": "Amazon",          "query": "Supply Chain Planning Amazon",          "color": "232F3E", "accent": "FF9900"},
    {"company": "Google",          "query": "Supply Chain Operations Google",         "color": "4285F4", "accent": "34A853"},
    {"company": "Tesla",           "query": "Supply Chain Planning Tesla",            "color": "CC0000", "accent": "1A1A1A"},
    {"company": "McKinsey",        "query": "Supply Chain Consultant McKinsey",       "color": "003366", "accent": "0066CC"},
    {"company": "EY",              "query": "Supply Chain Manager EY",                "color": "FFE600", "accent": "2E2E38"},
    {"company": "AWS",             "query": "Supply Chain Technical Program Manager AWS", "color": "232F3E", "accent": "FF9900"},
    {"company": "Intel",           "query": "Supply Chain Planning Intel",            "color": "0071C5", "accent": "00AEEF"},
    {"company": "Applied Materials","query": "Supply Chain Applied Materials",        "color": "003F7D", "accent": "009FDF"},
    {"company": "Mastercard",      "query": "Supply Chain Operations Mastercard",     "color": "EB001B", "accent": "F79E1B"},
    {"company": "Home Depot",      "query": "Supply Chain Planning Home Depot",       "color": "F96302", "accent": "1C3A6A"},
    {"company": "Rivian",          "query": "Supply Chain Planning Rivian",           "color": "00A0B0", "accent": "1A1A1A"},
]

RESULTS_PER_COMPANY = 5
OUTPUT_DIR = "."

# ── Shared styles ───────────────────────────────────────────────────────────
WHITE = "FFFFFF"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def fill(color):
    return PatternFill("solid", start_color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink_cell(cell, url, label="View Job →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# ── API fetch ────────────────────────────────────────────────────────────────
def fetch_jobs(query, num_results=5):
    url = "https://jsearch.p.rapidapi.com/search"
    headers = {
        "x-rapidapi-key": RAPIDAPI_KEY,
        "x-rapidapi-host": "jsearch.p.rapidapi.com"
    }
    params = {
        "query": f"{query} United States",
        "page": "1",
        "num_pages": "1",
        "country": "us",
        "date_posted": "month"
    }
    resp = requests.get(url, headers=headers, params=params, timeout=15)
    resp.raise_for_status()
    data = resp.json().get("data", [])
    return data[:num_results]

# ── Excel builder ────────────────────────────────────────────────────────────
HEADERS = ["#", "Job Title", "Company", "Location", "Type", "Salary", "Posted", "Link"]
COL_WIDTHS = [4, 35, 20, 22, 18, 22, 14, 16]

def build_excel(company, jobs, color, accent):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{company} Jobs"

    # Title row
    ws.row_dimensions[1].height = 34
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    title = ws["A1"]
    title.value = f"{company.upper()} — Top Supply Chain & Planning Jobs  |  {datetime.now().strftime('%B %Y')}"
    title.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    title.alignment = center()
    title.fill = fill(color)
    title.border = BORDER

    # Accent bar
    ws.row_dimensions[2].height = 5
    ws.merge_cells(f"A2:{get_column_letter(len(HEADERS))}2")
    ws["A2"].fill = fill(accent)

    # Header row
    ws.row_dimensions[3].height = 22
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = fill(color)
        cell.alignment = center()
        cell.border = BORDER

    ws.freeze_panes = "A4"

    # Set column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    alt_fill_color = "F2F2F2"
    for i, job in enumerate(jobs):
        row = i + 4
        ws.row_dimensions[row].height = 48
        row_fill = fill(alt_fill_color if i % 2 == 0 else WHITE)

        salary = ""
        if job.get("job_min_salary") and job.get("job_max_salary"):
            salary = f"${job['job_min_salary']:,.0f} – ${job['job_max_salary']:,.0f}"
            if job.get("job_salary_period"):
                salary += f"\n({job['job_salary_period']})"
        elif job.get("job_salary_currency"):
            salary = "See listing"

        posted = ""
        if job.get("job_posted_at_datetime_utc"):
            try:
                dt = datetime.fromisoformat(job["job_posted_at_datetime_utc"].replace("Z", "+00:00"))
                posted = dt.strftime("%b %d, %Y")
            except Exception:
                posted = job.get("job_posted_at_datetime_utc", "")[:10]

        values = [
            i + 1,
            job.get("job_title", ""),
            job.get("employer_name", ""),
            f"{job.get('job_city', '')}, {job.get('job_state', '')}".strip(", "),
            job.get("job_employment_type", ""),
            salary,
            posted,
            None,  # hyperlink handled separately
        ]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = left_align() if c > 1 else center()
            cell.font = font(bold=(c == 2))

        # Hyperlink in last column
        link_cell = ws.cell(row=row, column=len(HEADERS))
        job_url = job.get("job_apply_link") or job.get("job_google_link", "")
        if job_url:
            hyperlink_cell(link_cell, job_url)
        link_cell.fill = row_fill
        link_cell.border = BORDER
        link_cell.alignment = center()

    # No jobs fallback
    if not jobs:
        ws.row_dimensions[4].height = 30
        cell = ws.cell(row=4, column=1, value="No results found — try a different query or check your API key.")
        cell.font = font(color="CC0000")
        ws.merge_cells(f"A4:{get_column_letter(len(HEADERS))}4")

    filename = os.path.join(OUTPUT_DIR, f"{company.lower().replace(' ', '_')}_jobs_live.xlsx")
    wb.save(filename)
    return filename

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    if RAPIDAPI_KEY == "YOUR_RAPIDAPI_KEY_HERE":
        print("ERROR: Set your RapidAPI key in RAPIDAPI_KEY or as env var RAPIDAPI_KEY")
        print("Get a free key at: https://rapidapi.com/letscrape-6bRBa3QguO5/api/jsearch")
        return

    print(f"Fetching jobs for {len(TARGETS)} companies...\n")
    for target in TARGETS:
        company = target["company"]
        try:
            print(f"  [{company}] Fetching...", end=" ", flush=True)
            jobs = fetch_jobs(target["query"], RESULTS_PER_COMPANY)
            filename = build_excel(company, jobs, target["color"], target["accent"])
            print(f"✓  {len(jobs)} jobs → {os.path.basename(filename)}")
        except requests.exceptions.HTTPError as e:
            print(f"✗  API error: {e}")
        except Exception as e:
            print(f"✗  Error: {e}")

    print("\nDone! Open the *_jobs_live.xlsx files to see your listings.")

if __name__ == "__main__":
    main()

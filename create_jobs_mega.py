from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Shared Styles ──────────────────────────────────────────────────────────
WHITE      = "FFFFFF"
DARK_BLUE  = "1F4E79"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hf(color="1F4E79"):
    return Font(name="Arial", bold=True, color=WHITE, size=10)

def hfill(color="1F4E79"):
    return PatternFill("solid", start_color=color)

def nf(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink(cell, url, label="View Job →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

def write_header_row(ws, row, headers, fill_color, text_color=WHITE):
    ws.row_dimensions[row].height = 22
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=text_color, size=10)
        cell.fill = PatternFill("solid", start_color=fill_color)
        cell.alignment = center()
        cell.border = border

def write_title(ws, title, fill_color, text_color=WHITE, cols=12, row=1, height=32):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, color=text_color, size=14)
    cell.alignment = center()
    cell.fill = PatternFill("solid", start_color=fill_color)

def write_data_row(ws, r, row_data, fill, link_col=None):
    ws.row_dimensions[r].height = 38
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.border = border
        cell.font = nf()
        cell.alignment = center() if c == 1 else left()
        if c == link_col and val and val.startswith("http"):
            hyperlink(cell, val)
            cell.fill = fill
        else:
            cell.value = val

def tier_fill(tier):
    colors = {"⭐⭐ PERFECT": "C6EFCE", "⭐ HIGH": "FFEB9C", "✅ GOOD": "DDEBF7", "🔵 RELEVANT": "F2F2F2"}
    return PatternFill("solid", start_color=colors.get(tier, "FFFFFF"))

def tier_font(tier):
    colors = {"⭐⭐ PERFECT": "375623", "⭐ HIGH": "7F6000", "✅ GOOD": "1F4E79", "🔵 RELEVANT": "595959"}
    return Font(name="Arial", bold=True, size=9, color=colors.get(tier, "000000"))

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER LIST (All 50+ Jobs)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "🎯 All 50+ Jobs"

COLS = 11
write_title(ws1, "🎯 MEGA JOB LIST — o9 · OMP · SC Planning · Solutions Consulting — 52 Roles (US, 2026)", "1F4E79", cols=COLS)
ws1.row_dimensions[2].height = 6
write_header_row(ws1, 3, ["#","Job Title","Company","Location","Type","Salary","Tier","Category","Link","Recruiter / HM","Notes"], "2E75B6")

# (row#, title, company, location, type, salary, tier, category, url, recruiter, notes)
ALL_JOBS = [
    # ─── o9 / OMP DIRECT ─────────────────────────────────────────────────────
    (1,  "Lead o9 Consultant ⭐",             "Infosys",              "Richardson, TX",        "On-site · FT", "Not listed",        "⭐⭐ PERFECT", "o9 Direct",
     "https://www.linkedin.com/jobs/view/4378154088/",  "Managed off LinkedIn",  "3 connections · Actively reviewing · Easy Apply"),
    (2,  "Senior o9 Consultant ⭐",           "Infosys",              "Richardson, TX",        "On-site · FT", "Not listed",        "⭐⭐ PERFECT", "o9 Direct",
     "https://www.linkedin.com/jobs/view/4375271313/",  "Managed off LinkedIn",  "Actively reviewing · Easy Apply"),
    (3,  "Principal O9 Consultant ⭐",        "Infosys",              "Raleigh, NC",           "On-site · FT", "Not listed",        "⭐⭐ PERFECT", "o9 Direct",
     "https://www.linkedin.com/jobs/view/4388512492/",  "Managed off LinkedIn",  "14 o9 alumni at Infosys · Easy Apply"),
    (4,  "O9 Solutions Consultant ⭐",        "InfoVision Inc.",       "New York, US",          "Remote · FT",  "Not listed",        "⭐⭐ PERFECT", "o9 Direct",
     "https://www.linkedin.com/jobs/view/4383145463/",  "Not listed",            "Easy Apply · Actively reviewing applicants"),
    (5,  "O9 Supply Chain Consultant (SCM) ⭐","InfoVision Inc.",      "United States",         "Remote · FT",  "Not listed",        "⭐⭐ PERFECT", "o9 Direct",
     "https://www.linkedin.com/jobs/view/4385958110/",  "Not listed",            "Easy Apply · Actively reviewing applicants"),
    (6,  "O9 Technical Consultant",           "Xaxis Solutions",      "Berkeley Heights, NJ",  "On-site · FT", "Not listed",        "⭐ HIGH",      "o9 Direct",
     "https://www.linkedin.com/jobs/view/4398494386/",  "Not listed",            "Easy Apply · Review time ~1 week"),
    # ─── DELOITTE ─────────────────────────────────────────────────────────────
    (7,  "SC Planning Manager ⭐⭐ (o9/OMP/BY/JDA)","Deloitte",        "Charlotte, NC+",        "Hybrid · FT",  "Not listed",        "⭐⭐ PERFECT", "Big 4",
     "https://apply.deloitte.com/careers/JobDetail/Supply-Chain-Planning-Manager-experience-with-o9-OMP-or-Blue-Yonder-JDA/119515",
     "Not listed",  "8+ yrs SC · o9/OMP/BY directly in title! · 50% travel"),
    (8,  "SC Planning Sr Consultant ⭐ (o9/OMP/BY)","Deloitte",        "Huntsville, AL",        "Hybrid · FT",  "$95K–$135K est.",   "⭐⭐ PERFECT", "Big 4",
     "https://www.linkedin.com/jobs/view/4373925993/",  "Not listed",            "5+ yrs SC · o9/OMP/BY in title! · S&OP/IBP"),
    (9,  "SC Planning Consultant (o9/OMP/BY)", "Deloitte",            "Huntsville, AL",        "Hybrid · FT",  "$67K–$111K est.",   "⭐ HIGH",      "Big 4",
     "https://www.linkedin.com/jobs/view/4373929959/",  "Not listed",            "3+ yrs SC · o9/OMP/BY in title! · 1 wk ago"),
    (10, "SCNO Sr Procurement Mgr ⭐",        "Deloitte",             "Charlotte, NC",         "Hybrid · FT",  "$159K–$293K 💰",    "⭐ HIGH",      "Big 4",
     "https://www.linkedin.com/jobs/view/4377149810/",  "Not listed",            "28 o9 alumni · Deadline 5/30/26 · 2 days ago"),
    (11, "SCNO Sr Procurement Mgr ⭐",        "Deloitte",             "Atlanta, GA",           "Hybrid · FT",  "$159K–$293K 💰",    "⭐ HIGH",      "Big 4",
     "https://www.linkedin.com/jobs/view/4377163250/",  "Not listed",            "28 o9 alumni · Deadline 5/30/26 · 2 days ago"),
    (12, "SC Sourcing Consulting Mgr",        "Deloitte",             "New York, NY",          "Hybrid · FT",  "$130K–$241K 💰",    "✅ GOOD",      "Big 4",
     "https://www.linkedin.com/jobs/view/4367928244/",  "Not listed",            "28 o9 alumni · Coupa/Ariba · 4 days ago"),
    (13, "SC Sourcing Consulting Mgr",        "Deloitte",             "Philadelphia, PA",      "Hybrid · FT",  "$130K–$241K 💰",    "✅ GOOD",      "Big 4",
     "https://www.linkedin.com/jobs/view/4367915410/",  "Not listed",            "28 o9 alumni · 4 days ago"),
    (14, "Oracle Cloud SC Manager",           "Deloitte",             "Jersey City, NJ",       "Hybrid · FT",  "Not listed",        "✅ GOOD",      "Big 4",
     "https://www.linkedin.com/jobs/view/4383839935/",  "Not listed",            "28 o9 alumni · Be an early applicant · 1 wk ago"),
    # ─── ACCENTURE ────────────────────────────────────────────────────────────
    (15, "SC Planning & Exec Sr Mgr ⭐⭐\n(O9/Kinaxis/BY/SAP IBP)",
     "Accenture", "USA (Multiple)",  "Hybrid · FT", "$140K–$200K est.", "⭐⭐ PERFECT", "Big 4",
     "https://www.accenture.com/us-en/careers/jobdetails?id=R00174637_en",
     "Not listed", "5+ yrs o9/Kinaxis/BY/SAP IBP · SC transformation · Location flexible"),
    (16, "SC Planning Mgmt Consultant ⭐⭐\n(o9/Kinaxis/SAP IBP/BY)",
     "Accenture", "Midwest / Chicago, IL", "Hybrid · FT", "$110K–$160K est.", "⭐⭐ PERFECT", "Big 4",
     "https://www.accenture.com/us-en/careers/jobdetails?id=R00209619_en",
     "Not listed", "2+ yrs o9/Kinaxis/SAP IBP/BY · SC transformation · Consulting exp"),
    (17, "Life Sciences SC Transformation Mgr ⭐\n(Kinaxis/SAP IBP/OMP)",
     "Accenture", "USA (Multiple)", "Hybrid · FT", "$120K–$180K est.", "⭐⭐ PERFECT", "Big 4",
     "https://www.accenture.com/us-en/careers/jobdetails?id=R00281859_en",
     "Not listed", "OMP + Kinaxis + SAP IBP · Life sciences focus"),
    (18, "Energy SC Planning Consulting Mgr", "Accenture",            "Austin, TX",            "Hybrid · FT",  "$120K–$170K est.", "⭐ HIGH",      "Big 4",
     "https://www.glassdoor.com/job-listing/supply-chain-and-operations-energy-supply-chain-planning-consulting-manager-accenture-JV_IC1139761_KO0,75_KE76,85.htm?jl=1010030762925",
     "Not listed", "IBP/o9/OMP/Kinaxis planning tech · Feb 2026 posting"),
    (19, "Energy SC Planning Consulting Mgr", "Accenture",            "Nashville, TN",         "Hybrid · FT",  "$120K–$170K est.", "⭐ HIGH",      "Big 4",
     "https://www.glassdoor.com/job-listing/supply-chain-and-operations-energy-supply-chain-planning-consulting-manager-accenture-JV_IC1144541_KO0,75_KE76,85.htm?jl=1010030762932",
     "Not listed", "IBP/o9/OMP/Kinaxis planning tech · Feb 2026 posting"),
    (20, "Energy SC Planning Consulting Mgr", "Accenture",            "Los Angeles, CA",       "Hybrid · FT",  "$130K–$180K est.", "⭐ HIGH",      "Big 4",
     "https://www.glassdoor.com/job-listing/supply-chain-and-operations-energy-supply-chain-planning-consulting-manager-accenture-JV_IC1146821_KO0,75_KE76,85.htm?jl=1010030762891",
     "Not listed", "IBP/o9/OMP/Kinaxis planning tech · LA market premium"),
    # ─── EY ───────────────────────────────────────────────────────────────────
    (21, "SC Tech - OMP Planning Mgr 🌍\n(Location OPEN)",
     "EY", "USA (Remote-Flexible)", "Hybrid · FT", "Not listed", "⭐⭐ PERFECT", "Big 4",
     "https://seasonalworks.labor.ny.gov/albany-ny/supply-chain-tech-omp-planning-manager-location-open/4AA121E9ECFB4969AB54DA59E1622059/job/",
     "Not listed", "OMP in job title! · Location OPEN · S&OP/IBP · SC planning transformation"),
    (22, "SC Tech - Blue Yonder Planning Mgr 🌍\n(Location OPEN)",
     "EY", "USA (Remote-Flexible)", "Hybrid · FT", "Not listed", "⭐ HIGH",      "Big 4",
     "https://www.linkedin.com/jobs/view/3173102187",
     "Not listed", "BY expertise required · Full lifecycle SC planning impl"),
    (23, "Tech & Arch - SC Technical Mgr 🌍\n(Location OPEN)",
     "EY", "USA (Remote-Flexible)", "Hybrid · FT", "Not listed", "⭐ HIGH",      "Big 4",
     "https://www.linkedin.com/jobs/view/3371223485",
     "Not listed", "10+ yrs SC tech · Blue Yonder/Kinaxis/Logility · Consulting"),
    (24, "SC Tech - Planning Sr Manager",     "EY",                   "Des Moines, IA",        "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "Big 4",
     "https://www.linkedin.com/jobs/view/3520240287",
     "Not listed", "BY/Kinaxis · 2–3 full lifecycle SC planning implementations"),
    # ─── KPMG ─────────────────────────────────────────────────────────────────
    (25, "Manager, SC Consulting –\nDemand & Supply Planning",
     "KPMG", "Dallas, TX", "Hybrid · FT", "$84K–$169K", "⭐ HIGH",      "Big 4",
     "https://kpmgcareers.dejobs.org/dallas-tx/manager-supply-chain-consulting-demand-supply-planning/4424E5A7628F45A799BE13BDD2F9249C/job/?vs=28",
     "Not listed", "SC planning transformation · S&OP/IBP focus"),
    (26, "Manager, SC Consulting – Supply Planning", "KPMG",          "Multiple US",           "Hybrid · FT",  "$84K–$169K",        "⭐ HIGH",      "Big 4",
     "https://us-jobs.kpmg.com/careers/JobDetail?jobId=54762&srcCat=Internet&specSrc=Indeed",
     "Not listed", "Demand & Supply planning · 96 KPMG SC jobs available"),
    # ─── CAPGEMINI ─────────────────────────────────────────────────────────────
    (27, "SAP IBP Supply & Response Lead Architect", "Capgemini",      "Atlanta, GA",           "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "Consulting",
     "https://www.linkedin.com/jobs/view/4236151912",
     "Not listed", "10+ yrs SC planning · 2-3 full IBP implementations · Lead architect"),
    (28, "SAP IBP Inventory Optimization Lead", "Capgemini",           "Houston, TX",           "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "Consulting",
     "https://www.ziprecruiter.com/c/Capgemini-North-America/Job/SAP-IBP-Inventory-Optimization-Lead/-in-Houston,TX?jid=f29c976b84a58f42",
     "Not listed", "10+ yrs SC planning · IBP inventory optimization specialist"),
    (29, "Demand & SC Planning Consultant",   "Capgemini",            "USA",                   "Hybrid · FT",  "Not listed",        "✅ GOOD",      "Consulting",
     "https://www.capgemini.com/jobs/332907-en_GB+sap_btp/",
     "Not listed", "SAP BTP/IBP · SC planning transformation"),
    (30, "Business Transformation Consultant –\nSC Planning",
     "IBM", "USA (Multiple)", "Hybrid · FT", "Not listed", "✅ GOOD",      "Consulting",
     "https://careers.ibm.com/en_US/careers/JobDetail/Business-Transformation-Consultant-Supply-Chain-Planning/88311",
     "Not listed", "SC planning transformation · IBP · Digital SC"),
    # ─── BLUE YONDER (Vendor) ─────────────────────────────────────────────────
    (31, "Sr Functional Solution Architect ⭐", "Blue Yonder",         "Coppell, TX (Remote)",  "Remote · FT",  "Not listed",        "⭐⭐ PERFECT", "SC Vendor",
     "https://www.linkedin.com/jobs/view/4394225134/",
     "Not listed",  "HIGH MATCH · Be an early applicant · Vendor direct"),
    (32, "Solution Architect - WMS",          "Blue Yonder",          "Coppell, TX (Remote)",  "Remote · FT",  "Not listed",        "⭐ HIGH",      "SC Vendor",
     "https://www.linkedin.com/jobs/view/4321921503/",
     "Network: Kevin + others",  "99 applicants · Responses managed off LinkedIn"),
    (33, "Blue Yonder Project Manager 🟢",    "Maven Workforce Inc.", "California",            "Remote",        "Not listed",        "✅ GOOD",      "SC Vendor",
     "https://www.linkedin.com/jobs/view/4400574301/",
     "Not listed",  "Easy Apply · 1 hr ago · Blue Yonder project"),
    # ─── KINAXIS (Vendor) ─────────────────────────────────────────────────────
    (34, "Solutions Architect – Kinaxis ⭐",  "Genpact",              "USA (Remote)",          "Remote · FT",  "$150K–$180K 💰",    "⭐⭐ PERFECT", "SC Vendor",
     "https://www.linkedin.com/jobs/view/4330227649/",
     "Vipul Gupta\nGlobal TA @ Genpact",  "Easy Apply · Actively reviewing · HIGH MATCH"),
    (35, "Sr Solution Consultant – Kinaxis 🟢","Pacer Group",          "USA (Remote)",          "Remote · Contract","$82–$92/hr 💰",  "⭐ HIGH",      "SC Vendor",
     "https://www.linkedin.com/jobs/view/4393864119/",
     "Not listed",  "~$170K–$191K annualized · Easy Apply · 45 applicants"),
    (36, "Platform Architect – Kinaxis",      "Kinaxis (Direct)",     "USA (Remote)",          "Remote · FT",  "Not listed",        "⭐ HIGH",      "SC Vendor",
     "https://www.linkedin.com/jobs/view/4398156298/",
     "Not listed",  "Direct from Kinaxis · 46 mins ago at time of search"),
    # ─── SAP IBP ──────────────────────────────────────────────────────────────
    (37, "Senior SAP IBP Consultant 🚀",      "SELECCIÓN Consulting", "East Brunswick, NJ",    "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "SC Vendor",
     "https://www.linkedin.com/jobs/view/4400587273/",
     "Krishna S. (visible)",  "10 mins old · 0 applicants! · Easy Apply · APPLY FIRST"),
    (38, "SAP IBP Consultant",                "via Dice",             "USA (Remote)",          "Remote",        "Not listed",        "✅ GOOD",      "SC Vendor",
     "https://www.linkedin.com/jobs/view/4398118825/",
     "Not listed",  "Easy Apply"),
    # ─── KÖRBER ───────────────────────────────────────────────────────────────
    (39, "Senior Solution Designer ⭐",       "Körber Supply Chain",  "USA (Remote)",          "Remote · FT",  "Not listed",        "⭐⭐ PERFECT", "SC Vendor",
     "https://www.linkedin.com/jobs/view/4382197450/",
     "Michael Unser\nDir Solution Design, Körber NA",  "HIGH MATCH · SC platform consulting"),
    # ─── INDUSTRY PLANNING ROLES ─────────────────────────────────────────────
    (40, "Sr Supply Demand Planner ⭐",       "Intuitive",            "Sunnyvale, CA",         "On-site · FT", "$118.7K–$170.7K 💰","⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4345274465/",
     "Not listed",  "2 o9 alumni work here · Responses managed off LinkedIn"),
    (41, "Sr Supply & Demand Planner ⭐",     "Kohler Co.",           "Kohler, WI",            "On-site · FT", "Not listed",        "⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4366386259/",
     "Taylor Magri, CRD\nRecruiter @ Kohler",  "o9 alumni in network · Message recruiter!"),
    (42, "Senior Demand Planner ⭐",          "Intel",                "Santa Clara, CA",       "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4394482391/",
     "Rutvi (in network)",  "381 alumni · Mentions Blue Yonder! · 4 days ago"),
    (43, "Senior Demand Planner ⭐",          "Intel",                "Folsom, CA",            "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4394478440/",
     "Rutvi (in network)",  "381 alumni · Mentions Blue Yonder! · 4 days ago"),
    (44, "Senior Supply Chain Planner ⭐",    "Micron Technology",    "Boise, ID",             "On-site · FT", "Not listed",        "⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4386872480/",
     "Not listed",  "50 school alumni · Semiconductor SC planning"),
    (45, "Sr Analyst, Supply Chain Planning ⭐","Analog Devices",     "Wilmington, MA",        "Hybrid · FT",  "Not listed",        "⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4371933606/",
     "Not listed",  "140 school alumni · IBP/S&OP focus"),
    (46, "Supply Chain Analysts – Sr ⭐",     "Cummins Inc.",         "Atlanta, GA",           "On-site · FT", "Not listed",        "⭐ HIGH",      "Industry",
     "https://www.linkedin.com/jobs/view/4393461958/",
     "Not listed",  "73 school alumni · SC analytics & planning"),
    (47, "Senior S&OP Planner",               "Amazon Web Services",  "Seattle, WA",           "On-site · FT", "Not listed",        "✅ GOOD",      "Industry",
     "https://www.linkedin.com/jobs/view/4381119799/",
     "Not listed",  "11 o9 alumni work here · AWS infrastructure S&OP"),
    (48, "Sr Demand Planner (Ecomm & Wholesale)","Velvet Caviar",      "New York, US",          "Remote",        "$90K–$120K",        "✅ GOOD",      "Industry",
     "https://www.linkedin.com/jobs/view/4391579652/",
     "Not listed",  "Easy Apply · Actively reviewing applicants"),
    (49, "SC Solutions Architect",            "Texas Instruments",    "Dallas, TX",            "On-site · FT", "Not listed",        "✅ GOOD",      "Industry",
     "https://www.linkedin.com/jobs/view/4382493831/",
     "Not listed",  "SC planning + tech architecture blend"),
    (50, "Sr Supply Chain Planner",           "Sr Supply & Demand Planner ⭐",  "Kohler Co.", "Kohler, WI",  "On-site · FT", "Not listed", "⭐ HIGH", "Industry",
     "https://www.linkedin.com/jobs/view/4366386259/",
     "Taylor Magri, CRD",  "o9 alumni in network"),
    # ─── EY ACTIVE ROLES ─────────────────────────────────────────────────────
    (45, "SC Tech - OMP Planning Mgr ⭐⭐\n(Location OPEN)",
     "EY", "USA (Remote-Flexible)", "Hybrid · FT", "Not listed", "⭐⭐ PERFECT", "Big 4",
     "https://www.linkedin.com/jobs/view/3485225708",
     "Not listed", "OMP in job title! · Location OPEN · S&OP/IBP transformation"),
    (46, "SAP - SC Planning IBP PPDS Mgr ⭐\n(Location OPEN)",
     "EY", "Chicago, IL (Hybrid)", "Hybrid · FT", "$141K–$258K 💰", "⭐ HIGH", "Big 4",
     "https://careers.ey.com/ey/job/Chicago-SAP-Supply-Chain-Planning-IBP-PPDS-Manager-Consulting-Location-OPEN-IL-60606/1108397501/",
     "Not listed", "5+ yrs SAP IBP/APO · S&OP/IBP · $141K–$258K · Location OPEN"),
    (47, "SC Health - Perf Improvement Mgr",  "EY", "Nashville, TN (On-site)", "On-site · FT", "Not listed", "✅ GOOD", "Big 4",
     "https://www.linkedin.com/jobs/view/4393135679/",
     "Not listed", "30 o9 alumni · SC performance improvement consulting"),
    (48, "SC & Ops Senior Consultant",        "EY", "Nashville, TN (On-site)", "On-site · FT", "Not listed", "✅ GOOD", "Big 4",
     "https://www.linkedin.com/jobs/view/4397937369/",
     "Not listed", "30 o9 alumni · Be an early applicant! · SC & Operations"),
    (49, "SC Procurement Mgr – Multiple\n(Positions 1696063)",
     "EY", "Atlanta, GA (On-site)", "On-site · FT", "Not listed", "✅ GOOD", "Big 4",
     "https://www.linkedin.com/jobs/view/4391469844/",
     "Not listed", "30 o9 alumni · Multiple positions open · SC procurement & planning"),
    # ─── KPMG ACTIVE ROLES ───────────────────────────────────────────────────
    (50, "Manager, SAP Supply Chain (IBP)",   "KPMG", "Multiple US", "Hybrid · FT", "Not listed", "⭐ HIGH", "Big 4",
     "https://www.kpmguscareers.com/jobdetail/?jobId=106915",
     "Not listed", "SAP IBP implementation · SC planning background · 6000+ SAP consultants"),
    # ─── PwC ACTIVE ROLES ────────────────────────────────────────────────────
    (51, "SAP SC - IBP Solution Architect\nSr. Manager ⭐",
     "PwC", "New York, NY + 11 locations", "Hybrid · FT", "$130K–$256K 💰", "⭐⭐ PERFECT", "Big 4",
     "https://jobs.us.pwc.com/job/new-york/sap-supply-chain-ibp-solution-architect-sr-manager/932/67751281280",
     "Not listed", "IBP Demand/Supply/Inventory/S&OP · Solution Architect Sr Mgr · $130K–$256K!"),
    (52, "SAP IBP/TM/PP Senior Manager ⭐",   "PwC", "Atlanta, GA (Hybrid)", "Hybrid · FT", "$130K–$256K 💰", "⭐ HIGH", "Big 4",
     "https://jobs.us.pwc.com/job/atlanta/sap-ibp-tm-pp-senior-manager/932/78996929616",
     "Not listed", "SAP IBP + Transportation Mgmt + PP · Senior Manager level"),
    (53, "SAP IBP Senior Manager ⭐",         "PwC", "Dallas, TX (Hybrid)", "Hybrid · FT", "$130K–$256K 💰", "⭐ HIGH", "Big 4",
     "https://jobs.us.pwc.com/job/dallas/sap-integrated-business-planning-ibp-senior-manager/932/92430021024",
     "Not listed", "SAP IBP focused · $130K–$256K · Dallas location"),
    # ─── NTT DATA ────────────────────────────────────────────────────────────
    (54, "o9 Consulting Manager ⭐⭐",         "NTT DATA North America", "Georgia (Remote)", "Remote · FT", "Not listed", "⭐⭐ PERFECT", "Consulting",
     "https://www.linkedin.com/jobs/view/4390662733/",
     "Not listed", "o9 IN TITLE! · Remote · NTT DATA SC Consulting · 13 alumni · Be early applicant"),
    # ─── STAFFING / CONTRACT ─────────────────────────────────────────────────
    (55, "SC Planning Solutions Consultant ⭐⭐\n(BY/Kinaxis/OMP/O9)",
     "Accenture", "St. Louis, MO / US", "Hybrid · FT", "Not listed", "⭐⭐ PERFECT", "Big 4",
     "https://success.recruitmilitary.com/job/35788570/fromstatic",
     "Not listed", "BY, Kinaxis, OMP, O9 all in title! · Veterans preferred"),
    (56, "O9 Supply Chain Technical Lead",    "Net2Source Inc.",       "Austin, TX",            "On-site · Contract","Not listed",  "⭐⭐ PERFECT", "Staffing",
     "https://www.linkedin.com/jobs/view/o9-supply-chain-technical-lead-at-net2source-inc-3606649798",
     "Not listed",  "7+ yrs o9 · Demand/Supply/S&OP · Technical + functional"),
    (57, "Consultant – SC & Ops",             "Rios Partners",         "Arlington, VA",         "On-site · FT", "$110K–$130K 💰",    "✅ GOOD",      "Consulting",
     "https://www.linkedin.com/jobs/view/4385530240/",
     "Not listed",  "SC consulting boutique · $110K–$130K base"),
]

# Deduplicate by row # (fix row 50 duplicate)
seen = set()
unique_jobs = []
for j in ALL_JOBS:
    key = j[8]  # URL as unique key
    if key not in seen:
        seen.add(key)
        unique_jobs.append(j)

# Re-number
unique_jobs = [(i+1,) + j[1:] for i, j in enumerate(unique_jobs)]

COL_HEADERS = ["#","Job Title","Company","Location","Type","Salary","Tier","Category","Link","Recruiter / HM","Notes"]
COL_WIDTHS  = [4,  36,         18,        18,         14,    16,      12,    10,         14,    28,              38]

# Category color bands for zebra
CAT_COLORS = {
    "o9 Direct":  "FFF2CC",
    "Big 4":      "DEEAF1",
    "Consulting": "E2EFDA",
    "SC Vendor":  "FCE4D6",
    "Industry":   "F2F2F2",
    "Staffing":   "EDE7F6",
}

for i, job in enumerate(unique_jobs):
    r = i + 4
    cat = job[7]
    base_color = CAT_COLORS.get(cat, "FFFFFF")
    # Slightly alternate lightness per row
    fill = PatternFill("solid", start_color=base_color)
    ws1.row_dimensions[r].height = 40

    for c, val in enumerate(job, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 7:  # Tier
            cell.fill = tier_fill(str(val))
            cell.font  = tier_font(str(val))
            cell.alignment = center()
            cell.value = val
        elif c == 9:  # Link
            if val and val.startswith("http"):
                hyperlink(cell, val)
                cell.fill = fill
            else:
                cell.value = val
                cell.font  = nf()
                cell.alignment = left()
        else:
            cell.value = val
            cell.font  = nf()
            cell.alignment = center() if c == 1 else left()

ws1.freeze_panes = "A4"
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Legend row at bottom ───────────────────────────────────────────────────
legend_row = len(unique_jobs) + 5
ws1.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=COLS)
leg = ws1.cell(row=legend_row, column=1,
    value="🟡 o9 Direct   🔵 Big 4 / Consulting   🟠 SC Vendor   ⚪ Industry   ⭐⭐PERFECT = o9/OMP in job title   ⭐HIGH = SC planning tech   ✅GOOD = SC consulting")
leg.font = Font(name="Arial", italic=True, size=9, color="595959")
leg.alignment = center()
leg.fill = PatternFill("solid", start_color="F2F2F2")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — o9 / OMP DIRECT
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("⭐ o9 & OMP Direct")
write_title(ws2, "⭐ o9 / OMP Direct Roles — US Job Listings 2026", "D4A017", text_color="FFFFFF", cols=10)
ws2.row_dimensions[2].height = 6
write_header_row(ws2, 3, ["#","Job Title","Company","Location","Type","Salary","Match","Link","Recruiter / HM","Notes"], "7F6000", text_color="FFFFFF")

o9_jobs = [j for j in unique_jobs if j[7] in ("o9 Direct",) or "PERFECT" in str(j[6])]
for i, job in enumerate(o9_jobs):
    r = i + 4
    fill = PatternFill("solid", start_color="FFF9C4" if i % 2 == 0 else "FFFDE7")
    row_data = (job[0], job[1], job[2], job[3], job[4], job[5], job[6], job[8], job[9], job[10])
    ws2.row_dimensions[r].height = 40
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.fill = fill; cell.border = border
        if c == 7:
            cell.fill = tier_fill(str(val)); cell.font = tier_font(str(val)); cell.alignment = center(); cell.value = val
        elif c == 8 and val and val.startswith("http"):
            hyperlink(cell, val); cell.fill = fill
        else:
            cell.value = val; cell.font = nf(); cell.alignment = center() if c == 1 else left()
ws2.freeze_panes = "A4"
for i, w in enumerate([4,36,18,18,14,16,12,14,28,38], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BIG 4 / CONSULTING
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🏢 Big 4 & Consulting")
write_title(ws3, "🏢 Big 4 & Consulting Firm SC Planning Roles — 2026", "1F4E79", cols=10)
ws3.row_dimensions[2].height = 6
write_header_row(ws3, 3, ["#","Job Title","Company","Location","Type","Salary","Match","Link","Deadline","Notes"], "2E75B6")

consulting_jobs = [j for j in unique_jobs if j[7] in ("Big 4", "Consulting")]
for i, job in enumerate(consulting_jobs):
    r = i + 4
    fill = PatternFill("solid", start_color="DEEAF1" if i % 2 == 0 else "FFFFFF")
    row_data = (job[0], job[1], job[2], job[3], job[4], job[5], job[6], job[8], "Open", job[10])
    ws3.row_dimensions[r].height = 42
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c)
        cell.fill = fill; cell.border = border
        if c == 7:
            cell.fill = tier_fill(str(val)); cell.font = tier_font(str(val)); cell.alignment = center(); cell.value = val
        elif c == 8 and val and val.startswith("http"):
            hyperlink(cell, val); cell.fill = fill
        else:
            cell.value = val; cell.font = nf(); cell.alignment = center() if c == 1 else left()
ws3.freeze_panes = "A4"
for i, w in enumerate([4,38,14,18,14,18,12,14,14,42], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SC TECH VENDORS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🔧 SC Tech Vendors")
write_title(ws4, "🔧 SC Tech Vendor Roles — Blue Yonder · Kinaxis · SAP IBP · Körber — 2026", "C00000", cols=10)
ws4.row_dimensions[2].height = 6
write_header_row(ws4, 3, ["#","Job Title","Company","Location","Type","Salary","Match","Link","Recruiter / HM","Notes"], "C00000")

vendor_jobs = [j for j in unique_jobs if j[7] in ("SC Vendor","Staffing")]
for i, job in enumerate(vendor_jobs):
    r = i + 4
    fill = PatternFill("solid", start_color="FCE4D6" if i % 2 == 0 else "FFFFFF")
    row_data = (job[0], job[1], job[2], job[3], job[4], job[5], job[6], job[8], job[9], job[10])
    ws4.row_dimensions[r].height = 40
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c)
        cell.fill = fill; cell.border = border
        if c == 7:
            cell.fill = tier_fill(str(val)); cell.font = tier_font(str(val)); cell.alignment = center(); cell.value = val
        elif c == 8 and val and val.startswith("http"):
            hyperlink(cell, val); cell.fill = fill
        else:
            cell.value = val; cell.font = nf(); cell.alignment = center() if c == 1 else left()
ws4.freeze_panes = "A4"
for i, w in enumerate([4,36,18,18,14,16,12,14,28,38], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — INDUSTRY PLANNING ROLES
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🏭 Industry Planning")
write_title(ws5, "🏭 Industry Planning Roles — Demand · Supply · S&OP — 2026", "375623", cols=10)
ws5.row_dimensions[2].height = 6
write_header_row(ws5, 3, ["#","Job Title","Company","Location","Type","Salary","Match","Link","Recruiter / HM","Notes"], "375623")

industry_jobs = [j for j in unique_jobs if j[7] == "Industry"]
for i, job in enumerate(industry_jobs):
    r = i + 4
    fill = PatternFill("solid", start_color="E2EFDA" if i % 2 == 0 else "FFFFFF")
    row_data = (job[0], job[1], job[2], job[3], job[4], job[5], job[6], job[8], job[9], job[10])
    ws5.row_dimensions[r].height = 40
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c)
        cell.fill = fill; cell.border = border
        if c == 7:
            cell.fill = tier_fill(str(val)); cell.font = tier_font(str(val)); cell.alignment = center(); cell.value = val
        elif c == 8 and val and val.startswith("http"):
            hyperlink(cell, val); cell.fill = fill
        else:
            cell.value = val; cell.font = nf(); cell.alignment = center() if c == 1 else left()
ws5.freeze_panes = "A4"
for i, w in enumerate([4,34,20,18,14,18,12,14,28,40], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
out = "C:/Users/amolp/Prometheus/job_listings_mega.xlsx"
wb.save(out)
total = len(unique_jobs)
print(f"Done! job_listings_mega.xlsx saved with {total} jobs across 5 sheets.")

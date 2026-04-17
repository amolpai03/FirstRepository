from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bot = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="404040"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hyperlink(cell, url, label="View Job →"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Amazon brand palette
AMZ_ORANGE  = "FF9900"   # Amazon signature orange
AMZ_DARK    = "232F3E"   # Amazon dark navy
AMZ_BLUE    = "146EB4"   # Amazon blue
AMZ_L_ORG  = "FFF3E0"   # Light orange tint
AMZ_L_BLUE = "E8F4FD"   # Light blue tint
AMZ_L_GRN  = "E8F5E9"   # Light green tint
AMZ_L_YLW  = "FFFDE7"   # Light yellow tint
AMZ_L_PURP = "F3E5F5"   # Light purple tint

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — TOP 5 AMAZON JOBS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Amazon Top 5 Jobs"

ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "AMAZON — Top 5 Supply Chain & Planning Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=AMZ_DARK)
t.border = border

ws1.row_dimensions[2].height = 5
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=AMZ_ORANGE)

ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "39 o9 alumni at Amazon  ·  994 school alumni  ·  1,639 open SC/Planning jobs at Amazon US  ·  LinkedIn Company ID: 1586"
s.font = Font(name="Arial", italic=True, size=9, color=AMZ_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color="FFF8EE")
s.border = border

HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match",
           "Status / Posted", "Experience Req.", "Key Skills & Tools", "Link", "Network Edge", "Why Apply / Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AMZ_DARK)
    cell.alignment = center()
    cell.border = thick_bot

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Sr. Technical PM — Supply Chain Planning\n(Robotics Supply Chain PMO)",
        "location": "Boston, MA\nor Seattle, WA\n(On-site)",
        "type": "On-site · Full-time",
        "salary": "$148,700 – $201,200\n(highest of top 5!)",
        "match": "PERFECT MATCH",
        "match_bg": "C6EFCE", "match_fg": "375623",
        "status": "1 wk ago\nBe an early applicant!\n39 o9 alumni",
        "exp": "5+ yrs SC Planning ops\n5+ yrs planning app impl.\n5+ yrs capability frameworks",
        "skills": "KINAXIS — explicitly required!\nOracle · SAP ERP · PLM\nS&OP · Demand Planning\nSupply Planning · Capability roadmaps\nPMP / APICS preferred",
        "url": "https://www.linkedin.com/jobs/view/4393981588/",
        "network": "39 o9 Solutions\nalumni at Amazon",
        "why": "KINAXIS IS IN THE JD — your #1 tool is explicitly listed. 'Kinaxis supply planning experience' = preferred qual. S&OP + Demand + Supply Planning capability framework = your daily work. $148K-$201K is exceptional. Early applicant window OPEN — apply today.",
        "fill": AMZ_L_GRN,
    },
    {
        "num": 2,
        "title": "Supply Chain Manager,\nSales & Operations Planning",
        "location": "Austin, TX  /  Bellevue, WA\nor Tempe, AZ\n(On-site)",
        "type": "On-site · Full-time",
        "salary": "$100,000 – $150,000\n(est. for Austin/Bellevue)",
        "match": "PERFECT MATCH",
        "match_bg": "C6EFCE", "match_fg": "375623",
        "status": "4 days ago\nBe an early applicant!\nMultiple locations",
        "exp": "3+ yrs program mgmt\n2+ yrs supply chain\nBachelor's degree",
        "skills": "S&OP — in the title!\nNetwork capacity planning\nSupply & demand balancing\nExcel (Pivot Tables, VLOOKUP)\nSQL · Data-driven decisions\nStakeholder management",
        "url": "https://www.linkedin.com/jobs/view/4399479658/",
        "network": "39 o9 alumni\n994 school alumni\nat Amazon",
        "why": "S&OP IS your core domain — this is literally your job title background. 3 US locations = 3x application opportunities. 4 days old + early applicant = low competition window. Your 7 yrs S&OP experience far exceeds the 3-yr minimum — you'll stand out instantly.",
        "fill": AMZ_L_ORG,
    },
    {
        "num": 3,
        "title": "Senior Planning Manager:\nGlobal Supply & Inventory Planning",
        "location": "Sunnyvale, CA\n(On-site)",
        "type": "On-site · Full-time",
        "salary": "$133,300 – $185,000",
        "match": "HIGH MATCH",
        "match_bg": "FFEB9C", "match_fg": "7F6000",
        "status": "Recent posting\n39 o9 alumni",
        "exp": "Bachelor's + 7+ yrs\nOR 5+ yrs Excel\n+ SQL experience",
        "skills": "Global supply & inventory planning\nDemand planning · Production scheduling\nInventory allocation across regions\nRisk mitigation · Data analysis\nSQL · AI/ML exposure preferred",
        "url": "https://www.linkedin.com/jobs/view/4393527651/",
        "network": "39 o9 Solutions\nalumni at Amazon",
        "why": "Amazon Devices Accessories = complex multi-region planning — directly maps to your demand/supply planning background. $133K-$185K. 7+ yrs requirement = your seniority band. Preferred qual includes AI/ML exposure — highlight any GenAI or analytics work.",
        "fill": AMZ_L_BLUE,
    },
    {
        "num": 4,
        "title": "Sr. Supply Chain PM,\nHub Delivery",
        "location": "New York, NY\n(On-site)",
        "type": "On-site · Full-time",
        "salary": "$104,000 – $181,000",
        "match": "HIGH MATCH",
        "match_bg": "FFEB9C", "match_fg": "7F6000",
        "status": "1 wk ago\nBe an early applicant!\n994 school alumni",
        "exp": "Senior PM level\n5+ yrs SC planning\nStrategic + tactical",
        "skills": "SC strategic planning\nHub Delivery network expansion\nData analysis & modeling\nCross-functional leadership\nAmbiguous problem solving\nExecutive communication",
        "url": "https://www.linkedin.com/jobs/view/4393989232/",
        "network": "994 school alumni\nat Amazon (NYC)",
        "why": "Early applicant window open — go now! Hub Delivery = last-mile logistics SC planning. Your planning background (S&OP, capacity, demand) translates directly. NYC = one of Amazon's fastest-growing delivery markets. 994 school alumni = massive referral network available.",
        "fill": AMZ_L_YLW,
    },
    {
        "num": 5,
        "title": "Supply Chain Manager,\nSales & Operations Planning\n(Consumables division)",
        "location": "New York, NY\n(On-site)",
        "type": "On-site · Full-time",
        "salary": "$100,000 – $150,000\n(estimated NYC rate)",
        "match": "STRONG MATCH",
        "match_bg": "DDEBF7", "match_fg": "1F4E79",
        "status": "Viewed recently\n994 school alumni",
        "exp": "3+ yrs program mgmt\n2+ yrs supply chain\nBachelor's degree",
        "skills": "S&OP planning ownership\nInventory optimization\nTransportation & logistics\nLabor capacity management\nKPI metrics & reporting\nCross-functional stakeholder mgmt",
        "url": "https://www.linkedin.com/jobs/view/4398951070/",
        "network": "994 school alumni\nat Amazon (NYC)",
        "why": "Same S&OP role in the Consumables vertical — apply to BOTH #2 and #5 for dual coverage within Amazon. Consumables = high-volume, fast-moving goods = sophisticated S&OP cycles. NYC-based. 994 school alumni = highest referral probability of all 5 roles.",
        "fill": AMZ_L_PURP,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 105
    base_fill = PatternFill("solid", start_color=job["fill"])

    vals = [job["num"], job["title"], job["location"], job["type"], job["salary"],
            job["match"], job["status"], job["exp"], job["skills"],
            job["url"], job["network"], job["why"]]

    for c, val in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = base_fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=AMZ_ORANGE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=AMZ_DARK)
            cell.alignment = left()
        elif c == 5:
            cell.value = val
            cell.font = nf(bold=True, size=9, color="155724")
            cell.alignment = left()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_early = "early applicant" in str(val).lower()
            cell.font = Font(name="Arial", bold=is_early, size=9,
                           color="C00000" if is_early else "595959")
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = base_fill
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

COL_WIDTHS = [4, 30, 20, 14, 20, 14, 18, 22, 34, 12, 18, 46]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RESUME TAILORING
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING — What to Highlight for Each Amazon Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=AMZ_DARK)

ws2.row_dimensions[2].height = 5
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=AMZ_ORANGE)

HEADERS2 = ["Role", "Must Highlight in Resume", "Amazon Keywords to Use",
            "Lead With This", "Potential Gap", "Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AMZ_DARK)
    cell.alignment = center()
    cell.border = thick_bot

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Sr. Technical PM\nSupply Chain Planning\n(Robotics PMO)",
        "• Kinaxis implementation projects\n• SC planning capability roadmaps\n• S&OP / Demand / Supply planning\n• Cross-functional program leadership\n• ERP integration experience",
        "Kinaxis, supply chain planning, S&OP, demand planning, supply planning, capability framework, roadmap, ERP, Oracle, SAP, implementation, program management, cross-functional",
        "Lead with specific Kinaxis project: 'Led end-to-end Kinaxis implementation for [client], built planning capability framework covering S&OP, demand, and supply planning for $Xbn supply chain'",
        "Amazon Robotics is hardware/manufacturing focused — different from consumer goods SC. Role requires ERP depth (Oracle/SAP) beyond planning tools alone",
        "Frame your o9/Kinaxis work as 'large-scale enterprise planning system implementation'. Highlight any manufacturing or ops-side clients. PMP/APICS cert = major advantage here. Apply IMMEDIATELY — early window open",
        "APPLY NOW!\nPERFECT MATCH\n$148K-$201K"
    ),
    (
        "Supply Chain Manager\nSales & Operations Planning\n(Austin / Bellevue / Tempe)",
        "• S&OP process ownership\n• Demand / supply balancing\n• Inventory optimization KPIs\n• Network capacity planning\n• Data-driven decision making",
        "S&OP, sales and operations planning, network capacity, supply chain, inventory, demand planning, fulfillment, distribution, process improvement, SQL, Excel, stakeholder management",
        "Lead with S&OP program ownership: 'Managed end-to-end S&OP cycle for [X] product categories across [Y] locations, reducing inventory excess by Z%'. Show Amazon-scale thinking",
        "3 yr min req — you have 7 yrs (massive overqualification — frame as depth, not excess). Amazon's S&OP = fulfillment network, not product S&OP",
        "Tailor resume to Amazon fulfillment language: 'network capacity', 'fulfillment efficiency', 'inventory flow'. Apply to ALL 3 locations for 3x chances. Use 'Be an early applicant' advantage now",
        "APPLY ALL 3\nLOCATIONS!\n4 days old"
    ),
    (
        "Senior Planning Manager\nGlobal Supply &\nInventory Planning",
        "• Multi-region inventory planning\n• Production planning experience\n• Demand change analysis\n• Risk mitigation strategies\n• Supplier production coordination",
        "Supply planning, inventory allocation, production planning, demand planning, order fulfillment, risk mitigation, data-driven, cross-functional, NPI, end-of-life planning, SQL",
        "Lead with your global/multi-region planning scope. Quantify: # SKUs managed, $ inventory value, forecast accuracy %. Amazon Devices = consumer electronics SC = show any tech/electronics planning",
        "Devices Accessories has very short product lifecycles (NPI to EOL in months). AI/ML preferred qual — Amazon is heavy on data science",
        "Highlight any product lifecycle planning (NPI or EOL). Mention any SQL or Python work. Note: 'AI systems evaluation' is preferred — even discussing AI tools in planning (o9's AI features) counts here",
        "HIGH MATCH\n$133K-$185K"
    ),
    (
        "Sr. Supply Chain PM\nHub Delivery\n(New York)",
        "• SC strategic planning\n• Last-mile / delivery logistics\n• Large-scale program management\n• Ambiguous problem solving\n• Executive stakeholder comms",
        "Supply chain, program management, hub delivery, last-mile, logistics, strategic planning, data analysis, network expansion, capacity, stakeholder management, cross-functional",
        "Lead with your largest strategic SC program. Show how you tackled ambiguous SC problems and drove results. Hub Delivery = new territory — show your 'think big' and 'invent and simplify' (Amazon LPs)",
        "Hub Delivery is Amazon-specific last-mile — no direct consulting analogy. NYC on-site required",
        "Map your work to Amazon Leadership Principles: Customer Obsession (forecast accuracy), Think Big (network expansion), Dive Deep (data analysis), Deliver Results (KPIs). Use LP language in resume bullets",
        "EARLY BIRD\nApply NOW!\n$104K-$181K"
    ),
    (
        "Supply Chain Manager\nS&OP — Consumables\n(New York)",
        "• S&OP planning ownership\n• High-volume SKU management\n• Inventory & labor capacity\n• Transportation optimization\n• Cross-team coordination",
        "S&OP, supply chain, inventory optimization, consumables, transportation, labor capacity, fulfillment, distribution, planning, metrics, KPI, stakeholder, data analysis",
        "Lead with S&OP experience in high-volume / fast-moving goods. Any FMCG or consumer goods SC exposure is a huge plus. Show breadth: demand, supply, inventory, logistics all in one role",
        "Consumables = Amazon Fresh/grocery/everyday items — high velocity, tight margins. Role may overlap significantly with Job #2",
        "Apply to BOTH #2 and #5 simultaneously — different teams, same Amazon LP culture. 994 school alumni = highest referral density. LinkedIn-message 2-3 NCSU alumni at Amazon for a referral code",
        "DUAL APPLY\nWith Job #2\n994 alumni!"
    ),
]

ROW_FILLS  = [AMZ_L_GRN, AMZ_L_ORG, AMZ_L_BLUE, AMZ_L_YLW, AMZ_L_PURP]
PRI_COLORS = [("C6EFCE","375623"),("C6EFCE","375623"),("FFEB9C","7F6000"),("FFEB9C","7F6000"),("DDEBF7","1F4E79")]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 115
    fill = PatternFill("solid", start_color=ROW_FILLS[i])
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=AMZ_DARK)
            cell.fill = fill; cell.alignment = left()
        elif c == 7:
            bg, fg = PRI_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill; cell.alignment = left()

for i, w in enumerate([26, 40, 40, 44, 32, 40, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — AMAZON QUICK FACTS + LP CHEAT SHEET
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Amazon Facts + LP Guide")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "AMAZON — Company Facts & Leadership Principles Interview Cheat Sheet"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=AMZ_DARK)
t3.border = border

ws3.row_dimensions[2].height = 5
ws3.merge_cells("A2:D2")
ws3["A2"].fill = PatternFill("solid", start_color=AMZ_ORANGE)

FACTS = [
    ("── COMPANY OVERVIEW ──", ""),
    ("Company",       "Amazon.com, Inc. — World's largest e-commerce + cloud (AWS) + logistics company"),
    ("HQ",            "Seattle, WA (also major hubs: NYC, Austin TX, Nashville TN, Bellevue WA, Boston MA)"),
    ("Revenue",       "$638B (2024) · Fortune #2 globally · NASDAQ: AMZN · ~1.5M employees worldwide"),
    ("SC Scale",      "Largest private SC in the world: 200+ fulfillment centers, 100K+ delivery vehicles, same-day delivery"),
    ("Planning Tech", "Kinaxis (Robotics PMO), Oracle, SAP, internal tools, heavy SQL/Python/Databricks/AWS analytics"),
    ("o9 Alumni",     "39 o9 Solutions alumni at Amazon — strong network for referrals"),
    ("School Alumni", "994 school alumni at Amazon — highest alumni count across all companies searched"),
    ("Salary Range",  "Sr Technical PM: $148K-$201K | Sr Planning Mgr: $133K-$185K | Sr SC PM: $104K-$181K"),
    ("Work Style",    "All 5 roles are 100% ON-SITE. Amazon has a strict 5-day RTO (return-to-office) policy"),
    ("── AMAZON INTERVIEW STYLE ──", ""),
    ("Format",        "Behavioral interviews using STAR method mapped to Amazon's 16 Leadership Principles (LPs)"),
    ("Key LPs for SC","Customer Obsession · Dive Deep · Deliver Results · Invent & Simplify · Think Big · Ownership"),
    ("── TOP 6 LEADERSHIP PRINCIPLES FOR SC ROLES ──", ""),
    ("1. Customer Obsession",   "Start with customer needs → SC = on-time delivery, inventory availability, cost efficiency"),
    ("2. Dive Deep",            "Data-driven decisions → show SQL queries, metrics ownership, root cause analysis"),
    ("3. Deliver Results",      "Quantify EVERYTHING → 'Reduced forecast error by 18%', 'Saved $2M in inventory'"),
    ("4. Invent & Simplify",    "Process improvements → show how you simplified SC processes or automated planning"),
    ("5. Think Big",            "Strategic vision → show how you scaled programs, planned for 2-3 yr horizons"),
    ("6. Ownership",            "End-to-end accountability → you owned the S&OP cycle, not just contributed to it"),
    ("── AMAZON-SPECIFIC TIPS ──", ""),
    ("Resume Format",   "1 page preferred · 3-5 bullet points per role · Every bullet starts with action verb + metric"),
    ("Bar Raiser",      "Every loop has a Bar Raiser — an unbiased interviewer. Prepare extra LP examples"),
    ("S&OP Language",   "Amazon calls it 'Sales & Operations Planning' or 'network capacity management' — use their vocabulary"),
    ("Kinaxis Tip",     "Job #1 says 'Kinaxis supply chain experience' in preferred quals — mention Kinaxis in first 3 lines of resume"),
    ("Apply Strategy",  "Apply to Jobs #2 AND #5 simultaneously (same S&OP role, different teams) + Job #1 for highest ROI"),
]

ws3.row_dimensions[3].height = 22
for c, h in enumerate(["Category", "Details", "", ""], 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=AMZ_DARK)
    cell.alignment = center()
    cell.border = thick_bot

for i, (cat, detail) in enumerate(FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 26
    is_header = detail == ""
    if is_header:
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cell = ws3.cell(row=r, column=1, value=cat)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", start_color=AMZ_ORANGE)
        cell.alignment = center()
        cell.border = border
    else:
        alt_fill = PatternFill("solid", start_color="FFF8EE" if i % 2 == 0 else WHITE)
        c1 = ws3.cell(row=r, column=1, value=cat)
        c1.font = nf(bold=True, size=9, color=AMZ_DARK)
        c1.fill = alt_fill; c1.alignment = left(); c1.border = border
        ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ws3.cell(row=r, column=2, value=detail)
        c2.font = nf(size=9); c2.fill = alt_fill
        c2.alignment = left(); c2.border = border

ws3.freeze_panes = "A4"
for i, w in enumerate([26, 84, 1, 1], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
output = "C:/Users/amolp/Prometheus/amazon_top5_jobs.xlsx"
wb.save(output)
print("Done! amazon_top5_jobs.xlsx saved with 3 sheets.")

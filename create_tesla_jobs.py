from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="595959"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def hyperlink(cell, url, label="View Job ->"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Tesla brand colors
TESLA_RED      = "CC0000"
TESLA_DARK_RED = "8B0000"
TESLA_DARK     = "1B1B1B"
TESLA_SILVER   = "AAAAAA"
TESLA_LIGHT    = "F9F9F9"
TESLA_RED_LIGHT= "FFEBEE"
TESLA_GRAY     = "F0F0F0"
TESLA_CHARCOAL = "3D3D3D"

GREEN_BADGE  = "C6EFCE"
GREEN_DARK   = "375623"
YELLOW_BADGE = "FFEB9C"
YELLOW_DARK  = "7F6000"
BLUE_BADGE   = "DDEBF7"
BLUE_DARK    = "1F4E79"

# ============================================================
# SHEET 1 — TESLA TOP 5 JOBS
# ============================================================
ws1 = wb.active
ws1.title = "Tesla Top 5 Jobs"

ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "TESLA -- Top 5 Supply Chain & Planning Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=TESLA_DARK)
t.border = border

ws1.row_dimensions[2].height = 7
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=TESLA_RED)

ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "2 o9 alumni at Tesla  |  Job 2 posted 14 HOURS ago  |  Job 5 is 5 days old (early applicant)  |  49 targeted SC/Planning roles  |  Tesla LinkedIn Company ID: 15564"
s.font = Font(name="Arial", italic=True, size=9, color=TESLA_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color=TESLA_RED_LIGHT)
s.border = border

HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match Level",
           "Status / Alumni", "Posted", "Key Skills Required", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=TESLA_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Supply Chain Planning\nChannel Planning Manager",
        "location": "Fremont, CA\n(On-site)\nTesla SC HQ",
        "type": "Full-Time\nOn-site",
        "salary": "~$130,000 - $185,000\n(est. Tesla SC Mgr range)\n+ Stock + Bonus",
        "match": "PERFECT MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "2 o9 alumni at Tesla\nViewed\nFremont HQ",
        "posted": "Recently posted",
        "skills": "SC planning + channel planning\nDemand/supply planning process\nChannel/distribution planning\nS&OP or IBP methodology\nCross-functional coordination\nPlanning tools (o9/Kinaxis/BY/IBP)",
        "url": "https://www.linkedin.com/jobs/view/4335127658/",
        "network": "2 o9 alumni\nat Tesla\n= referral pathway",
        "strategy": "CHANNEL PLANNING MANAGER = your exact sweet spot. Tesla's channel planning covers vehicle demand across sales regions, delivery, and service -- essentially S&OP for a product with 250K+ units/quarter. Lead with demand planning methodology and cross-functional S&OP ownership. Message an o9 alum at Tesla ASAP for a referral.",
        "fill": TESLA_RED_LIGHT,
    },
    {
        "num": 2,
        "title": "Materials Program Manager\nNPI Materials",
        "location": "Fremont, CA\n(On-site)",
        "type": "Full-Time\nOn-site",
        "salary": "~$96,000 - $159,000\n(Tesla NPI PM range)\n+ Stock + Bonus",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "POSTED 14 HRS AGO!\nBe an Early Applicant\n2 o9 alumni at Tesla",
        "posted": "14 HOURS AGO",
        "skills": "NPI materials planning\nBOM & materials management\nSC program management\nSupplier development for new products\nCross-functional NPI coordination\nManufacturing ramp planning",
        "url": "https://www.linkedin.com/jobs/view/4401550597/",
        "network": "2 o9 alumni\nat Tesla\nFremont campus",
        "strategy": "FRESHEST POSTING -- only 14 hours old! Apply IMMEDIATELY before applicant pool grows. NPI materials program management = managing new vehicle model SC launches from BOM to production ramp. Frame consulting engagements as NPI-equivalent programs: new tool deployments with supplier onboarding, BOM configuration, and ramp planning.",
        "fill": "FFF3E0",
    },
    {
        "num": 3,
        "title": "Manager\nSupply Chain Systems",
        "location": "Fremont, CA\n(On-site)",
        "type": "Full-Time\nOn-site",
        "salary": "~$130,000 - $185,000\n(est. Tesla SC Mgr range)\n+ Stock + Bonus",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "4 weeks ago\n2 o9 alumni at Tesla",
        "posted": "4 weeks ago",
        "skills": "Supply chain systems management\nAPS / ERP / WMS systems expertise\nSC technology roadmap ownership\nCross-functional IT + SC leadership\nData & analytics for SC ops\nProcess design + system implementation",
        "url": "https://www.linkedin.com/jobs/view/4385576964/",
        "network": "2 o9 alumni\nat Tesla\nSC systems team",
        "strategy": "SC SYSTEMS MANAGER = your solutions architect background in APS/ERP tools maps directly. Tesla builds many internal SC tools (vs. commercial APS) so they want someone who can bridge business SC process with systems design. Lead with your track record of designing and implementing SC planning systems. o9 alumni network for referral.",
        "fill": "EBF5FB",
    },
    {
        "num": 4,
        "title": "Supply Chain\nProgram Manager",
        "location": "Palo Alto, CA\n(On-site)\nTesla Corp HQ",
        "type": "Full-Time\nOn-site",
        "salary": "~$97,000 - $217,000\n(wide band)\n+ Stock + Bonus",
        "match": "HIGH MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "1 month ago\n2 o9 alumni at Tesla\nPalo Alto HQ",
        "posted": "1 month ago",
        "skills": "Supply chain program management\nCross-functional SC leadership\nSupplier management + NPI\nProgram tracking & risk management\nData-driven SC decision making\nExcel / SQL / BI tools",
        "url": "https://www.linkedin.com/jobs/view/4386579262/",
        "network": "2 o9 alumni\nat Tesla\nPalo Alto location",
        "strategy": "SC PROGRAM MANAGER at Tesla Corp HQ (Palo Alto) = cross-functional SC leadership role managing programs across suppliers, engineering, and manufacturing. Your consulting background = program management depth across SC workstreams. Wide salary band ($97K-$217K) suggests flexibility on seniority level. Contact o9 alum for referral.",
        "fill": "F0F4F8",
    },
    {
        "num": 5,
        "title": "Supply Chain Manager\nStructures",
        "location": "Palo Alto, CA\n(On-site)",
        "type": "Full-Time\nOn-site",
        "salary": "~$120,000 - $185,000\n(est. Tesla SC Mgr range)\n+ Stock + Bonus",
        "match": "HIGH MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "5 days ago\nBe an Early Applicant!\n2 o9 alumni at Tesla",
        "posted": "5 days ago",
        "skills": "SC management for structures/body parts\nSupplier relationship management\nDemand/supply balancing\nCost negotiation & sourcing\nNPI materials planning support\nManufacturing SC operations",
        "url": "https://www.linkedin.com/jobs/view/4399846066/",
        "network": "2 o9 alumni\nat Tesla\nPalo Alto campus",
        "strategy": "FRESH POSTING (5 days) + early applicant advantage. Supply Chain Manager for Structures = managing the body/frame component supply chain for Tesla vehicles. Strong direct SC management role. Frame consulting client SC projects as equivalent commodity management experience. Apply quickly + leverage o9 network for warm introduction.",
        "fill": TESLA_GRAY,
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 95
    fill = PatternFill("solid", start_color=job["fill"])

    data = [
        job["num"], job["title"], job["location"], job["type"],
        job["salary"], job["match"], job["status"], job["posted"],
        job["skills"], job["url"], job["network"], job["strategy"]
    ]

    for c, val in enumerate(data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=TESLA_RED)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=TESLA_DARK)
            cell.alignment = left()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_hot = "14 HRS" in str(val) or "5 days" in str(val)
            cell.font = Font(name="Arial", bold=is_hot, size=9,
                           color=TESLA_RED if "14 HRS" in str(val) else
                           ("375623" if "Early" in str(val) else "595959"))
            cell.alignment = center()
        elif c == 8:
            cell.value = val
            cell.font = Font(name="Arial", bold="HOURS" in str(val), size=9,
                           color=TESLA_RED if "HOURS" in str(val) else "595959")
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = fill
        elif c == 12:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

COL_WIDTHS = [4, 26, 18, 14, 22, 14, 22, 13, 34, 12, 18, 48]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 2 — RESUME TAILORING GUIDE
# ============================================================
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING -- What to Highlight for Each Tesla Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=TESLA_DARK)
t2.border = border

ws2.row_dimensions[2].height = 7
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=TESLA_RED)

HEADERS2 = ["Role", "Must Highlight in Resume", "Keywords to Include",
            "Experience to Lead With", "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=TESLA_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Supply Chain Planning\nChannel Planning Manager",
        "- Channel demand/supply planning ownership\n- S&OP process leadership and facilitation\n- Planning tool expertise (IBP/BY/o9/Kinaxis)\n- Cross-regional demand coordination\n- Forecast accuracy KPIs you've owned\n- Executive S&OP presentation experience",
        "Channel planning, demand planning, supply planning, S&OP, IBP, o9, Kinaxis, Blue Yonder, forecast accuracy, channel inventory, distribution planning, sell-through, sales ops, consensus forecast",
        "Lead with S&OP program ownership and demand planning examples at scale. Show forecast accuracy %, # channels or regions planned, $ inventory optimized. Frame consulting clients as 'channels' -- each client's SC = a demand/supply planning system you owned.",
        "Tesla uses internal SC systems (not commercial APS like o9/Kinaxis). 'Channel' at Tesla means vehicle delivery channels -- auto industry SC knowledge may be limited.",
        "Your SC planning process expertise is more important than tool-specific experience -- Tesla will train on their systems. Emphasize methodology over specific tool names. Add any automotive or EV-adjacent client references in your resume.",
        "APPLY FIRST\n(PERFECT TITLE\nMATCH)"
    ),
    (
        "Materials Program Manager\nNPI Materials",
        "- NPI program management experience\n- Materials planning for new product launches\n- BOM management + supplier onboarding\n- Manufacturing ramp coordination\n- Cross-functional NPI program ownership\n- SC readiness for product launches",
        "NPI, new product introduction, materials planning, BOM, bill of materials, supplier qualification, ramp, program management, materials readiness, supply chain readiness, cross-functional, manufacturing launch",
        "Lead with new product/implementation launch programs you've managed: tool go-lives, new client SC deployments, system cutovers. Frame APS tool implementations as 'NPI-equivalent': you onboard new SC planning capability, qualify data/BOM, manage ramp, and ensure supply readiness.",
        "14-hour-old posting -- apply IMMEDIATELY. NPI Materials is manufacturing-specific (vehicle model launches); consulting PM background is adjacent but not identical.",
        "Speed is everything for this 14-hour posting. Apply NOW. In your cover note: 'I have managed complex SC system implementations from BOM design to go-live ramp -- directly analogous to NPI materials readiness.' Quantify: # SKUs, # suppliers onboarded, ramp timelines hit.",
        "APPLY NOW!\n(14 HRS OLD\n= Apply Today)"
    ),
    (
        "Manager\nSupply Chain Systems",
        "- APS/ERP/WMS system expertise\n- SC technology roadmap ownership\n- Business + IT bridge role experience\n- SC process design + system config\n- Data pipeline + analytics for SC\n- Implementation program leadership",
        "Supply chain systems, APS, ERP, WMS, S/4HANA, IBP, o9, Kinaxis, systems manager, technology roadmap, SC technology, data architecture, process design, system integration, implementation",
        "Lead with SC system implementations: name the systems (o9, Kinaxis, BY, IBP), modules configured, integrations built, # users supported, business outcomes. Show you own SC systems strategy, not just use them. 'Manager, SC Systems' at Tesla = owning their internal planning tool ecosystem.",
        "Tesla builds many proprietary SC tools -- experience with commercial APS is only partially transferable. Role may require managing internal software engineers.",
        "Position yourself as the SC domain expert who guides technology design: 'I translate SC business requirements into system architecture.' Your APS consulting background = understanding what good SC systems look like -- Tesla needs that expertise to build their own. Strong fit for your solutions architect background.",
        "APPLY -- HIGH\nMATCH (SC\nArchitect fit)"
    ),
    (
        "Supply Chain\nProgram Manager",
        "- SC program management track record\n- Cross-functional stakeholder management\n- Supplier management + NPI support\n- Risk identification + mitigation in SC\n- Data-driven SC reporting and metrics\n- End-to-end SC program ownership",
        "Supply chain program manager, program management, cross-functional, supplier management, NPI, SC risk, stakeholder, SC operations, metrics, reporting, delivery, execution",
        "Lead with the biggest, most complex SC programs you've managed: $ value, # stakeholders, timeline, risks managed. Show outcomes: cost savings, on-time delivery %, supplier performance improvement. Consulting background = natural program manager -- you've managed multiple concurrent client programs.",
        "Wide salary band ($97K-$217K) is unclear on seniority -- could be various levels. Palo Alto = Tesla Corp HQ, may be more strategy/planning focused vs. operational.",
        "Position at Sr. level (aim for top of the range). Palo Alto HQ role = more visibility to leadership. Frame consulting project leadership as equivalent to internal program management with 15-20 stakeholders. Broad SC scope is exactly what consultants bring.",
        "APPLY -- GOOD\nFIT + o9\nalumni network"
    ),
    (
        "Supply Chain Manager\nStructures",
        "- Supply chain management for a commodity\n- Supplier relationship management\n- Demand/supply balancing at volume\n- Cost negotiations + sourcing strategy\n- SC risk and contingency management\n- Manufacturing SC operations support",
        "Supply chain manager, structures, body-in-white, stamping, supplier management, sourcing, cost reduction, demand planning, supply planning, manufacturing, automotive, SC risk",
        "Lead with direct SC management experience: commodity ownership, supplier base managed, $ spend, cost reductions achieved. Any manufacturing or automotive SC client work is highly relevant. Show NPI support experience -- Tesla launches frequent vehicle variants.",
        "Commodity-specific role (Structures = body/frame parts). Requires automotive/manufacturing supplier management experience. 5-day posting = still competitive.",
        "Frame consulting client SC projects as commodity management: 'managed SC for [client] across [# suppliers] delivering [$ value] in components.' Apply quickly (5 days old, early applicant still possible). Reach out to o9 alumni at Tesla for a warm introduction.",
        "APPLY -- FRESH\nPOSTING + EARLY\nAPPLICANT"
    ),
]

ROW_FILLS2 = [TESLA_RED_LIGHT, "FFF3E0", "EBF5FB", "F0F4F8", TESLA_GRAY]
PRIORITY_COLORS = [
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (YELLOW_BADGE, YELLOW_DARK),
    (YELLOW_BADGE, YELLOW_DARK),
]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 115
    fill = PatternFill("solid", start_color=ROW_FILLS2[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=TESLA_DARK)
            cell.fill = fill
            cell.alignment = left()
        elif c == 7:
            bg, fg = PRIORITY_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

COL_WIDTHS2 = [24, 40, 40, 44, 32, 40, 16]
for i, w in enumerate(COL_WIDTHS2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 3 — TESLA QUICK FACTS + INTERVIEW PREP
# ============================================================
ws3 = wb.create_sheet("Tesla Facts + Interview Prep")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "TESLA QUICK FACTS + Interview Prep -- Supply Chain & Planning Roles"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=TESLA_DARK)
t3.border = border

ws3.row_dimensions[2].height = 7
ws3.merge_cells("A2:H2")
ws3["A2"].fill = PatternFill("solid", start_color=TESLA_RED)

# Facts section
ws3.row_dimensions[3].height = 22
ws3.merge_cells("A3:H3")
s3 = ws3["A3"]
s3.value = "TESLA COMPANY FACTS -- Critical Context Before Your Interview"
s3.font = Font(name="Arial", bold=True, color=WHITE, size=11)
s3.alignment = center()
s3.fill = PatternFill("solid", start_color=TESLA_RED)
s3.border = border

TESLA_FACTS = [
    ("Company Overview", "Tesla, Inc.: ~$98B revenue (2024). ~125,000 employees. EV + Energy Storage + Solar. Nasdaq: TSLA. CEO: Elon Musk."),
    ("SC Scale", "Tesla manufactured ~1.77M vehicles in 2024. Manages global SC for EVs, batteries (4680 cells), powertrains, energy products (Megapack, Powerwall)."),
    ("Manufacturing Locations", "Fremont CA (Model S/3/X/Y), Austin TX (Cybertruck/Model Y), Sparks NV (Gigafactory 1 - batteries), Berlin GE (Model Y), Shanghai CN (Model 3/Y)."),
    ("SC Approach", "Vertical integration strategy: Tesla makes its own battery cells (4680), seats, software, chips. Fewer external suppliers than traditional OEMs."),
    ("SC Systems", "Tesla uses proprietary internal SC/planning tools -- NOT commercial APS like o9/Kinaxis/BY. They value SC process expertise + willingness to learn internal systems."),
    ("Tesla SC Team", "SC org covers: vehicle SC managers (per commodity), channel planning, NPI materials, global supply managers, SC systems/data, logistics, and procurement."),
    ("Work Culture", "Move fast, no excuses culture. Elon's first principles thinking. Flat hierarchy. Expect to be hands-on, data-driven, and own problems end-to-end with minimal oversight."),
    ("Compensation", "Base + RSU (stock) + bonus. Tesla RSUs vest monthly after 1-year cliff. Total comp is 20-30% above base. Stock grants are significant part of TC."),
    ("Interview Style", "Behavioral + case-based. Expect: 'Tell me about a time when...' + SC scenario questions. Focus on ownership, speed, data-driven decisions, and quantified outcomes."),
    ("LinkedIn Company ID", "15564  |  49 targeted SC/Planning jobs found  |  607 broader SC results  |  Use f_C=15564 for future LinkedIn searches"),
]

for i, (label, val) in enumerate(TESLA_FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 30
    bg = TESLA_RED_LIGHT if i % 2 == 0 else TESLA_GRAY

    lc = ws3.cell(row=r, column=1, value=label)
    lc.font = Font(name="Arial", bold=True, size=9, color=TESLA_DARK)
    lc.fill = PatternFill("solid", start_color=bg)
    lc.alignment = left()
    lc.border = border

    ws3.merge_cells(f"B{r}:H{r}")
    vc = ws3.cell(row=r, column=2, value=val)
    vc.font = nf(size=9)
    vc.fill = PatternFill("solid", start_color=bg)
    vc.alignment = left()
    vc.border = border

# Interview prep section
int_start = len(TESLA_FACTS) + 5
ws3.row_dimensions[int_start].height = 22
ws3.merge_cells(f"A{int_start}:H{int_start}")
sec2 = ws3[f"A{int_start}"]
sec2.value = "TESLA INTERVIEW PREP -- How to Nail SC Interviews at Tesla"
sec2.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec2.alignment = center()
sec2.fill = PatternFill("solid", start_color=TESLA_DARK)
sec2.border = border

INT_HDR = int_start + 1
ws3.row_dimensions[INT_HDR].height = 22
for c, h in enumerate(["#", "Interview Topic", "What Tesla Tests", "How to Answer (SC Consultant Angle)"], 1):
    cell = ws3.cell(row=INT_HDR, column=c)
    cell.value = h
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=TESLA_CHARCOAL)
    cell.alignment = center()
    cell.border = thick_bottom

INTERVIEW_QS = [
    (1, "Why Tesla?",
     "Genuine passion for Tesla's mission: accelerating sustainable energy. Culture fit for fast-paced, mission-driven work.",
     "Don't just say 'EVs are cool.' Talk about Tesla's vertical integration SC model being uniquely complex -- you want to work on the hardest SC problem in manufacturing. Reference Megapack energy storage growth as a SC challenge you're excited to tackle. Show you've studied their SC."),
    (2, "Speed + Ownership Stories",
     "Tesla moves fast. They want people who own problems without waiting for permission. Show bias for action.",
     "Prepare stories where you identified a SC problem, took ownership without being asked, and moved fast to fix it. Quantify the impact and timeline. Tesla hates slow, bureaucratic thinkers. Use: 'I saw X problem, I immediately did Y, the result was Z within N days.'"),
    (3, "SC Planning Methodology",
     "How you structure demand planning, supply planning, and S&OP processes at volume and speed.",
     "Walk through how you build a demand plan at scale: data inputs, statistical baseline, market adjustments, consensus process. For Tesla: think about planning 1.77M vehicles/year across 5+ global factories. Show you can operate at EV-speed (weekly or daily planning cycles, not monthly)."),
    (4, "Data & Analytics Depth",
     "Tesla is data-obsessed. All SC decisions are driven by data. They expect you to pull your own data and build your own analysis.",
     "Show SQL, Excel, Tableau, or Python skills. Have a story about a time you built a SC dashboard, automated a planning report, or found a supply insight from raw data. Tesla SC planners code and query -- position yourself as data-fluent, not just data-aware."),
    (5, "Dealing with Supplier Constraints",
     "Tesla has historically had supply constraints (semiconductors, battery cells, structural castings). How you manage constrained supply.",
     "Share a story about managing supply shortfall: how you triaged demand, communicated allocation decisions, expedited key components, and built recovery plans. Show you stay calm under pressure, use data to prioritize, and communicate clearly to stakeholders."),
    (6, "First Principles Thinking",
     "Elon Musk's core philosophy. Tesla looks for people who question assumptions and reason from fundamentals, not by analogy.",
     "Prepare a story where you questioned a standard SC assumption and found a better approach. Example: 'Instead of using historical forecast as-is, I re-built the demand model from first principles using [leading indicators], which improved accuracy by X%.' Tesla rewards intellectual courage."),
]

INT_FILLS = [TESLA_RED_LIGHT, TESLA_GRAY, TESLA_RED_LIGHT, TESLA_GRAY, TESLA_RED_LIGHT, TESLA_GRAY]

for i, (num, topic, tests, how) in enumerate(INTERVIEW_QS):
    r = INT_HDR + 1 + i
    ws3.row_dimensions[r].height = 60
    bg = INT_FILLS[i % len(INT_FILLS)]
    fill = PatternFill("solid", start_color=bg)

    for c, val in enumerate([num, topic, tests, how], 1):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=13, color=TESLA_RED)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=TESLA_DARK)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 62
ws3.freeze_panes = f"A{int_start + 2}"

# ============================================================
# SAVE
# ============================================================
output = "C:/Users/amolp/Prometheus/tesla_top5_jobs.xlsx"
wb.save(output)
print("Done! tesla_top5_jobs.xlsx saved with 3 sheets.")

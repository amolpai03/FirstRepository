import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

wb = openpyxl.Workbook()

# ── EY Brand Colors ──────────────────────────────────────────────
EY_YELLOW   = "FFE600"   # iconic EY yellow
EY_DARK     = "2E2E38"   # near-black navy
EY_GRAY     = "747480"   # slate gray
EY_LGRAY    = "F2F2F2"   # light background
EY_WHITE    = "FFFFFF"
EY_GOLD     = "C9A84C"   # warm gold accent for highlights
EY_GREEN    = "168736"   # green for match indicators
EY_AMBER    = "F5821F"   # amber for salary callouts
EY_BLUE     = "1A5276"   # dark blue for links

# ── Helpers ──────────────────────────────────────────────────────
def fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium")
    return Border(left=t, right=t, top=t, bottom=t)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — TOP 5 JOBS
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "EY Top 5 Jobs"
ws1.sheet_properties.tabColor = EY_YELLOW

# ── Freeze panes ──
ws1.freeze_panes = "A5"

# ── Main title banner ──
ws1.merge_cells("A1:J1")
c = ws1["A1"]
c.value = "🏢  EY (Ernst & Young) — Top 5 Supply Chain Jobs in the USA"
c.fill = fill(EY_DARK)
c.font = Font(name="Calibri", bold=True, size=16, color=EY_YELLOW)
c.alignment = align("center", "center")
set_row_height(ws1, 1, 32)

# ── Sub-banner with profile match ──
ws1.merge_cells("A2:J2")
c2 = ws1["A2"]
c2.value = ("Curated for: 7-yr SC Planning Consultant  •  Tools: o9 | OMP | Kinaxis | Blue Yonder | SAP IBP  "
            "•  S&OP | Demand | Supply | Inventory  •  USA Only  •  Pulled: April 16, 2026")
c2.fill = fill(EY_YELLOW)
c2.font = Font(name="Calibri", bold=True, size=10, color=EY_DARK)
c2.alignment = align("center", "center")
set_row_height(ws1, 2, 22)

# ── Spacer row ──
ws1.merge_cells("A3:J3")
ws1["A3"].fill = fill(EY_LGRAY)
set_row_height(ws1, 3, 6)

# ── Column headers ──
headers = [
    ("A", "#",                  4),
    ("B", "Job Title",          36),
    ("C", "Level",              14),
    ("D", "Location(s)",        22),
    ("E", "Salary (USD)",       16),
    ("F", "Posted",             12),
    ("G", "Key Requirements",   40),
    ("H", "Tool Match 🎯",      26),
    ("I", "EY Practice",        22),
    ("J", "Apply Link",         18),
]
for col, hdr, width in headers:
    c = ws1[f"{col}4"]
    c.value = hdr
    c.fill = fill(EY_DARK)
    c.font = Font(name="Calibri", bold=True, size=11, color=EY_YELLOW)
    c.alignment = align("center", "center")
    c.border = border()
    set_col_width(ws1, col, width)
set_row_height(ws1, 4, 22)

# ── Job Data ──────────────────────────────────────────────────
jobs = [
    {
        "num": 1,
        "title": "Technology Consulting – Sr. Manager, Digital Supply Chain Planning",
        "level": "Senior Manager",
        "location": "Charlotte, NC\n(Location Open – USA)",
        "salary": "$275,335 / yr",
        "posted": "Apr 10, 2026",
        "reqs": ("8 yrs consulting post-BA; 6 yrs SC planning delivery; "
                 "4 yrs hands-on with advanced planning platforms; "
                 "2+ full-lifecycle implementations; "
                 "S&OP / Demand & Supply Planning SME; "
                 "onshore/offshore delivery; 80% travel"),
        "tools": "Blue Yonder ✅\no9 ✅\nOMP ✅\nKinaxis ✅\nSAP IBP ✅\n(ALL 5 tools match!)",
        "practice": "EY Technology\nConsulting – SC Tech",
        "url": "https://careers.ey.com/ey/job/Charlotte-Tech-Con-Tech-and-Platforms-S-C-Tech-Digi-S-C-Planning-Senior-Manager-NC-28202/1373227233/",
        "row_fill": EY_YELLOW,
        "font_color": EY_DARK,
        "note": "★ BEST MATCH — All 5 tools explicitly listed",
    },
    {
        "num": 2,
        "title": "Supply Chain & Operations – Manager / Senior Manager, Business Consulting",
        "level": "Manager / Sr. Mgr",
        "location": "Anywhere in USA\n(Multiple cities)",
        "salary": "Competitive\n(~$126K–$200K)",
        "posted": "Feb 23, 2026",
        "reqs": ("6+ yrs SC consulting; deliver large-scale SC "
                 "transformation; IBP / S&OP systems expertise; "
                 "demand/supply/logistics planning; "
                 "client management & business development"),
        "tools": "SAP IBP ✅\nS&OP / IBP ✅\nDemand Planning ✅",
        "practice": "EY Business\nConsulting – SC&O",
        "url": "https://careers.ey.com/ey/job/Supply-Chain-&-Operations,-ManagerSenior-Manager,-Business-Consulting-048583/698835801/",
        "row_fill": "FFFDE7",
        "font_color": EY_DARK,
        "note": "",
    },
    {
        "num": 3,
        "title": "EY-Parthenon – Strategy & Execution, Supply Chain – Director",
        "level": "Director",
        "location": "NYC, Atlanta, Boston,\nChicago, Dallas, Denver,\nDetroit, Houston, LA,\nMcLean, Philly, SF, Seattle",
        "salary": "$205,000–$235,000",
        "posted": "Mar 24, 2026",
        "reqs": ("5 yrs post-BA or 3 yrs post-grad; MBA preferred; "
                 "M&A supply chain integration & divestiture experience; "
                 "plan/source/make/deliver expertise; "
                 "direct materials, planning/inventory, logistics optimization; "
                 "80% travel"),
        "tools": "SC Strategy ✅\nM&A Integration ✅\nInventory Optimization ✅",
        "practice": "EY-Parthenon\nStrategy & Execution",
        "url": "https://careers.ey.com/ey/job/New-York-EY-Parthenon-Strategy-and-Execution-Supply-Chain-Director-Multiple-Locations-NY-10001-8604/1218623401/",
        "row_fill": "FFF3CD",
        "font_color": EY_DARK,
        "note": "",
    },
    {
        "num": 4,
        "title": "EY-Parthenon – Strategy & Execution, Supply Chain – Manager (Multiple Positions)",
        "level": "Manager",
        "location": "Detroit, MI\n(Multiple positions)",
        "salary": "$235,005 / yr",
        "posted": "Mar 20, 2026",
        "reqs": ("5 yrs post-BA or 3 yrs post-grad; "
                 "2 yrs M&A integrations/divestitures; "
                 "3 yrs SC infrastructure assessments; "
                 "2 yrs SC integration/separation of client operations; "
                 "18 months consulting in M&A; "
                 "80% travel"),
        "tools": "SC Strategy ✅\nM&A / Carve-out ✅\nSC Assessment ✅",
        "practice": "EY-Parthenon\nStrategy & Execution",
        "url": "https://careers.ey.com/ey/job/Detroit-EY-Parthenon-Strategy-and-Execution-Supply-Chain-Manager-Multiple-Positions-MI-48226/1365990733/",
        "row_fill": "FFFDE7",
        "font_color": EY_DARK,
        "note": "",
    },
    {
        "num": 5,
        "title": "EY-Parthenon – Strategy & Execution, Supply Chain – Sr. Associate / Consultant",
        "level": "Sr. Associate / Consultant",
        "location": "NYC, Atlanta, Dallas,\nMcLean, Boston, Houston,\nChicago, LA, SF",
        "salary": "$130,000–$185,000",
        "posted": "Apr 1, 2026",
        "reqs": ("2 yrs post-BA or 18 months post-grad; "
                 "supply chain functional expertise; "
                 "M&A SC diligence/integration/separation analysis; "
                 "quantitative SC modeling; "
                 "up to 80% travel; multiple locations"),
        "tools": "SC Consulting ✅\nM&A Diligence ✅\nAnalytics ✅",
        "practice": "EY-Parthenon\nStrategy & Execution",
        "url": "https://careers.ey.com/ey/job/New-York-EY-Parthenon-Strategy-and-Execution-Supply-Chain-Senior-Associate-Consultant-NY-10001-8604/1285432501/",
        "row_fill": "F5F5F5",
        "font_color": EY_DARK,
        "note": "",
    },
]

for idx, job in enumerate(jobs):
    row = 5 + idx
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]

    row_data = [
        job["num"],
        job["title"],
        job["level"],
        job["location"],
        job["salary"],
        job["posted"],
        job["reqs"],
        job["tools"],
        job["practice"],
        "→ Apply Here",
    ]

    for ci, col in enumerate(cols):
        cell = ws1[f"{col}{row}"]
        cell.value = row_data[ci]
        cell.fill = fill(job["row_fill"])
        cell.font = Font(name="Calibri", size=10, color=job["font_color"],
                         bold=(col == "B"))
        cell.alignment = align("left" if col != "A" else "center", "center", wrap=True)
        cell.border = border()

    # Hyperlink on J column
    j_cell = ws1[f"J{row}"]
    j_cell.hyperlink = job["url"]
    j_cell.font = Font(name="Calibri", size=10, color=EY_BLUE, bold=True,
                       underline="single")
    j_cell.alignment = align("center", "center")

    # Special highlight for Job 1 (best match)
    if idx == 0:
        for col in cols:
            ws1[f"{col}{row}"].border = thick_border()
        ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12,
                                   color=EY_DARK)

    # Row height
    set_row_height(ws1, row, 70)

# ── Best Match callout banner ──
ws1.merge_cells("A10:J10")
c_bm = ws1["A10"]
c_bm.value = ("★  JOB #1 HIGHLIGHT:  Digital SC Planning Senior Manager — The ONLY active EY role explicitly "
               "listing ALL 5 of your exact tools (Blue Yonder + o9 + OMP + Kinaxis + SAP IBP).  "
               "$275K base.  Posted April 10, 2026.  ★")
c_bm.fill = fill(EY_DARK)
c_bm.font = Font(name="Calibri", bold=True, size=10, color=EY_YELLOW)
c_bm.alignment = align("center", "center", wrap=True)
set_row_height(ws1, 10, 28)

# ── Footer ──
footer_row = 12
ws1.merge_cells(f"A{footer_row}:J{footer_row}")
cf = ws1[f"A{footer_row}"]
cf.value = ("Data sourced directly from careers.ey.com  •  All jobs confirmed active in USA as of April 16, 2026  "
            "•  EY accepts ongoing applications — apply promptly as openings close frequently  "
            "•  Salaries shown are base; total comp may include bonus, benefits, and equity")
cf.fill = fill(EY_LGRAY)
cf.font = Font(name="Calibri", size=9, italic=True, color=EY_GRAY)
cf.alignment = align("center", "center", wrap=True)
set_row_height(ws1, footer_row, 30)

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — RESUME TAILORING GUIDE
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resume Tailoring Guide")
ws2.sheet_properties.tabColor = EY_DARK

# Banner
ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "✏️  EY Resume Tailoring — Align Your 7-Year SC Planning Background"
c.fill = fill(EY_DARK)
c.font = Font(name="Calibri", bold=True, size=14, color=EY_YELLOW)
c.alignment = align("center", "center")
set_row_height(ws2, 1, 30)

ws2.merge_cells("A2:G2")
c2 = ws2["A2"]
c2.value = "EY values: Exceptional client service | Teamwork | Integrity | Respect | Professional Growth"
c2.fill = fill(EY_YELLOW)
c2.font = Font(name="Calibri", bold=True, size=10, color=EY_DARK)
c2.alignment = align("center", "center")
set_row_height(ws2, 2, 20)

ws2.merge_cells("A3:G3")
ws2["A3"].fill = fill(EY_LGRAY)
set_row_height(ws2, 3, 5)

resume_headers = [
    ("A", "Job #",        8),
    ("B", "Job Title (Short)",   28),
    ("C", "EY Keyword / Theme",  26),
    ("D", "Your Experience to Highlight",  36),
    ("E", "Power Phrases to Use",          36),
    ("F", "Skills to Front-Load",          30),
    ("G", "Watch Out For",                 28),
]
for col, hdr, width in resume_headers:
    c = ws2[f"{col}4"]
    c.value = hdr
    c.fill = fill(EY_DARK)
    c.font = Font(name="Calibri", bold=True, size=10, color=EY_YELLOW)
    c.alignment = align("center", "center", wrap=True)
    c.border = border()
    set_col_width(ws2, col, width)
set_row_height(ws2, 4, 22)

resume_data = [
    (
        "1",
        "Digital SC Planning\nSr. Manager",
        "SC Tech Architecture\nAdvanced Planning Systems\nFull-lifecycle implementation",
        "4+ yrs hands-on with EVERY tool they listed: BY, o9, OMP, Kinaxis, SAP IBP — "
        "Quantify: # implementations, $ value, team size led, clients served",
        '"Led 3 full-lifecycle Kinaxis implementations..."\n'
        '"Architected SAP IBP demand planning solution for $2B revenue client..."\n'
        '"Delivered onshore/offshore SC transformation reducing planning cycle 40%"',
        "Solution Architecture | Full-Lifecycle Implementation | "
        "Blue Yonder | Kinaxis | SAP IBP | o9 | OMP | Demand/Supply Planning SME | "
        "Onshore-Offshore Delivery | S&OP | Value Chain",
        "This is a SENIOR MANAGER role — emphasize leadership of teams and client "
        "relationships, not just individual contributor skills. 80% travel is required; "
        "mention flexibility in your cover letter.",
    ),
    (
        "2",
        "SC&O Manager /\nSenior Manager",
        "SC Transformation\nS&OP Process Redesign\nIBP / Planning Systems",
        "7 yrs consulting SC planning across industries — "
        "S&OP design, IBP implementation, org re-engineering, "
        "performance improvement; client-facing delivery",
        '"Delivered end-to-end SC transformation for Fortune 500 client..."\n'
        '"Redesigned S&OP process reducing excess inventory by $X..."\n'
        '"Implemented IBP system increasing forecast accuracy to X%"',
        "Supply Chain Transformation | S&OP | IBP | Demand/Supply Planning | "
        "Process Re-engineering | Client Delivery | ERP Systems | SAP IBP",
        "EY SC&O group serves multiple industries — tailor each application by "
        "matching your industry SC experience to the target industry for the role.",
    ),
    (
        "3",
        "EY-Parthenon\nSC Director",
        "M&A Supply Chain\nDue Diligence\nIntegration / Carve-out",
        "If you have M&A or integration project experience, highlight it first. "
        "Otherwise position SC expertise in context of rapid assessments and "
        "synergy identification across plan/source/make/deliver",
        '"Conducted rapid SC assessment across 14-country M&A integration..."\n'
        '"Identified $30M inventory rationalization opportunity in 6-week diligence..."\n'
        '"Led SC integration workstream for post-merger Day 1 readiness"',
        "M&A SC Integration | Carve-out | Divestiture | SC Due Diligence | "
        "Plan-Source-Make-Deliver | Synergy Identification | MBA Preferred",
        "EY-Parthenon is pure strategy/M&A — do NOT lead with tools/systems. "
        "Lead with SC strategy, business impact, and financial metrics. "
        "MBA is preferred; note if you have one.",
    ),
    (
        "4",
        "EY-Parthenon\nSC Manager",
        "M&A SC Integration\nInfrastructure Assessment\nOperations Separation",
        "SC consulting background, project/engagement management, "
        "ability to assess SC infrastructure quickly, "
        "cross-functional team leadership",
        '"Managed $XM SC integration engagement from sign-to-close..."\n'
        '"Delivered SC separation playbook for $500M carve-out in 8 weeks..."\n'
        '"Led cross-functional team of X in supply chain infrastructure assessment"',
        "SC Integration | Business Divestiture | Carve-out Planning | "
        "Engagement Management | Cross-functional Leadership | SC Assessment",
        "Location listed as Detroit, MI — confirm willingness to relocate or commute. "
        "Role is very M&A specific; if weak on M&A, consider leaning on #2 or #1 first.",
    ),
    (
        "5",
        "EY-Parthenon\nSC Sr. Associate",
        "SC Strategy & Analytics\nM&A Diligence Support\nQuantitative Modeling",
        "Analytical SC work, data-driven modeling, "
        "supply chain assessments, problem framing, "
        "junior team support — easier entry point than Director",
        '"Developed SC optimization model reducing inventory carrying cost by $XM..."\n'
        '"Supported SC diligence across 3 M&A transactions in 6 months..."\n'
        '"Built demand forecast model improving accuracy from X% to Y%"',
        "SC Analytics | Quantitative Modeling | M&A Diligence | Demand Forecasting | "
        "Inventory Optimization | Supply Planning | Python/Excel Analytics",
        "This is a more junior title than your 7-yr experience warrants — "
        "use as a BACKUP option or if targeting EY-Parthenon specifically. "
        "$130K-$185K is below your likely market value.",
    ),
]

for idx, row_data in enumerate(resume_data):
    row = 5 + idx
    row_fill = EY_YELLOW if idx == 0 else ("FFFDE7" if idx % 2 == 0 else EY_LGRAY)
    for ci, col in enumerate(["A", "B", "C", "D", "E", "F", "G"]):
        c = ws2[f"{col}{row}"]
        c.value = row_data[ci]
        c.fill = fill(row_fill)
        c.font = Font(name="Calibri", size=9,
                      color=EY_DARK if ci < 2 else "333333",
                      bold=(ci == 0))
        c.alignment = align("left" if ci > 0 else "center", "top", wrap=True)
        c.border = border()
    set_row_height(ws2, row, 90)

# ── EY Cover Letter Tips ──────────────────────────────────────
tip_row = 11
ws2.merge_cells(f"A{tip_row}:G{tip_row}")
ct = ws2[f"A{tip_row}"]
ct.value = "📝  EY COVER LETTER & APPLICATION TIPS"
ct.fill = fill(EY_DARK)
ct.font = Font(name="Calibri", bold=True, size=12, color=EY_YELLOW)
ct.alignment = align("center", "center")
set_row_height(ws2, tip_row, 24)

tips = [
    ("Opening Hook",
     'Lead with a specific SC planning outcome: "I reduced a $450M client\'s forecast error from 28% to 11% '
     'using SAP IBP — that\'s the kind of measurable impact I\'d bring to EY." EY interviewers are drawn to '
     'quantified business outcomes over generic consulting narrative.'),
    ("EY Differentiator",
     'Mention EY\'s SC practice breadth: "EY\'s unique position spanning technology consulting AND '
     'EY-Parthenon strategy aligns with my ability to bridge systems implementation with SC strategy." '
     'Shows you understand EY\'s service line structure.'),
    ("Tool Alignment (Job #1)",
     'For the Digital SC Planning role: "Having delivered full-lifecycle implementations across all five '
     'platforms listed in the JD — Blue Yonder, o9, OMP, Kinaxis, and SAP IBP — I bring immediate '
     'client-facing tool depth that most candidates can\'t offer." Direct and compelling.'),
    ("EY Values Language",
     'Weave in EY\'s values: "exceptional client service," "building a better working world," '
     '"inclusive culture," and "continuous learning." Reference EY\'s brand promise in 1-2 sentences.'),
    ("Application Tip",
     'Apply directly via careers.ey.com (not Indeed) — EY tracks source. Connect with EY SC partners on '
     'LinkedIn BEFORE applying and reference the connection in your cover letter. EY has strong internal '
     'referral culture; a warm introduction multiplies callback odds 5x.'),
]

for ti, (label, content) in enumerate(tips):
    r = tip_row + 1 + ti
    ws2.merge_cells(f"A{r}:B{r}")
    lc = ws2[f"A{r}"]
    lc.value = label
    lc.fill = fill("E8E057")
    lc.font = Font(name="Calibri", bold=True, size=10, color=EY_DARK)
    lc.alignment = align("center", "center", wrap=True)
    lc.border = border()
    ws2.merge_cells(f"C{r}:G{r}")
    tc = ws2[f"C{r}"]
    tc.value = content
    tc.fill = fill(EY_LGRAY)
    tc.font = Font(name="Calibri", size=9, color="333333")
    tc.alignment = align("left", "center", wrap=True)
    tc.border = border()
    set_row_height(ws2, r, 55)

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — EY QUICK FACTS + INTERVIEW PREP
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("EY Facts & Interview Prep")
ws3.sheet_properties.tabColor = EY_GOLD

# Banner
ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "🎯  EY Interview Prep & Company Quick Facts — Supply Chain Practice"
c.fill = fill(EY_DARK)
c.font = Font(name="Calibri", bold=True, size=14, color=EY_YELLOW)
c.alignment = align("center", "center")
set_row_height(ws3, 1, 30)

ws3.merge_cells("A2:F2")
c2 = ws3["A2"]
c2.value = '"Building a better working world" — EY\'s purpose statement. Use it.'
c2.fill = fill(EY_YELLOW)
c2.font = Font(name="Calibri", bold=True, size=11, color=EY_DARK, italic=True)
c2.alignment = align("center", "center")
set_row_height(ws3, 2, 22)

ws3.merge_cells("A3:F3")
ws3["A3"].fill = fill(EY_LGRAY)
set_row_height(ws3, 3, 5)

# Column widths
for col, w in [("A", 26), ("B", 46), ("C", 26), ("D", 46), ("E", 22), ("F", 36)]:
    set_col_width(ws3, col, w)

# ── Section: Company Quick Facts ────────────────────────────
def s3_section(ws, row, title, color=EY_DARK, fcolor=EY_YELLOW):
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = title
    c.fill = fill(color)
    c.font = Font(name="Calibri", bold=True, size=11, color=fcolor)
    c.alignment = align("center", "center")
    set_row_height(ws, row, 22)

def s3_row(ws, row, label, value, label_fill=EY_YELLOW, val_fill=EY_LGRAY,
           label_start="A", label_end="B", val_start="C", val_end="F"):
    ws.merge_cells(f"{label_start}{row}:{label_end}{row}")
    lc = ws[f"{label_start}{row}"]
    lc.value = label
    lc.fill = fill(label_fill)
    lc.font = Font(name="Calibri", bold=True, size=10, color=EY_DARK)
    lc.alignment = align("left", "center", wrap=True)
    lc.border = border()
    ws.merge_cells(f"{val_start}{row}:{val_end}{row}")
    vc = ws[f"{val_start}{row}"]
    vc.value = value
    vc.fill = fill(val_fill)
    vc.font = Font(name="Calibri", size=10, color="333333")
    vc.alignment = align("left", "center", wrap=True)
    vc.border = border()
    set_row_height(ws, row, 42)

s3_section(ws3, 4, "📊  EY AT A GLANCE")
s3_row(ws3, 5, "Full Name & HQ", "Ernst & Young Global Limited | London, UK HQ | US HQ: New York, NY")
s3_row(ws3, 6, "Revenue (2024)", "$51.2B global revenue | ~$17B+ in Americas | Big 4 accounting & consulting firm")
s3_row(ws3, 7, "Employees", "~395,000 globally | ~80,000 in the USA | Strong campus recruiting culture")
s3_row(ws3, 8, "Service Lines", "Assurance | Tax | Strategy & Transactions (EY-Parthenon) | Consulting | Technology Consulting")
s3_row(ws3, 9, "SC Practice Focus",
       ("EY Supply Chain & Operations (SC&O) sits within Business Consulting. "
        "Covers: SC strategy, S&OP/IBP, network design, SC risk, digital SC transformation, "
        "procurement, and tech implementations (SAP IBP, Kinaxis, Blue Yonder, o9, OMP)."))
s3_row(ws3, 10, "EY-Parthenon",
       ("EY-Parthenon is EY's pure strategy arm — focuses on corporate strategy and M&A SC "
        "integration/divestiture. Analyst-to-Director career path. Higher strategy rigor, "
        "lower tool implementation focus vs. Business Consulting."))
s3_row(ws3, 11, "Tech Consulting",
       ("Technology Consulting houses the Digital SC Planning practice (Job #1). "
        "This is where SAP IBP, Kinaxis, Blue Yonder, o9, OMP implementations live. "
        "Your tools expertise is the #1 differentiator for these roles."))
s3_row(ws3, 12, "Culture / Values",
       "Exceptional client service | Integrity | Respect | Teaming | Building a better working world | "
       "Inclusiveness | High performance — reference these in interviews explicitly")

set_row_height(ws3, 4, 22)

# ── Interview Prep Questions ─────────────────────────────────
s3_section(ws3, 14, "🎙️  LIKELY INTERVIEW QUESTIONS + ANSWER FRAMEWORKS")

qs = [
    ("Q: Walk me through a supply chain planning transformation you led.",
     "STAR Method: Situation (client/industry context) → Task (what was broken: forecast accuracy, "
     "excess inventory, long planning cycles) → Action (IBP/S&OP redesign, system implementation: "
     "SAP IBP/Kinaxis/o9) → Result (quantify: % improvement, $ savings, cycle time reduction). "
     "Mention the specific tool. EY loves quantified outcomes."),
    ("Q: Why EY specifically, and not Deloitte or Accenture?",
     '"EY\'s SC practice sits at the intersection of strategy (EY-Parthenon) and technology '
     '(Tech Consulting) — that breadth matches my profile exactly. I also value EY\'s purpose-led '
     'culture and \'building a better working world\' mission, which resonates with SC work that '
     'creates real operational impact for clients." Be specific — mention their practice structure.'),
    ("Q: Tell me about a time you implemented an advanced planning system.",
     "Pick your strongest implementation (Kinaxis / SAP IBP / Blue Yonder). Cover: client size & "
     "industry, business problem, your specific role in design and config, change management "
     "challenges, go-live outcome. Highlight full-lifecycle ownership and quantify results. "
     "For Job #1, this is your single most important story — rehearse it until flawless."),
    ("Q: How do you handle a client who resists process change?",
     "EY SC consulting involves major org change. Framework: 1) Listen first — understand their "
     "'why' 2) Align change to their business metrics / incentives 3) Show small wins early "
     "4) Build internal champions 5) Use data to de-personalize resistance. "
     "Cite a real example where you turned a skeptical stakeholder into a project champion."),
    ("Q: Describe a complex S&OP implementation you've led.",
     "Cover: # stakeholders, cross-functional scope (demand/supply/finance/exec), cadence design, "
     "tool selected and why, process change challenges, time to stabilize, measurable improvement. "
     "For EY, emphasize how the S&OP connected to client's financial planning (IBP = integrated "
     "business planning, not just SC planning)."),
    ("Q: How would you sell EY's SC practice to a new client?",
     "Show business development instinct: 1) Identify client's SC pain (inventory, service level, "
     "cost) via data benchmarking 2) Map EY's capabilities (transformation + tech + Parthenon) "
     "to their need 3) Propose phased approach with quick wins in 8 weeks 4) Reference comparable "
     "client outcome. EY managers are expected to contribute to BD — mention your appetite for it."),
    ("Q: What's your experience with offshore/onshore delivery models?",
     "Relevant for Job #1 (4 yrs onshore/offshore required). Cover: how you managed time zone "
     "gaps, task allocation between onsite leads and offshore teams, quality oversight, "
     "communication cadence. Frame it as a leadership challenge you solved, not just a logistics fact."),
    ("Q: What's the biggest trend in digital supply chain planning?",
     "Strong answer: 'The convergence of AI-driven demand sensing (o9, Kinaxis) with real-time "
     "supply visibility is reshaping the S&OP cycle from monthly batch processes to continuous "
     "planning. The move from statistical forecasting to ML-based demand shaping, combined with "
     "autonomous supply responses, is where the next $10B of SC value will be unlocked.' "
     "Shows thought leadership, not just tool familiarity."),
]

for qi, (q, a) in enumerate(qs):
    r = 15 + qi
    ws3.merge_cells(f"A{r}:B{r}")
    qc = ws3[f"A{r}"]
    qc.value = q
    qc.fill = fill("E8E057" if qi % 2 == 0 else EY_YELLOW)
    qc.font = Font(name="Calibri", bold=True, size=9, color=EY_DARK)
    qc.alignment = align("left", "top", wrap=True)
    qc.border = border()
    ws3.merge_cells(f"C{r}:F{r}")
    ac = ws3[f"C{r}"]
    ac.value = a
    ac.fill = fill(EY_LGRAY if qi % 2 == 0 else EY_WHITE)
    ac.font = Font(name="Calibri", size=9, color="333333")
    ac.alignment = align("left", "top", wrap=True)
    ac.border = border()
    set_row_height(ws3, r, 65)

set_row_height(ws3, 14, 22)

# ── EY Salary Benchmarks ──────────────────────────────────────
s_row = 24
s3_section(ws3, s_row, "💰  EY SALARY BENCHMARKS — Supply Chain Consulting (USA, 2026)")
salary_data = [
    ("Senior Associate / Consultant", "$85K–$130K base", "Entry SC consulting at EY; EY-Parthenon ~$130K-$185K"),
    ("Manager", "$128K–$235K base", "Wide range: Business Consulting ~$128K-$200K; EY-Parthenon Manager ~$235K"),
    ("Senior Manager", "$150K–$300K base", "Digital SC Planning Sr. Mgr = $275K; varies by practice/geography"),
    ("Director / Principal", "$205K–$350K+", "EY-Parthenon Director = $205K-$235K base; tech consulting Directors higher"),
    ("Partner", "$500K–$1.5M+ total", "Equity + base; highly performance-dependent; 10-15 yr track"),
    ("Bonus / Total Comp", "+15–40% bonus typical", "EY performance bonuses range 15-40%; strong performers exceed midpoint"),
]
for si, (level, salary, note) in enumerate(salary_data):
    r = s_row + 1 + si
    ws3.merge_cells(f"A{r}:B{r}")
    lc = ws3[f"A{r}"]
    lc.value = level
    lc.fill = fill(EY_YELLOW if si % 2 == 0 else "FFFDE7")
    lc.font = Font(name="Calibri", bold=True, size=10, color=EY_DARK)
    lc.alignment = align("center", "center", wrap=True)
    lc.border = border()
    ws3.merge_cells(f"C{r}:D{r}")
    sc2 = ws3[f"C{r}"]
    sc2.value = salary
    sc2.fill = fill("D5F5E3" if si % 2 == 0 else "EAFAF1")
    sc2.font = Font(name="Calibri", bold=True, size=11, color=EY_DARK)
    sc2.alignment = align("center", "center")
    sc2.border = border()
    ws3.merge_cells(f"E{r}:F{r}")
    nc = ws3[f"E{r}"]
    nc.value = note
    nc.fill = fill(EY_LGRAY)
    nc.font = Font(name="Calibri", size=9, color="555555", italic=True)
    nc.alignment = align("left", "center", wrap=True)
    nc.border = border()
    set_row_height(ws3, r, 32)

# ── EY SC Practice Differences ────────────────────────────────
diff_row = 31
s3_section(ws3, diff_row, "🗂️  EY SC PRACTICE GUIDE — Which Track Is Right For You?")
tracks = [
    ("EY Business Consulting\n(SC&O Practice)",
     ("Best if: You want to do IBP/S&OP process consulting + system implementations. "
      "Jobs #1 and #2 are in this track. Day-to-day: client SC transformation, S&OP "
      "redesign, technology selection & deployment. Career: Consultant → Manager → Sr. Mgr → Director → Partner. "
      "Your tools expertise (Kinaxis, SAP IBP, o9, BYD, OMP) is your ticket in.")),
    ("EY Technology Consulting\n(Digital SC Planning)",
     ("Best if: You are a hands-on platform implementer with full-lifecycle project history. "
      "Job #1 (Digital SC Planning Sr. Mgr) is here. Day-to-day: architecting and deploying "
      "APS platforms, leading onshore/offshore teams, client training. "
      "Most direct match to your tool skills. $275K base reflects tool scarcity in market.")),
    ("EY-Parthenon\n(Strategy & Execution)",
     ("Best if: You have M&A exposure or want to pivot to pure SC strategy at premium pay. "
      "Jobs #3, #4, #5 are here. Day-to-day: M&A diligence, integration planning, carve-out "
      "execution — less tool-heavy, more rapid SC assessment. "
      "MBA preferred. If no M&A background, lead with SC assessment speed and financial metrics.")),
]
for ti, (track, desc) in enumerate(tracks):
    r = diff_row + 1 + ti
    ws3.merge_cells(f"A{r}:B{r}")
    tc = ws3[f"A{r}"]
    tc.value = track
    tc.fill = fill(EY_YELLOW if ti == 1 else "FFFDE7")
    tc.font = Font(name="Calibri", bold=True, size=10, color=EY_DARK)
    tc.alignment = align("center", "center", wrap=True)
    tc.border = border()
    ws3.merge_cells(f"C{r}:F{r}")
    dc = ws3[f"C{r}"]
    dc.value = desc
    dc.fill = fill(EY_LGRAY)
    dc.font = Font(name="Calibri", size=9, color="333333")
    dc.alignment = align("left", "top", wrap=True)
    dc.border = border()
    set_row_height(ws3, r, 72)

# ── Final footer ──────────────────────────────────────────────
last_row = 36
ws3.merge_cells(f"A{last_row}:F{last_row}")
cf = ws3[f"A{last_row}"]
cf.value = ("Jobs sourced from careers.ey.com  •  Salary data from EY public postings + Glassdoor benchmarks  "
            "•  All positions confirmed USA-based as of April 16, 2026  •  EY Big 4 — competitive market; apply within 2 weeks of posting")
cf.fill = fill(EY_LGRAY)
cf.font = Font(name="Calibri", size=8, italic=True, color=EY_GRAY)
cf.alignment = align("center", "center", wrap=True)
set_row_height(ws3, last_row, 24)

# ── Freeze Sheet 3 top rows ──
ws3.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────
output_path = r"C:\Users\amolp\Prometheus\ey_top5_jobs.xlsx"
wb.save(output_path)
print(f"[OK]  Saved: {output_path}")
print("[OK]  Sheets: EY Top 5 Jobs | Resume Tailoring Guide | EY Facts & Interview Prep")
print("[OK]  Job #1: Digital SC Planning Sr. Manager -- $275,335 | ALL 5 tools match")
print("[OK]  Job #3: EY-Parthenon Director -- $205K-$235K | NYC + 13 US cities")

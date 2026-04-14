from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

WHITE = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="595959"))

def nf(bold=False, color="1A1A1A", size=9):
    return Font(name="Arial", bold=bold, size=size, color=color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def hyperlink(cell, url, label="View Job ->"):
    cell.hyperlink = url
    cell.value = label
    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

# Home Depot brand colors
HD_ORANGE       = "F96302"
HD_DARK         = "1C1C1C"
HD_LIGHT_ORANGE = "FEF0E7"
HD_LIGHT_GRAY   = "F5F5F5"
HD_DARK_ORANGE  = "C94F00"
HD_MEDIUM_ORANGE= "FDD9BC"

GREEN_BADGE  = "C6EFCE"
GREEN_DARK   = "375623"
YELLOW_BADGE = "FFEB9C"
YELLOW_DARK  = "7F6000"
BLUE_BADGE   = "DDEBF7"
BLUE_DARK    = "1F4E79"

# ============================================================
# SHEET 1 -- HOME DEPOT TOP 5 JOBS
# ============================================================
ws1 = wb.active
ws1.title = "Home Depot Top 5 Jobs"

ws1.row_dimensions[1].height = 38
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "THE HOME DEPOT -- Top 5 Supply Chain & Planning Jobs  |  LinkedIn Search  |  April 2026"
t.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t.alignment = center()
t.fill = PatternFill("solid", start_color=HD_DARK)
t.border = border

ws1.row_dimensions[2].height = 7
ws1.merge_cells("A2:L2")
ws1["A2"].fill = PatternFill("solid", start_color=HD_ORANGE)

ws1.row_dimensions[3].height = 20
ws1.merge_cells("A3:L3")
s = ws1["A3"]
s.value = "110 school alumni at Home Depot  |  Job 2 posted 3 days ago (early applicant!)  |  All roles Atlanta HQ  |  35 targeted SC/Planning results  |  HD LinkedIn Company ID: 1534"
s.font = Font(name="Arial", italic=True, size=9, color=HD_DARK)
s.alignment = center()
s.fill = PatternFill("solid", start_color=HD_LIGHT_ORANGE)
s.border = border

HEADERS = ["#", "Job Title", "Location", "Type", "Salary Range", "Match Level",
           "Status / Alumni", "Posted", "Key Skills Required", "Link", "Network Edge", "Apply Strategy"]
ws1.row_dimensions[4].height = 24
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=HD_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws1.freeze_panes = "A5"

JOBS = [
    {
        "num": 1,
        "title": "Manager\nSupply Chain PMO\n(Project Mgmt Office)",
        "location": "Atlanta, GA\n(On-site)\nHD Corp HQ",
        "type": "Full-Time\nOn-site",
        "salary": "$119,000 - $176,000\n(SC Manager range)\n+ Bonus",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "Recently posted\n110 school alumni\nat Home Depot",
        "posted": "Recently posted",
        "skills": "SC project / program management\nCross-functional stakeholder leadership\nProject governance + reporting\nSC strategy implementation\nExecutive communication\n5+ yrs SC or PM experience",
        "url": "https://www.linkedin.com/jobs/view/4389315453/",
        "network": "110 NCSU alumni\nat Home Depot\n= large network",
        "strategy": "SC PMO MANAGER = your consulting program management background is a near-perfect fit. PMO drives key cross-functional SC projects (the exact work consultants do). Role reports into SC leadership, giving visibility to Directors and VPs. Lead with your history of managing complex multi-workstream SC programs. Use the 110 alumni network to get a warm intro.",
        "fill": HD_LIGHT_ORANGE,
    },
    {
        "num": 2,
        "title": "Sr. Manager\nDropShip Vendor Operations\n& Strategy",
        "location": "Atlanta, GA\n(On-site)\nHD Corp HQ",
        "type": "Full-Time\nOn-site",
        "salary": "$151,000 - $255,000\n(Sr. Manager range)\n+ Bonus",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "3 DAYS AGO!\nBe an Early Applicant\n110 school alumni",
        "posted": "3 days ago",
        "skills": "Dropship / 3P vendor operations\nVendor SC strategy + performance\nOmnichannel SC (store/online)\nEDI, compliance, vendor SLAs\nCross-functional SC strategy\n7+ yrs SC or vendor ops experience",
        "url": "https://www.linkedin.com/jobs/view/4395960349/",
        "network": "110 NCSU alumni\nat Home Depot\nAtlanta HQ",
        "strategy": "FRESHEST STRATEGIC ROLE -- 3 days old, highest salary ($151K-$255K), early applicant window still open. DropShip Vendor Ops = managing 3P seller supply chains + strategy (omnichannel SC). Your SC planning consulting experience maps to vendor performance optimization and SC strategy design. Apply TODAY + message alumni for referral before pool builds.",
        "fill": HD_MEDIUM_ORANGE,
    },
    {
        "num": 3,
        "title": "Sr. Analyst\nInventory Planning",
        "location": "Atlanta, GA\n(On-site)\nHD Corp HQ",
        "type": "Full-Time\nOn-site",
        "salary": "$94,000 - $130,000\n(Sr. Analyst range)\n+ Bonus",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "Recently posted\n110 school alumni\nat Home Depot",
        "posted": "Recently posted",
        "skills": "Inventory planning & optimization\nBlue Yonder / JDA (HD uses BY)\nReplenishment logic + safety stock\nDemand planning inputs + signals\nSQL / Tableau / data analysis\nRetail SC inventory management",
        "url": "https://www.linkedin.com/jobs/view/4375797166/",
        "network": "110 NCSU alumni\nat Home Depot\nSC planning team",
        "strategy": "HOME DEPOT USES BLUE YONDER (JDA) for inventory planning -- your BY expertise is a direct hit! Sr. Analyst, Inventory Planning = owning replenishment parameters, safety stock models, and inventory optimization for a $150B+ retailer. Lead with BY/JDA tool depth + inventory optimization outcomes (% fill rate, turns, stockout reduction). Strong match.",
        "fill": "FEF9F5",
    },
    {
        "num": 4,
        "title": "Manager\nSupply Chain Analytics",
        "location": "Atlanta, GA\n(On-site)\nHD Corp HQ",
        "type": "Full-Time\nOn-site",
        "salary": "$100,000 - $155,000\n(SC Manager range)\n+ Bonus",
        "match": "MEDIUM MATCH",
        "match_bg": YELLOW_BADGE,
        "match_fg": YELLOW_DARK,
        "status": "Recently posted\n110 school alumni\n3 direct reports",
        "posted": "Recently posted",
        "skills": "SC analytics + data science\nTableau + SQL + Cognos + Hadoop\nStatistical analysis techniques\nTeam leadership (3 direct reports)\nSC operational + financial targets\nM.S. in CS/Math/Engineering preferred",
        "url": "https://www.linkedin.com/jobs/view/4390900071/",
        "network": "110 NCSU alumni\nat Home Depot\nAnalytics team",
        "strategy": "SC ANALYTICS MANAGER = leads a team of 3 driving analytical processes across SC. Reports to Director, SC Analytics. Your consulting data skills (SC data models, KPI dashboards, forecasting) are relevant. Gap: requires deep statistical + database skills (Hadoop, Cognos). Best fit if you have strong BI/analytics background alongside SC expertise. Tableau + SQL are must-haves.",
        "fill": HD_LIGHT_GRAY,
    },
    {
        "num": 5,
        "title": "Analyst\nInventory Planning\n& Replenishment",
        "location": "Atlanta, GA\n(On-site)\nHD Corp HQ",
        "type": "Full-Time\nOn-site",
        "salary": "$80,000 - $105,000\n(Analyst range)\n+ Bonus",
        "match": "HIGH MATCH",
        "match_bg": GREEN_BADGE,
        "match_fg": GREEN_DARK,
        "status": "1 week ago\n110 school alumni\nat Home Depot",
        "posted": "1 week ago",
        "skills": "Inventory planning + replenishment\nBlue Yonder / JDA replenishment\nDemand signals + order management\nSafety stock + reorder point logic\nSQL / Excel / data analysis\nRetail SC domain knowledge",
        "url": "https://www.linkedin.com/jobs/view/4371581269/",
        "network": "110 NCSU alumni\nat Home Depot\nInventory team",
        "strategy": "REPLENISHMENT ANALYST = your BY/JDA + IBP tool knowledge is a direct qualification. Home Depot's inventory planning team manages replenishment for 2,300+ stores using Blue Yonder. Note: analyst-level salary ($80K-$105K) may be below target -- negotiate for Sr. Analyst title given 7 yrs experience. Apply alongside Job #3 (Sr. Analyst) for best shot at right level.",
        "fill": "F5FBF5",
    },
]

for i, job in enumerate(JOBS):
    r = i + 5
    ws1.row_dimensions[r].height = 95
    fill = PatternFill("solid", start_color=job["fill"])

    data = [
        job["num"], job["title"], job["location"], job["type"],
        job["salary"], job["match"], job["status"], job["posted"],
        job["skills"], job["url"], job["network"], job["strategy"]
    ]

    for c, val in enumerate(data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill

        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=16, color=HD_ORANGE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=HD_DARK)
            cell.alignment = left()
        elif c == 6:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=job["match_fg"])
            cell.fill = PatternFill("solid", start_color=job["match_bg"])
            cell.alignment = center()
        elif c == 7:
            cell.value = val
            is_hot = "3 DAYS" in str(val)
            cell.font = Font(name="Arial", bold=is_hot, size=9,
                           color=HD_DARK_ORANGE if is_hot else "595959")
            cell.alignment = center()
        elif c == 8:
            cell.value = val
            cell.font = Font(name="Arial", bold="3 days" in str(val), size=9,
                           color=HD_DARK_ORANGE if "3 days" in str(val) else "595959")
            cell.alignment = center()
        elif c == 10:
            hyperlink(cell, val)
            cell.fill = fill
        elif c == 12:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left() if c > 2 else center()

COL_WIDTHS = [4, 26, 18, 13, 22, 14, 22, 13, 34, 12, 18, 48]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 2 -- RESUME TAILORING GUIDE
# ============================================================
ws2 = wb.create_sheet("Resume Tailoring Guide")

ws2.row_dimensions[1].height = 38
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "RESUME TAILORING -- What to Highlight for Each Home Depot Role"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.alignment = center()
t2.fill = PatternFill("solid", start_color=HD_DARK)
t2.border = border

ws2.row_dimensions[2].height = 7
ws2.merge_cells("A2:G2")
ws2["A2"].fill = PatternFill("solid", start_color=HD_ORANGE)

HEADERS2 = ["Role", "Must Highlight in Resume", "Keywords to Include",
            "Experience to Lead With", "Potential Gap", "Gap Mitigation", "Priority"]
ws2.row_dimensions[3].height = 24
for c, h in enumerate(HEADERS2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=HD_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

ws2.freeze_panes = "A4"

RESUME_DATA = [
    (
        "Manager\nSupply Chain PMO",
        "- SC program / project management depth\n- Cross-functional stakeholder leadership\n- Project governance + status reporting\n- Strategic SC initiative execution\n- Executive presentation skills\n- PMP or SC management background",
        "PMO, program management, project governance, supply chain, cross-functional, stakeholder, strategic initiatives, project planning, status reporting, risk management, SC transformation, implementation",
        "Lead with the biggest SC programs you've managed: multi-workstream implementations, SC process transformations, cross-functional rollouts. Show: # workstreams, # stakeholders, $ impact, timeline adherence. PMO at Home Depot manages strategic SC projects -- your consulting experience IS PMO work.",
        "Home Depot PMO is internal-facing (not client-facing consulting). Adjusting from external advisor to internal program owner is a mindset shift.",
        "Frame consulting engagements as internal programs: 'I led a 12-month SC transformation program for [client] across 5 functional workstreams delivering $X in savings.' Remove consulting language, use operational outcomes language. Strong fit overall.",
        "APPLY -- HIGH\nMATCH (PMO =\nconsulting work)"
    ),
    (
        "Sr. Manager\nDropShip Vendor Ops\n& Strategy",
        "- Vendor / 3P SC operations experience\n- SC strategy development + execution\n- Omnichannel retail SC knowledge\n- Vendor performance management\n- EDI, compliance, SLA management\n- Senior leadership + team management",
        "DropShip, 3P vendor, omnichannel, retail SC, vendor operations, vendor strategy, drop ship compliance, EDI, SLA, vendor performance, fulfillment, e-commerce SC, last-mile",
        "Lead with any vendor management or 3P supply chain experience from consulting clients. Home Depot's dropship program serves their online marketplace -- show you understand the link between vendor SC performance and customer experience. Any e-commerce or retail SC project work is highly relevant.",
        "DropShip ops is very specific to retail/e-commerce SC. May require direct experience managing 3P vendor relationships at scale.",
        "Apply immediately (3 days old!). Frame consulting SC design work as 'vendor operations architecture' -- you've designed how SC systems should work with 3P vendors. Retail SC knowledge from any retail/CPG client is directly transferable. Highest-paying role in this list ($151K-$255K).",
        "APPLY NOW!\n(3 DAYS OLD\n+ HIGHEST PAY)"
    ),
    (
        "Sr. Analyst\nInventory Planning",
        "- Inventory planning tool expertise\n- Blue Yonder (JDA) configuration + use\n- Replenishment logic + safety stock models\n- Demand planning signals + data inputs\n- SQL / Tableau for inventory analytics\n- Retail inventory optimization KPIs",
        "Inventory planning, replenishment, Blue Yonder, JDA, safety stock, reorder point, min/max, demand signal, fill rate, inventory turns, days of supply, stockout, ABC classification, SKU planning",
        "Lead with Blue Yonder / JDA implementation and configuration experience. Show inventory optimization outcomes: fill rate %, inventory turns improved, stockout reduction %. Home Depot uses BY for all store replenishment -- your BY expertise is the #1 qualification here.",
        "Sr. Analyst = individual contributor role. May feel like a step back from consulting lead. Salary ($94K-$130K) lower than SC manager roles.",
        "Position as 'domain expert who wants to go deep on inventory science at $150B scale.' BY expertise makes you immediately productive from Day 1. Apply alongside Job #5 (Analyst, IPR) -- mention Sr. Analyst preference in cover note. Strong foot-in-the-door at HD with BY as your ace card.",
        "STRONG MATCH\n(Blue Yonder\nexpertise = ace!)"
    ),
    (
        "Manager\nSupply Chain Analytics",
        "- SC analytics + data science skills\n- Tableau + SQL expertise (required)\n- Team leadership (3 direct reports)\n- Statistical analysis + forecasting\n- SC operational + financial analytics\n- Advanced data tools (Cognos, Hadoop)",
        "Supply chain analytics, Tableau, SQL, Cognos, Hadoop, statistical analysis, Python, data mining, SC performance, KPIs, analytical processes, team management, data visualization, business intelligence",
        "Lead with SC analytics projects: dashboards built, statistical models developed, data-driven SC decisions enabled. Show SQL proficiency and Tableau dashboard portfolio. Managing a team of 3 = show your experience mentoring analysts or junior consultants.",
        "Requires M.S. preferred + deep statistical/database skills (Hadoop, IBM Cognos). Heavy analytics engineering focus -- not pure SC planning.",
        "Apply if you have strong BI/analytics background. Your SC domain knowledge + analytics skills = rare combo that HD values. If Hadoop/Cognos are gaps, highlight Tableau + SQL strength. Frame consulting data analysis work as 'building analytical SC processes' -- exactly what the JD says.",
        "APPLY -- GOOD\nFIT if strong\nSQL/Tableau"
    ),
    (
        "Analyst\nInventory Planning\n& Replenishment",
        "- Inventory planning + replenishment tools\n- Blue Yonder / JDA replenishment modules\n- Safety stock + reorder point logic\n- SKU-level planning methodology\n- SQL / Excel for inventory analytics\n- Retail SC demand patterns",
        "Inventory planning, replenishment, Blue Yonder, JDA, safety stock, reorder point, demand planning, replenishment parameters, order management, fill rate, days on hand, retail SC",
        "Lead with BY/JDA replenishment module experience: show you've configured parameters, tuned safety stock, and improved fill rates. Entry point into Home Depot's core planning team. With 7 years experience, negotiate for Sr. Analyst title during offer stage.",
        "Analyst-level role -- salary ($80K-$105K) likely below 7-yr experience target. Job title may not reflect seniority level sought.",
        "Apply WITH Job #3 (Sr. Analyst) simultaneously. In interview: position yourself at Sr. level and let HD adjust. Your BY expertise + 7 yrs consulting = significantly overqualified for pure analyst. If offered analyst, negotiate immediately for Sr. Analyst title + commensurate salary.",
        "APPLY + NEGOTIATE\nSr. Analyst\ntitle + pay"
    ),
]

ROW_FILLS2 = [HD_LIGHT_ORANGE, HD_MEDIUM_ORANGE, "FEF9F5", HD_LIGHT_GRAY, "F5FBF5"]
PRIORITY_COLORS = [
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (GREEN_BADGE, GREEN_DARK),
    (YELLOW_BADGE, YELLOW_DARK),
    (YELLOW_BADGE, YELLOW_DARK),
]

for i, row_data in enumerate(RESUME_DATA):
    r = i + 4
    ws2.row_dimensions[r].height = 115
    fill = PatternFill("solid", start_color=ROW_FILLS2[i])

    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=10, color=HD_DARK)
            cell.fill = fill
            cell.alignment = left()
        elif c == 7:
            bg, fg = PRIORITY_COLORS[i]
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = center()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.fill = fill
            cell.alignment = left()

COL_WIDTHS2 = [24, 40, 40, 44, 32, 40, 16]
for i, w in enumerate(COL_WIDTHS2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 3 -- HOME DEPOT FACTS + INTERVIEW PREP
# ============================================================
ws3 = wb.create_sheet("Home Depot Facts + Interview")

ws3.row_dimensions[1].height = 38
ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "HOME DEPOT QUICK FACTS + Interview Prep -- Supply Chain & Planning Roles"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.alignment = center()
t3.fill = PatternFill("solid", start_color=HD_DARK)
t3.border = border

ws3.row_dimensions[2].height = 7
ws3.merge_cells("A2:H2")
ws3["A2"].fill = PatternFill("solid", start_color=HD_ORANGE)

ws3.row_dimensions[3].height = 22
ws3.merge_cells("A3:H3")
s3 = ws3["A3"]
s3.value = "HOME DEPOT COMPANY FACTS -- Know Before Your Interview"
s3.font = Font(name="Arial", bold=True, color=WHITE, size=11)
s3.alignment = center()
s3.fill = PatternFill("solid", start_color=HD_ORANGE)
s3.border = border

HD_FACTS = [
    ("Company Overview", "The Home Depot: ~$159B revenue (FY2024). ~460,000 associates. World's largest home improvement retailer. 2,300+ stores in US/Canada/Mexico. NYSE: HD."),
    ("SC Scale", "One of the most complex retail SC networks: 25+ flatbed distribution centers, 150+ stocking distribution centers, RDCs (Rapid Deployment Centers), direct fulfillment centers."),
    ("SC Transformation", "HD completed a $1.2B SC transformation (Project Sync) building a direct-fulfillment network. Acquired SRS Distribution (2024) for $18.25B adding pro contractor SC."),
    ("Planning Tools", "Home Depot uses BLUE YONDER (formerly JDA) for inventory planning and replenishment. If you know BY/JDA, you are immediately relevant for any planning role here."),
    ("Omnichannel SC", "BOPIS (Buy Online, Pick Up In Store), same-day delivery, and DropShip are core to HD's SC strategy. 2,300+ stores = last-mile fulfillment nodes for online orders."),
    ("Pro Customer Focus", "~50% of HD revenue comes from Pro (contractor) customers. Pro customers require more complex SC planning: bulk orders, jobsite delivery, dedicated service levels."),
    ("Corp HQ", "All corporate SC roles are based in Atlanta, GA (HQ: 2455 Paces Ferry Rd NW). Hybrid or on-site typically required for SC leadership roles."),
    ("Compensation", "Base + annual bonus (10-15% target) + RSP (Restricted Stock Program for manager+). Total comp typically 15-25% above base. Good 401k match."),
    ("Interview Culture", "HD values: Customer First, Doing The Right Thing, Excellent Customer Service, Respect for All People. Expect behavioral STAR-format questions aligned to HD values."),
    ("LinkedIn Company ID", "1534  |  35 targeted SC/Planning roles found  |  Use f_C=1534 for future LinkedIn searches at Home Depot"),
]

for i, (label, val) in enumerate(HD_FACTS):
    r = i + 4
    ws3.row_dimensions[r].height = 30
    bg = HD_LIGHT_ORANGE if i % 2 == 0 else HD_LIGHT_GRAY

    lc = ws3.cell(row=r, column=1, value=label)
    lc.font = Font(name="Arial", bold=True, size=9, color=HD_DARK)
    lc.fill = PatternFill("solid", start_color=bg)
    lc.alignment = left()
    lc.border = border

    ws3.merge_cells(f"B{r}:H{r}")
    vc = ws3.cell(row=r, column=2, value=val)
    vc.font = nf(size=9)
    vc.fill = PatternFill("solid", start_color=bg)
    vc.alignment = left()
    vc.border = border

# Interview prep section
int_start = len(HD_FACTS) + 5
ws3.row_dimensions[int_start].height = 22
ws3.merge_cells(f"A{int_start}:H{int_start}")
sec2 = ws3[f"A{int_start}"]
sec2.value = "HOME DEPOT INTERVIEW PREP -- SC & Planning Role Interview Guide"
sec2.font = Font(name="Arial", bold=True, color=WHITE, size=11)
sec2.alignment = center()
sec2.fill = PatternFill("solid", start_color=HD_DARK)
sec2.border = border

INT_HDR = int_start + 1
ws3.row_dimensions[INT_HDR].height = 22
for c, h in enumerate(["#", "Interview Topic", "What HD Tests", "How to Answer (SC Consultant Angle)"], 1):
    cell = ws3.cell(row=INT_HDR, column=c)
    cell.value = h
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", start_color=HD_DARK)
    cell.alignment = center()
    cell.border = thick_bottom

INTERVIEW_QS = [
    (1, "Why Home Depot?",
     "Genuine interest in retail SC at scale + fit with HD's customer-first culture",
     "Talk about HD's SC transformation complexity: 2,300+ stores as fulfillment nodes, Blue Yonder-powered inventory planning, $159B in product flow. Show you understand the Pro customer SC challenge and the SRS acquisition opportunity. Tie to your passion for solving inventory planning at massive retail scale."),
    (2, "SC Planning Methodology",
     "How you approach inventory planning, replenishment design, and S&OP at retail scale",
     "Walk through how you set replenishment parameters (safety stock, reorder points, order quantities) using demand signals and service level targets. For HD: think in terms of 30,000+ SKUs across 2,300 stores with seasonal demand patterns (spring gardening, hurricane prep, holiday). Show BY/JDA methodology depth."),
    (3, "Blue Yonder / JDA Experience",
     "Hands-on familiarity with BY for replenishment, inventory optimization, or demand planning",
     "Be specific about BY modules you've implemented or configured: Demand, Replenishment, Allocation, Space & Category. Show client outcomes: fill rate %, inventory turns improvement, stockout reduction. HD's entire replenishment runs on BY -- this is the most important technical qualification to communicate clearly."),
    (4, "Cross-functional Leadership",
     "Ability to work with merchants, store ops, logistics, and IT teams simultaneously",
     "HD SC planners collaborate daily with Merchants (assortment/pricing), Store Ops (execution), Logistics (DC/transportation), and IT (systems). Share a story of driving alignment across 3-4 functions with competing priorities. Frame consulting multi-client experience as equivalent cross-functional leadership."),
    (5, "Data-Driven Decision Making",
     "SQL, Tableau, Excel skills and ability to turn SC data into actionable insights",
     "Prepare a story where you built a SC analysis that changed a business decision. Show data tools: SQL query, Tableau dashboard, Excel model. For HD: think inventory health dashboards, DC throughput analysis, demand forecast accuracy reporting. HD is highly data-driven -- show your analytical toolkit."),
    (6, "HD Values Behavioral Questions",
     "Customer First, Doing the Right Thing, Excellent Customer Service, Respect, Building Relationships",
     "HD's values framework drives behavioral interview questions. 'Customer First' = show how SC decisions impact store associates and end customers (not just cost/efficiency). 'Doing the Right Thing' = ethical SC sourcing, transparency. 'Building Relationships' = your cross-functional collaboration stories. Prep 2 STAR stories per value."),
]

INT_FILLS = [HD_LIGHT_ORANGE, HD_LIGHT_GRAY, HD_LIGHT_ORANGE, HD_LIGHT_GRAY, HD_LIGHT_ORANGE, HD_LIGHT_GRAY]

for i, (num, topic, tests, how) in enumerate(INTERVIEW_QS):
    r = INT_HDR + 1 + i
    ws3.row_dimensions[r].height = 60
    bg = INT_FILLS[i % len(INT_FILLS)]
    fill = PatternFill("solid", start_color=bg)

    for c, val in enumerate([num, topic, tests, how], 1):
        cell = ws3.cell(row=r, column=c)
        cell.border = border
        cell.fill = fill
        if c == 1:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=13, color=HD_ORANGE)
            cell.alignment = center()
        elif c == 2:
            cell.value = val
            cell.font = Font(name="Arial", bold=True, size=9, color=HD_DARK)
            cell.alignment = left()
        else:
            cell.value = val
            cell.font = nf(size=9)
            cell.alignment = left()

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 62
ws3.freeze_panes = f"A{int_start + 2}"

# ============================================================
# SAVE
# ============================================================
output = "C:/Users/amolp/Prometheus/homedepot_top5_jobs.xlsx"
wb.save(output)
print("Done! homedepot_top5_jobs.xlsx saved with 3 sheets.")

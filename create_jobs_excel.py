from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color

wb = Workbook()

# Colors
DARK_BLUE = "1F4E79"
MID_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

def header_font():
    return Font(name="Arial", bold=True, color=WHITE, size=11)

def title_font():
    return Font(name="Arial", bold=True, color=DARK_BLUE, size=14)

def normal_font():
    return Font(name="Arial", size=10)

def header_fill():
    return PatternFill("solid", start_color=DARK_BLUE)

def alt_fill():
    return PatternFill("solid", start_color=LIGHT_BLUE)

def white_fill():
    return PatternFill("solid", start_color=WHITE)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_row(ws, row, cols, fill, font=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        if font:
            cell.font = font
        cell.border = border

def add_hyperlink(cell, url, display):
    cell.hyperlink = url
    cell.value = display
    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

# ── SHEET 1: o9 OMP Roles ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "o9 OMP Roles"

ws1.row_dimensions[1].height = 30
ws1.merge_cells("A1:G1")
ws1["A1"] = "o9 / OMP Consultant Roles — US Job Listings"
ws1["A1"].font = title_font()
ws1["A1"].alignment = center()
ws1["A1"].fill = PatternFill("solid", start_color="DEEAF1")

ws1.row_dimensions[2].height = 8  # spacer

ws1.row_dimensions[3].height = 22
headers1 = ["#", "Job Title", "Company", "Location", "Type", "Link", "Recruiter / HM", "Notes"]
for c, h in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=c, value=h)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = center()
    cell.border = border

data1 = [
    (1, "Principal O9 Consultant", "Infosys", "Raleigh, NC", "On-site · Full-time",
     "https://www.linkedin.com/jobs/view/4388512492/", "Managed off LinkedIn", "Reach out via Infosys careers portal"),
    (2, "Lead o9 Consultant ⭐", "Infosys", "Richardson, TX", "On-site · Full-time",
     "https://www.linkedin.com/jobs/view/4378154088/", "Managed off LinkedIn", "HIGH MATCH — Reach out via Infosys careers portal"),
    (3, "O9 Solutions Consultant", "InfoVision Inc.", "New York, US", "Remote · Contract",
     "https://www.linkedin.com/jobs/view/4383145463/", "Not listed", "Easy Apply · Actively reviewing applicants"),
]

for i, row_data in enumerate(data1):
    r = i + 4
    ws1.row_dimensions[r].height = 20
    fill = alt_fill() if i % 2 == 0 else white_fill()
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c)
        cell.fill = fill
        cell.border = border
        cell.font = normal_font()
        if c == 1:
            cell.alignment = center()
        else:
            cell.alignment = left()
        if c == 6:
            add_hyperlink(cell, val, "View Job →")
            cell.fill = fill
        else:
            cell.value = val

ws1.freeze_panes = "A4"
col_widths1 = [5, 30, 18, 18, 18, 14, 30, 42]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 2: Supply Chain Architect Roles ─────────────────────────────────
ws2 = wb.create_sheet("Supply Chain Architect Roles")

ws2.row_dimensions[1].height = 30
ws2.merge_cells("A1:I1")
ws2["A1"] = "Supply Chain Architect Roles — US Job Listings"
ws2["A1"].font = title_font()
ws2["A1"].alignment = center()
ws2["A1"].fill = PatternFill("solid", start_color="DEEAF1")

ws2.row_dimensions[2].height = 8  # spacer

ws2.row_dimensions[3].height = 22
headers2 = ["#", "Job Title", "Company", "Location", "Type", "Salary", "Link", "Recruiter / HM", "Notes"]
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = center()
    cell.border = border

data2 = [
    (4, "Solutions Architect - Kinaxis ⭐", "Genpact", "USA", "Remote · Full-time", "$150K – $180K",
     "https://www.linkedin.com/jobs/view/4330227649/",
     "Vipul Gupta\nGlobal Talent Acquisition @ Genpact",
     "Easy Apply · Actively reviewing · HIGH MATCH"),
    (5, "Senior Solution Designer ⭐", "Körber Supply Chain", "USA", "Remote · Full-time", "Not listed",
     "https://www.linkedin.com/jobs/view/4382197450/",
     "Michael Unser\nDirector Solution Design, Körber NA",
     "HIGH MATCH"),
    (6, "Sr Functional Solution Architect ⭐", "Blue Yonder", "Coppell, TX", "Remote · Full-time", "Not listed",
     "https://www.linkedin.com/jobs/view/4394225134/",
     "Not listed publicly",
     "HIGH MATCH · Be an early applicant"),
    (7, "Supply Chain Solutions Architect", "Texas Instruments", "Dallas, TX", "On-site · Full-time", "Not listed",
     "https://www.linkedin.com/jobs/view/4382493831/",
     "Not listed publicly",
     ""),
]

for i, row_data in enumerate(data2):
    r = i + 4
    ws2.row_dimensions[r].height = 36
    fill = alt_fill() if i % 2 == 0 else white_fill()
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c)
        cell.fill = fill
        cell.border = border
        cell.font = normal_font()
        if c == 1:
            cell.alignment = center()
        else:
            cell.alignment = left()
        if c == 7:
            add_hyperlink(cell, val, "View Job →")
            cell.fill = fill
        else:
            cell.value = val

ws2.freeze_panes = "A4"
col_widths2 = [5, 32, 20, 14, 18, 16, 14, 32, 38]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 3: Last 24 Hours ────────────────────────────────────────────────
ws3 = wb.create_sheet("🔥 Last 24 Hours")

ws3.row_dimensions[1].height = 30
ws3.merge_cells("A1:J1")
ws3["A1"] = "🔥 Hot Jobs — Posted in Last 24 Hours (US) — " + "April 10, 2026"
ws3["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws3["A1"].alignment = center()
ws3["A1"].fill = PatternFill("solid", start_color="C00000")

ws3.row_dimensions[2].height = 8

ws3.row_dimensions[3].height = 22
headers3 = ["#", "Job Title", "Company", "Location", "Type", "Salary/Rate", "Posted", "Link", "Recruiter / HM", "Notes"]
for c, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = center()
    cell.border = border

data3 = [
    (1, "Senior SAP IBP Consultant 🚀",    "SELECCIÓN Consulting", "East Brunswick, NJ", "Hybrid · Full-time",  "—",              "10 mins ago · 0 applicants!",
     "https://www.linkedin.com/jobs/view/4400587273/", "Krishna S. (visible on search)", "Easy Apply · APPLY NOW — first mover advantage!"),
    (2, "Solution Architect - WMS ⭐",       "Blue Yonder",          "Coppell, TX",         "Remote · Full-time",  "—",              "4 hrs ago · 99 clicked",
     "https://www.linkedin.com/jobs/view/4321921503/", "Network: Kevin + others",        "HIGH MATCH · Responses managed off LinkedIn"),
    (3, "Kinaxis Sr Solution Consultant 🟢", "Pacer Group",          "USA",                 "Remote · Contract",   "$82–$92/hr",     "4 hrs ago · 45 applicants",
     "https://www.linkedin.com/jobs/view/4393864119/", "Not listed",                     "Easy Apply · ~$170K–$190K annualized"),
    (4, "Platform Architect - Semantics",    "Kinaxis (Direct)",     "USA",                 "Remote · Full-time",  "—",              "46 mins ago",
     "https://www.linkedin.com/jobs/view/4398156298/", "Not listed",                     "Direct from Kinaxis"),
    (5, "Blue Yonder Project Manager 🟢",    "Maven Workforce Inc.", "California",          "Remote",              "—",              "1 hr ago",
     "https://www.linkedin.com/jobs/view/4400574301/", "Not listed",                     "Easy Apply"),
    (6, "SAP IBP Consultant",                "Jobs via Dice",        "USA",                 "Remote",              "—",              "4 hrs ago",
     "https://www.linkedin.com/jobs/view/4398118825/", "Not listed",                     ""),
]

for i, row_data in enumerate(data3):
    r = i + 4
    ws3.row_dimensions[r].height = 36
    fill = alt_fill() if i % 2 == 0 else white_fill()
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c)
        cell.fill = fill
        cell.border = border
        cell.font = normal_font()
        if c == 1:
            cell.alignment = center()
        else:
            cell.alignment = left()
        if c == 8:
            add_hyperlink(cell, val, "View Job →")
            cell.fill = fill
        else:
            cell.value = val

ws3.freeze_panes = "A4"
col_widths3 = [5, 34, 22, 18, 18, 14, 22, 14, 28, 38]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 4: Demand Planner Roles ─────────────────────────────────────────
ws4 = wb.create_sheet("Demand Planner Roles")

ws4.row_dimensions[1].height = 30
ws4.merge_cells("A1:J1")
ws4["A1"] = "Demand Planner / Senior Demand Planner Roles — US Job Listings"
ws4["A1"].font = Font(name="Arial", bold=True, color=DARK_BLUE, size=14)
ws4["A1"].alignment = center()
ws4["A1"].fill = PatternFill("solid", start_color="E2EFDA")

ws4.row_dimensions[2].height = 8

ws4.row_dimensions[3].height = 22
headers4 = ["#", "Job Title", "Company", "Location", "Type", "Salary", "Match", "Link", "Recruiter / HM", "Notes"]
for c, h in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=c, value=h)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = center()
    cell.border = border

data4 = [
    (1, "Sr. Supply Demand Planner ⭐",           "Intuitive",           "Sunnyvale, CA",  "On-site · Full-time", "$118.7K–$170.7K", "HIGH",
     "https://www.linkedin.com/jobs/view/4345274465/", "Not listed",           "2 o9 alumni work here · Responses managed off LinkedIn"),
    (2, "Sr. Supply & Demand Planner ⭐",          "Kohler Co.",          "Kohler, WI",     "On-site · Full-time", "Not listed",       "HIGH",
     "https://www.linkedin.com/jobs/view/4366386259/", "Taylor Magri, CRD\nRecruiter @ Kohler", "o9 alumni in network · Message recruiter directly!"),
    (3, "Senior Supply Chain Planner ⭐",          "Micron Technology",   "Boise, ID",      "On-site · Full-time", "Not listed",       "HIGH",
     "https://www.linkedin.com/jobs/view/4386872480/", "Not listed",           "50 school alumni work here"),
    (4, "Senior Analyst, Supply Chain Planning ⭐", "Analog Devices",     "Wilmington, MA", "Hybrid · Full-time",  "Not listed",       "HIGH",
     "https://www.linkedin.com/jobs/view/4371933606/", "Not listed",           "140 school alumni work here"),
    (5, "Supply Chain Analysts – Sr #GA001 ⭐",    "Cummins Inc.",        "Atlanta, GA",    "On-site · Full-time", "Not listed",       "HIGH",
     "https://www.linkedin.com/jobs/view/4393461958/", "Not listed",           "73 school alumni work here"),
    (6, "Sr Demand Planner (Ecomm & Wholesale) 🟢","Velvet Caviar",       "New York, US",   "Remote",              "$90K–$120K",       "—",
     "https://www.linkedin.com/jobs/view/4391579652/", "Not listed",           "Easy Apply · Actively reviewing applicants"),
    (7, "Senior S&OP Planner",                     "Amazon Web Services", "Seattle, WA",    "On-site · Full-time", "Not listed",       "—",
     "https://www.linkedin.com/jobs/view/4381119799/", "Not listed",           "11 o9 alumni work here"),
]

GREEN_LIGHT = "E2EFDA"
for i, row_data in enumerate(data4):
    r = i + 4
    ws4.row_dimensions[r].height = 36
    fill = PatternFill("solid", start_color=GREEN_LIGHT) if i % 2 == 0 else white_fill()
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c)
        cell.fill = fill
        cell.border = border
        cell.font = normal_font()
        cell.alignment = center() if c == 1 else left()
        if c == 8:
            add_hyperlink(cell, val, "View Job →")
            cell.fill = fill
        else:
            cell.value = val
            # Bold HIGH match column
            if c == 7 and val == "HIGH":
                cell.font = Font(name="Arial", size=10, bold=True, color="375623")

ws4.freeze_panes = "A4"
col_widths4 = [5, 34, 20, 16, 18, 16, 8, 14, 30, 40]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

wb.save("C:/Users/amolp/Prometheus/job_listings_v2.xlsx")
print("Done: job_listings_v2.xlsx updated with Demand Planner sheet")
